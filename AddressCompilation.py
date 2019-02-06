# This file contains google mapping and address compilation functionality

import googlemaps
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant, QSortFilterProxyModel
from PyQt5.QtWidgets import QTableView
from openpyxl import load_workbook, Workbook
from datetime import datetime
import math
gmaps = googlemaps.Client(key='AIzaSyACdmhoHP3DVefOLizCPVuHtUULokCWCpI')


class RegionList(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data
        self.header_data = ["Existing Site Files"]

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return 1

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role != Qt.DisplayRole:
            return None

        # value = str(self._data[index.row()][index.column()])
        value = self._data[index.row()]

        return value

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]
        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable


class RegionView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, regions):
        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(True)
        self.sorter.setSourceModel(RegionList(regions))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)


class ExistingAddressBook:

    def __init__(self, site_name):

        self.site_name = site_name
        self.file_name = "Sites/" + site_name + ".xlsx"
        self.wb = load_workbook(self.file_name)

        self.existing_trips = []
        self.known_addresses = []
        self.populate_existing_trips()

    def populate_existing_trips(self):

        ws = self.wb["Trips"]
        max_row = ws.max_row
        x = 2

        while x <= max_row:
            stop_one = ws["A" + str(x)].value
            stop_two = ws["B" + str(x)].value
            address_one = ws["C" + str(x)].value
            address_two = ws["D" + str(x)].value
            distance = ws["E" + str(x)].value
            duration = ws["F" + str(x)].value
            tod = ws["G" + str(x)].value
            traffic = ws["H" + str(x)].value

            self.existing_trips.append([stop_one, stop_two, distance, math.ceil(duration), tod, traffic])
            self.known_addresses.append([stop_one, address_one])
            self.known_addresses.append([stop_two, address_two])
            x += 1

    def check_existing_trips(self, stop_one_name, stop_two_name):

        check_list = [x for x in self.existing_trips if x[0] == stop_one_name and x[1] == stop_two_name]
        if len(check_list) == 0:
            return False
        if len(check_list) == 1:
            return True
        if len(check_list) > 1:
            print("multiple trips found??")
            return True

    def add_a_trip(self, stop_one, stop_two, address_one, address_two):
        # print("adding a trip")
        ws = self.wb["Trips"]
        next_row = ws.max_row + 1

        ws["A" + str(next_row)].value = stop_one
        ws["B" + str(next_row)].value = stop_two

        try_google = True
        if address_one in (None, "", " "):
            address_one = self.find_an_address(stop_one)

        if address_two in (None, "", " "):
            address_two = self.find_an_address(stop_two)

        if not address_one or not address_two:
            try_google = False

        ws["C" + str(next_row)].value = address_one
        ws["D" + str(next_row)].value = address_two

        if try_google:
            try:
                dist, dur = self.google_map_it(address_one, address_two)
            except:
                print("google maps didn't work, trying to remove bad zip")
                try:
                    # try removing last four digits
                    address_one = address_one[:-5]
                    address_two = address_two[:-5]
                    dist, dur = self.google_map_it(address_one, address_two)
                except:
                    print("still not good!!! adding ten minutes and 1 mile to return to P&DC")
                    print(address_one)
                    print(address_two)
                    dist = 1609
                    dur = 600
        else:
            print("did not try google maps, one address was blank")
            dist = 1609
            dur = 600

        dur = max(1, int(math.ceil(dur/60)))

        ws["E" + str(next_row)].value = dist
        ws["F" + str(next_row)].value = dur
        ws["G" + str(next_row)].value = datetime.now()

        self.existing_trips.append([stop_one, stop_two, dist, dur, datetime.now(), None])
        self.wb.save(self.file_name)
        print("trip added")
        return dur, dist

    def add_a_schedule(self, schedule):

        for x, stop in enumerate(schedule.stops[:-1]):
            next_stop = schedule.stops[x+1]
            stop_one = stop.stop_name
            stop_two = next_stop.stop_name
            address_one = stop.address
            address_two = next_stop.address
            if not self.check_existing_trips(stop_one, stop_two):
                self.add_a_trip(stop_one, stop_two, address_one, address_two)

    def add_a_plate(self, schedules):

        for schedule in schedules:
            print(schedule.schedule_name)
            self.add_a_schedule(schedule)

    def get_a_trip(self, stop_one, stop_two):

        check_list = [x for x in self.existing_trips if x[0] == stop_one and x[1] == stop_two]
        if len(check_list) == 0:
            print("Trip not found")
            return None
        if len(check_list) == 1:
            return check_list[0]
        if len(check_list) > 1:
            print("multiple trips found??")
            return check_list[0]

    def find_an_address(self, stop_name):

        if stop_name in [x[0] for x in self.known_addresses]:
            return next(x[1] for x in self.known_addresses if x[0] == stop_name)

        return None

    @staticmethod
    def google_map_it(address1, address2):

        dist = gmaps.distance_matrix(address1, address2)
        temp = dist["rows"][0]["elements"][0]
        assert temp["status"] == "OK"
        dist_output = temp["distance"]["value"]
        dur_output = temp["duration"]["value"]

        return dist_output, dur_output


class NewAddressBook:

    def __init__(self, site_name):

        self.wb = Workbook()
        ws = self.wb.create_sheet("Trips")

        ws["A1"].value = "Stop 1"
        ws["B1"].value = "Stop 2"
        ws["C1"].value = "Stop 1 Address"
        ws["D1"].value = "Stop 2 Address"
        ws["E1"].value = "Distance"
        ws["F1"].value = "Duration"
        ws["G1"].value = "Time of Day"
        ws["H1"].value = "Traffic"

        self.wb.save("Sites/" + site_name + ".xlsx")
        self.wb.close()