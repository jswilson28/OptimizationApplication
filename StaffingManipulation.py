
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant, QSortFilterProxyModel
from PyQt5.QtWidgets import (QWidget, QHBoxLayout, QVBoxLayout, QLabel, QPushButton, QFileDialog,
                             QRadioButton, QLineEdit, QTableView)
from GeneralMethods import (cminute_to_tod, night_differential, sun_minutes, number, quick_color,
                            percent, today, now, number_decimal, day_from_num)
from openpyxl import load_workbook
import pandas as pd
import copy


class StaffingManipulationGUI(QWidget):

    def __init__(self, ip):
        super().__init__()

        # input passer
        self.ip = ip

        # tables
        self.employee_table = None
        self.totals_table = None
        self.others_table = None

        # top bar
        self.load_button = None
        self.load_previous_button = None
        self.restore_original_button = None
        self.show_schedules_radio = None
        self.show_details_radio = None
        self.show_details_bool = None
        self.clear_all_button = None

        # side bar
        self.original_sort_button = None
        self.new_sort_button = None
        self.add_employee_button = None
        self.remove_employees_button = None
        self.undo_change_button = None
        self.flatten_all_button = None
        self.all_flattened_bool = None
        self.round_all_button = None
        self.all_rounded_bool = None

        # bottom bar
        self.save_name_bar = None
        self.save_to_excel_button = None
        self.modify_existing_schedules_button = None

        # processor
        self.processor = None

        self.set_layout()
        self.add_button_listeners()

    def set_layout(self):

        outer_layout = QVBoxLayout()
        top_bar = QHBoxLayout()
        main_horizontal = QHBoxLayout()
        side_bar = QHBoxLayout()
        right_tables = QVBoxLayout()
        bottom_bar = QHBoxLayout()

        # top bar
        self.load_button = QPushButton("Load Staffing Files", self)
        self.load_previous_button = QPushButton("Load Manipulated File", self)
        self.restore_original_button = QPushButton("Restore Original Solution", self)
        self.show_schedules_radio = QRadioButton("Show Schedules")
        self.show_details_radio = QRadioButton("Show Details")
        self.clear_all_button = QPushButton("Clear All", self)

        self.show_schedules_radio.case = False
        self.show_details_radio.case = True
        self.show_schedules_radio.setChecked(True)
        self.show_details_bool = False

        top_bar.addWidget(self.load_button)
        top_bar.addWidget(self.load_previous_button)
        top_bar.addWidget(self.restore_original_button)
        top_bar.addWidget(self.show_schedules_radio)
        top_bar.addWidget(self.show_details_radio)
        top_bar.addWidget(self.clear_all_button)
        top_bar.addStretch(1)

        # side bar
        self.original_sort_button = QPushButton("Original Sort Order", self)
        self.new_sort_button = QPushButton("Re-sort", self)
        self.add_employee_button = QPushButton("Add Employee", self)
        self.remove_employees_button = QPushButton("Remove Blank Employees", self)
        self.flatten_all_button = QPushButton("Flatten All Start Times", self)
        self.round_all_button = QPushButton("Round All Start Times (15 min)", self)
        self.undo_change_button = QPushButton("Undo Last Change", self)

        side_bar.addWidget(self.original_sort_button)
        side_bar.addWidget(self.new_sort_button)
        side_bar.addWidget(self.add_employee_button)
        side_bar.addWidget(self.remove_employees_button)
        side_bar.addWidget(self.flatten_all_button)
        side_bar.addWidget(self.round_all_button)
        side_bar.addWidget(self.undo_change_button)
        side_bar.addStretch(1)

        # bottom bar
        save_label = QLabel("Enter Name for Output File:")

        self.save_name_bar = QLineEdit()
        self.save_name_bar.setText("****")
        self.save_name_bar.setFixedWidth(400)

        self.save_to_excel_button = QPushButton("Save Output to Excel", self)
        self.modify_existing_schedules_button = QPushButton("Modify Existing Schedules", self)

        bottom_bar.addWidget(save_label)
        bottom_bar.addWidget(self.save_name_bar)
        bottom_bar.addWidget(self.save_to_excel_button)
        bottom_bar.addWidget(self.modify_existing_schedules_button)
        bottom_bar.addStretch(1)

        # main table
        self.employee_table = EmployeeView()
        self.employee_table.doubleClicked.connect(self.employee_table_schedule_mover)

        # right tables
        self.totals_table = TotalsView()
        self.others_table = OthersView()

        right_tables.addWidget(self.totals_table, 3)
        right_tables.addWidget(self.others_table, 2)

        # main horizontal
        # main_horizontal.addLayout(side_bar)
        main_horizontal.addWidget(self.employee_table, 3)
        main_horizontal.addLayout(right_tables, 1)

        # outer layout
        outer_layout.addLayout(top_bar)
        outer_layout.addLayout(side_bar)
        outer_layout.addLayout(main_horizontal)
        outer_layout.addLayout(bottom_bar)

        self.setLayout(outer_layout)

    def add_button_listeners(self):

        # top bar
        self.load_button.clicked[bool].connect(self.load_files)
        self.load_previous_button.clicked[bool].connect(self.load_previous_file)
        self.restore_original_button.clicked[bool].connect(self.restore_original)
        self.show_schedules_radio.toggled.connect(self.show_details)
        self.show_details_radio.toggled.connect(self.show_details)
        self.clear_all_button.clicked[bool].connect(self.clear_all)

        # side bar
        self.original_sort_button.clicked[bool].connect(self.original_sort)
        self.new_sort_button.clicked[bool].connect(self.new_sort)
        self.add_employee_button.clicked[bool].connect(self.add_employee)
        self.remove_employees_button.clicked[bool].connect(self.remove_employees)
        self.flatten_all_button.clicked[bool].connect(self.flatten_all)
        self.round_all_button.clicked[bool].connect(self.round_all)
        self.undo_change_button.clicked[bool].connect(self.undo_change)

        # bottom bar
        self.save_to_excel_button.clicked[bool].connect(self.save_to_excel)
        self.modify_existing_schedules_button.clicked[bool].connect(self.modify_existing_schedules)

    def load_files(self):

        hua_file, _ = QFileDialog.getOpenFileName(self, "Select Hua's File", "",
                                                  "Excel Files (*.xls*);;Excel Files (*.xls*)")

        file, _ = QFileDialog.getOpenFileName(self, "Select Staffing Input Data File", "",
                                              "Excel Files (*.xlsx);;Excel Files (*.xlsx)")

        if not (hua_file and file):
            return

        self.processor = StaffingProcessor(hua_file, file, self.ip)
        self.save_name_bar.setText(hua_file.split("/")[-1])
        self.set_tables()

    def load_previous_file(self):

        file, _ = QFileDialog.getOpenFileName(self, "Select File Containing Existing Solution", "",
                                              "Excel Files (*.xlsx);;Excel Files (*.xlsx)")

        if not file:
            return

        self.processor = StaffingProcessor(None, file, self.ip)
        self.save_name_bar.setText(file.split("/")[-1])
        self.set_tables()

    def restore_original(self):

        if not self.processor:
            return

        self.processor.restore_originals()
        self.set_tables()

    def show_details(self):

        if not self.processor:
            return

        radiobutton = self.sender()

        if radiobutton.isChecked():
            self.show_details_bool = radiobutton.case

        self.set_tables()

    def undo_change(self):

        if not self.processor:
            return

        self.processor.undo_last_change()
        self.set_tables()

    def original_sort(self):

        if not self.processor:
            return

        self.employee_table.set_to_original_sort_order()
        # self.employee_table.update()

    def new_sort(self):

        if not self.processor:
            return

        self.processor.set_new_sort_order()
        self.set_tables()
        self.employee_table.set_to_new_sort_order()
        self.employee_table.update()

    def add_employee(self):

        if not self.processor:
            return

        self.processor.create_a_new_employee()
        self.set_tables()

    def remove_employees(self):

        if not self.processor:
            return

        self.processor.remove_blank_employees()
        self.set_tables()

    def flatten_all(self):

        if not self.processor:
            return

        if self.show_details_bool:
            return

        if self.all_flattened_bool:
            self.processor.unflatten_all()
            self.all_flattened_bool = False
        else:
            self.processor.flatten_all()
            self.all_flattened_bool = True

        self.set_tables()

    def round_all(self):

        if not self.processor:
            return

        if self.show_details_bool:
            return

        if self.all_rounded_bool:
            self.processor.unround_all()
            self.all_rounded_bool = False
        else:
            self.processor.round_all()
            self.all_rounded_bool = True

        self.set_tables()

    def save_to_excel(self):

        if not self.processor:
            return

        self.processor.print_to_excel(self.save_name_bar.text())

    def employee_table_listener(self, item):

        # this method saves employee type changes, time flattens, and 15 minute rounds
        if not self.processor:
            return

        index = self.employee_table.model().mapToSource(item)
        col = index.column()
        row = index.row()

        if col not in (1, 12, 13):
            return

        new_val = self.employee_table.model().sourceModel()._data[row][col]

        if col == 1:
            self.processor.change_an_employee_type(row, new_val)

        elif col == 12:
            if not new_val:
                self.processor.flatten_start_times(row)
            else:
                self.processor.unflatten_start_times(row)

        elif col == 13:
            if not new_val:
                self.processor.round_start_times(row)
            else:
                self.processor.unround_start_times(row)

        self.set_tables()

    def employee_table_schedule_mover(self, item):

        if not self.processor:
            return

        if self.show_details_bool:
            return

        col = item.column()
        if col not in (2, 3, 4, 5, 6, 7, 8):
            return

        if not self.employee_table.sorter.sourceModel().selection_mode:

            index = self.employee_table.model().mapToSource(item)

            row = index.row()

            schedule = self.processor.employees[row].schedules[col-2]
            eligible_rows, swappable_rows = self.processor.find_eligible_slots(schedule)

            self.employee_table.go_to_selection_mode(eligible_rows, swappable_rows, col-2, row)

        else:

            eligible_rows = self.employee_table.sorter.sourceModel().eligible_indices
            swappable_rows = self.employee_table.sorter.sourceModel().swappable_indices

            if not eligible_rows and not swappable_rows:
                self.employee_table.leave_selection_mode()
                return

            dow = self.employee_table.sorter.sourceModel().dow
            selected_index = self.employee_table.sorter.sourceModel().selected_index

            index = self.employee_table.model().mapToSource(item)

            if index.column() != (dow + 2):
                self.employee_table.leave_selection_mode()
                return

            if index.row() in eligible_rows:
                self.processor.move_a_schedule(selected_index, index.row(), dow)
                self.set_tables()
                return

            if index.row() in swappable_rows:
                self.processor.swap_two_schedules(selected_index, index.row(), dow)
                self.set_tables()
                return

            self.employee_table.leave_selection_mode()

    def set_tables(self):

        if not self.processor:
            return

        if self.show_details_bool:
            self.employee_table.set_model(self.processor.details_df, True)
        else:
            self.employee_table.set_model(self.processor.employee_df, False)

        self.employee_table.model().dataChanged.connect(self.employee_table_listener)
        self.totals_table.set_model(self.processor.totals_df)
        self.others_table.set_model(self.processor.others_df)
        self.employee_table.update()
        self.totals_table.update()
        self.others_table.update()

    def refresh_tables(self):

        self.employee_table.update()
        self.totals_table.update()
        self.others_table.update()

    def clear_all(self):

        print("Yeah, this doesn't actually work.")

    def modify_existing_schedules(self):

        if not self.processor:
            return

        existing_schedule_file, _ = QFileDialog.getOpenFileName(self, "Select File Containing Existing Solution", "",
                                                                "Excel Files (*.xlsx);;Excel Files (*.xlsx)")

        if not existing_schedule_file:
            return

        self.processor.modify_existing_schedules(existing_schedule_file)


class Employee:

    def __init__(self, **kwargs):

        self.employee_num = kwargs.get('employee_num')
        self.employee_type = kwargs.get('employee_type')
        self.schedules = kwargs.get('schedules')
        # self.original_schedules = copy.deepcopy(self.schedules)
        self.preset_schedules = copy.deepcopy(self.schedules)
        self.flattened_schedules = copy.deepcopy(self.schedules)
        self.rounded_schedules = copy.deepcopy(self.schedules)
        self.flattened_rounded = copy.deepcopy(self.schedules)

        self.round_minute = 15
        self.standby_minutes = 0
        self.round_minutes = 0
        self.fluff_minutes = 0

        self.nd_pm = kwargs.get('nd_pm')
        self.nd_am = kwargs.get('nd_am')

        self.max_days = None
        self.set_max_days()

        self.original_weekly_hours = None
        self.weekly_work_days = None

        self.weekly_hours = None
        self.weekly_worked_hours = None
        self.weekly_lunch_hours = None
        self.weekly_night_hours = None
        self.weekly_sunday_hours = None
        self.weekly_regular_hours = None
        self.weekly_overtime_hours = None
        self.weekly_penalty_hours = None

        self.annual_hours = None
        self.annual_worked_hours = None
        self.annual_lunch_hours = None
        self.annual_night_hours = None
        self.annual_sunday_hours = None
        self.annual_regular_hours = None
        self.annual_overtime_hours = None
        self.annual_penalty_hours = None

        self.annual_totals_list = []
        self.details_display_list = []
        self.schedules_display_list = []

        self.original_sort_order = None
        self.final_sort_order = None

        self.flattened_bool = False
        self.rounded_bool = False

        self.set_schedule_options()
        self.calculate_hours()

    def set_max_days(self):

        self.max_days = 5
        if self.employee_type in ("PTF", "PSE"):
            self.max_days = 6

    def calculate_hours(self):

        weekly_total_minutes = sum(x[3] for x in self.schedules if x)
        weekly_lunch_minutes = sum(x[6] for x in self.schedules if x)
        weekly_worked_minutes = weekly_total_minutes - weekly_lunch_minutes
        weekly_over_duration_minutes = max(0, weekly_worked_minutes - 2400)
        weekly_penalty_minutes = max(0, weekly_worked_minutes - 3360)

        work_days = 0
        over_duration_minutes = 0
        penalty_minutes = 0
        weekly_night_diff_minutes = 0
        weekly_sunday_minutes = 0

        for schedule in self.schedules:
            if schedule:
                night_diff_minutes = night_differential(self.nd_am, self.nd_pm, schedule[1], schedule[3])
                night_diff_lunch = night_differential(self.nd_am, self.nd_pm, schedule[4], schedule[6])
                night_diff_minutes -= night_diff_lunch

                sunday_minutes = sun_minutes(schedule[1], schedule[3])
                sunday_lunch = sun_minutes(schedule[4], schedule[6])
                sunday_minutes -= sunday_lunch

                weekly_night_diff_minutes += night_diff_minutes
                weekly_sunday_minutes += sunday_minutes
                work_days += 1
                over_duration_minutes += max(0, (schedule[3] - schedule[6]) - 480)
                penalty_minutes += max(0, (schedule[3] - schedule[6]) - 600)

        over_day_minutes = 0
        over_days = max(0, work_days - self.max_days)
        daily_minutes = [x[3] for x in self.schedules if x]
        daily_minutes.sort(key=int)
        for day in range(0, over_days):
            over_day_minutes += daily_minutes.pop(0)

        weekly_overtime_minutes = max(over_duration_minutes + over_day_minutes, weekly_over_duration_minutes, 0)
        weekly_penalty_minutes = max(weekly_penalty_minutes, penalty_minutes)
        weekly_regular_minutes = weekly_worked_minutes - weekly_overtime_minutes
        weekly_overtime_minutes -= weekly_penalty_minutes

        self.weekly_work_days = work_days
        if not self.original_weekly_hours:
            self.original_weekly_hours = weekly_total_minutes/60

        self.weekly_hours = weekly_total_minutes/60
        self.weekly_worked_hours = weekly_worked_minutes/60
        self.weekly_lunch_hours = weekly_lunch_minutes/60
        self.weekly_night_hours = weekly_night_diff_minutes/60
        self.weekly_sunday_hours = weekly_sunday_minutes/60
        self.weekly_regular_hours = weekly_regular_minutes/60
        self.weekly_overtime_hours = weekly_overtime_minutes/60
        self.weekly_penalty_hours = weekly_penalty_minutes/60

        self.annual_hours = self.weekly_hours * 52
        self.annual_worked_hours = self.weekly_worked_hours * 52
        self.annual_lunch_hours = self.weekly_lunch_hours * 52
        self.annual_night_hours = self.weekly_night_hours * 52
        self.annual_sunday_hours = self.weekly_sunday_hours * 52
        self.annual_regular_hours = self.weekly_regular_hours * 52
        self.annual_overtime_hours = self.weekly_overtime_hours * 52
        self.annual_penalty_hours = self.weekly_penalty_hours * 52

        self.set_lists()

    def set_lists(self):

        self.annual_totals_list = [self.annual_hours, self.annual_lunch_hours, self.annual_worked_hours,
                                   self.annual_regular_hours, self.annual_night_hours, self.annual_sunday_hours,
                                   self.annual_overtime_hours, self.annual_penalty_hours]

        self.details_display_list = [self.employee_num, self.employee_type] + self.annual_totals_list + \
                                    [self.check_good_week()]

        self.schedules_display_list = []
        self.schedules_display_list.append(self.employee_num)
        self.schedules_display_list.append(self.employee_type)

        for schedule in self.schedules:
            if schedule:
                start_time = cminute_to_tod(schedule[1])
                end_time = cminute_to_tod(schedule[2])
                time_string = "   " + str(start_time)[:-3] + "-" + str(end_time)[:-3] + "   "
                self.schedules_display_list.append(time_string)
            else:
                self.schedules_display_list.append(None)

        self.schedules_display_list.append(self.check_good_week())
        self.schedules_display_list.append(self.weekly_hours)
        self.schedules_display_list.append(self.weekly_work_days)
        self.schedules_display_list.append(self.flattened_bool)
        self.schedules_display_list.append(self.rounded_bool)
        self.schedules_display_list.append(self.original_sort_order)
        self.schedules_display_list.append(self.final_sort_order)

    def check_good_week(self):

        check_time = 720

        real_schedules = [x[1:3] for x in self.schedules if x]
        if not real_schedules:
            return True

        new_schedule = copy.deepcopy(real_schedules[0])
        new_schedule[0] += 10080
        new_schedule[1] += 10080
        real_schedules.append(new_schedule)

        for x, schedule in enumerate(real_schedules[:-1]):
            end_time = schedule[1]
            next_start_time = real_schedules[x+1][0]
            if (next_start_time - end_time) < check_time:
                for schedule in real_schedules:
                    print(schedule)
                return False

        if self.weekly_work_days > self.max_days:
            # print("here 2")
            return False

        return True

    def add_a_schedule(self, schedule):

        self.schedules = copy.deepcopy(self.preset_schedules)

        dow = int(schedule[0].split("_")[0]) - 1

        if self.schedules[dow]:
            return False
        else:
            self.schedules[dow] = schedule

        self.preset_schedules = copy.deepcopy(self.schedules)
        self.set_schedule_options()
        self.set_displayed_schedules()

        return True

    def remove_a_schedule(self, index):

        self.schedules = copy.deepcopy(self.preset_schedules)

        return_schedule = self.schedules[index]
        self.schedules[index] = None

        self.preset_schedules = copy.deepcopy(self.schedules)

        self.set_schedule_options()
        self.calculate_hours()

        return return_schedule

    def eligible_schedule_check(self, schedule):

        dow = int(schedule[0].split("_")[0]) - 1

        if self.schedules[dow]:
            return False

        # add check here for 12 hour buffer, multiple days?
        return True

    def change_type(self, new_type):

        self.employee_type = new_type
        self.set_max_days()
        self.calculate_hours()

    def set_schedule_options(self):

        self.flattened_schedules = copy.deepcopy(self.preset_schedules)

        start_times = [x[1] for x in self.schedules if x]
        start_times_adjusted = [(x + 240) % 1440 for x in start_times]

        if not start_times_adjusted:
            self.rounded_schedules = copy.deepcopy(self.preset_schedules)
            self.flattened_rounded = copy.deepcopy(self.flattened_schedules)
            return

        start_minute = min(start_times_adjusted)
        reductions = [x - start_minute for x in start_times_adjusted]
        x = 0
        for y, schedule in enumerate(self.flattened_schedules):
            if schedule:
                self.flattened_schedules[y][1] -= reductions[x]
                self.flattened_schedules[y][3] += reductions[x]
                self.standby_minutes += reductions[x]
                x += 1

        self.rounded_schedules = copy.deepcopy(self.preset_schedules)

        for y, schedule in enumerate(self.rounded_schedules):
            if schedule:
                current_minute = schedule[1] % self.round_minute
                if current_minute <= 10:
                    self.fluff_minutes += current_minute
                    self.rounded_schedules[y][1] -= current_minute
                    self.rounded_schedules[y][3] += current_minute
                else:
                    current_minute = self.round_minute - current_minute
                    self.fluff_minutes -= current_minute
                    self.rounded_schedules[y][1] += current_minute
                    self.rounded_schedules[y][3] -= current_minute

        self.flattened_rounded = copy.deepcopy(self.flattened_schedules)

        for y, schedule in enumerate(self.flattened_rounded):
            if schedule:
                current_minute = schedule[1] % self.round_minute
                if current_minute <= 10:
                    self.fluff_minutes += current_minute
                    self.flattened_rounded[y][1] -= current_minute
                    self.flattened_rounded[y][3] += current_minute
                else:
                    current_minute = self.round_minute - current_minute
                    self.fluff_minutes -= current_minute
                    self.flattened_rounded[y][1] += current_minute
                    self.flattened_rounded[y][3] -= current_minute

    def set_displayed_schedules(self):

        if self.flattened_bool and self.rounded_bool:
            self.schedules = copy.deepcopy(self.flattened_rounded)
        elif self.flattened_bool:
            self.schedules = copy.deepcopy(self.flattened_schedules)
        elif self.rounded_bool:
            self.schedules = copy.deepcopy(self.rounded_schedules)
        else:
            self.schedules = copy.deepcopy(self.preset_schedules)

        self.calculate_hours()

    def flatten_schedules(self):

        self.flattened_bool = True
        self.set_displayed_schedules()

    def unflatten_schedules(self):

        self.flattened_bool = False
        self.set_displayed_schedules()

    def round_schedules(self):

        self.rounded_bool = True
        self.set_displayed_schedules()

    def unround_schedules(self):

        self.rounded_bool = False
        self.set_displayed_schedules()

    def excel_print_schedules(self, ws, row):

        ws["B" + str(row)].value = self.employee_num
        ws["Q" + str(row)].value = self.weekly_hours
        ws["R" + str(row)].value = self.weekly_worked_hours
        ws["S" + str(row)].value = self.employee_type

        for x, schedule in enumerate(self.schedules):
            if schedule:
                y = 3 + (x * 2)
                ws.cell(row=row, column=y).value = cminute_to_tod(schedule[1])
                ws.cell(row=row, column=(y + 1)).value = cminute_to_tod(schedule[2])

    def excel_print_staffing(self, ws, row):

        ws["B" + str(row)].value = self.employee_num
        ws["C" + str(row)].value = self.employee_type

        for y in range(0, 7):
            if self.schedules[y]:
                ws.cell(row=row, column=(4 + y)).value = self.schedules[y][0]

    def excel_print_by_employee(self, ws, row):

        ws["A" + str(row)].value = self.employee_num
        ws["B" + str(row)].value = self.employee_type

        for x, item in enumerate(self.annual_totals_list):
            ws.cell(row=row, column=(3 + x)).value = item

    def set_original_sort_order(self, index):

        self.original_sort_order = index
        if not self.final_sort_order:
            self.final_sort_order = index
        # self.set_schedule_options()
        # self.calculate_hours()
        self.set_lists()

    def set_final_sort_order(self, index):

        self.final_sort_order = index
        # self.set_schedule_options()
        # self.calculate_hours()
        self.set_lists()


class StaffingProcessor:

    def __init__(self, hua_file, data_file, ip):

        # output dfs
        self.employees = []
        self.employee_df = None
        self.details_df = None
        self.totals_df = None
        self.others_df = None

        # output file name
        self.output_file_name = None

        # input files and load case (1 if loading from hua and data, 2 if loading from previous solution)
        self.load_case = 2
        if hua_file:
            self.results_file = load_workbook(hua_file)
            self.load_case = 1
        self.data_file = load_workbook(data_file)

        # night differential hours
        self.nd_am = ip.nd_am
        self.nd_pm = ip.nd_pm

        # schedule details of format [dow_schedule_num, start, stop, duration, lunch_start, lunch_end, lunch_minutes]
        self.sched_details = []
        # for printing in format [dow, schedule_num, start, stop, duration, lunch_minutes, lunch_start, lunch_end]
        self.data_print = []
        self.read_in_schedules()

        # results of format [employee num, part time, [mon_sched, tue_sched, wed_sched, ...]]
        self.initial_results = []
        if self.load_case == 1:
            self.read_in_initial_results()
        else:
            self.read_in_previous_solution()

        # undo change (1 = employee type, 2 = moved schedule, 3 = swapped schedule) variable
        self.last_change_method = 0

        # undo change employee type variables
        self.last_change_index = None
        self.last_change_type = None

        # undo moved schedule variables
        self.last_from_index = None
        self.last_to_index = None
        self.last_moved_dow = None

        # undo swapped schedule variables
        self.last_a_index = None
        self.last_b_index = None
        self.last_swapped_dow = None

        # setup
        self.set_employees()
        self.set_original_sort_order()
        self.original_employees = copy.deepcopy(self.employees)
        self.set_dfs()
        self.calculate_hours_totals()
        self.calculate_other_totals()

    def read_in_schedules(self):

        ws = self.data_file["Data"]
        max_row = ws.max_row

        x = 2
        while x <= max_row:
            if ws["A" + str(x)].value in (None, "", " "):
                return
            dow = ws["A" + str(x)].value
            schedule_num = ws["B" + str(x)].value
            dow_schedule_num = str(dow) + "_" + str(schedule_num)
            start = ws["C" + str(x)].value
            stop = ws["D" + str(x)].value
            duration = ws["E" + str(x)].value
            lunch_minutes = ws["F" + str(x)].value
            lunch_start = ws["G" + str(x)].value
            lunch_end = ws["H" + str(x)].value
            self.sched_details.append([dow_schedule_num, start, stop, duration, lunch_start, lunch_end, lunch_minutes])
            self.data_print.append([dow, schedule_num, start, stop, duration, lunch_minutes, lunch_start, lunch_end])
            x += 1

    def read_in_initial_results(self):

        ws = self.results_file["SCHEDULES"]

        x = 2
        max_row = ws.max_row

        all_rows = []
        while x <= max_row:
            if ws["A" + str(x)].value in (None, "", " "):
                break

            if ws["D" + str(x)].value == "SOURCE":
                x += 1
            else:
                employee_num = ws["A" + str(x)].value
                if employee_num in (None, "", " "):
                    break
                part_time = ws["B" + str(x)].value
                schedule = ws["D" + str(x)].value
                all_rows.append([employee_num, part_time, schedule])
                x += 1

        employee_nums = list(set([x[0] for x in all_rows]))

        for employee_num in employee_nums:
            rows = [x for x in all_rows if x[0] == employee_num]
            part_time = rows[0][1]
            schedules = [None, None, None, None, None, None, None]
            for row in rows:
                split = row[2].split("_")
                dow = int(split[0]) - 1
                schedules[dow] = row[2]
            self.initial_results.append([employee_num, part_time, schedules])

    def read_in_previous_solution(self):

        self.initial_results = []

        ws = self.data_file["Staffing"]

        x = 6
        max_row = ws.max_row

        while x <= max_row:
            if ws["B" + str(x)].value in (None, "", " "):
                break

            employee_num = ws["B" + str(x)].value
            if employee_num in (None, "", " "):
                break
            part_time = ws["C" + str(x)].value
            mon = ws["D" + str(x)].value
            tue = ws["E" + str(x)].value
            wed = ws["F" + str(x)].value
            thu = ws["G" + str(x)].value
            fri = ws["H" + str(x)].value
            sat = ws["I" + str(x)].value
            sun = ws["J" + str(x)].value
            schedules = [mon, tue, wed, thu, fri, sat, sun]

            self.initial_results.append([employee_num, part_time, schedules])
            x += 1

    def set_employees(self):

        self.employees = []

        for employee in self.initial_results:
            schedules = []
            for day_sched_num in employee[2]:
                if day_sched_num:
                    schedules.append(next(x for x in self.sched_details if x[0] == day_sched_num))
                else:
                    schedules.append(None)

            part_time = str(employee[1])

            if part_time in ("1", "PTF"):
                employee_type = "PTF"
            elif part_time in ("PSE"):
                employee_type = "PSE"
            else:
                employee_type = "FT"

            new_employee = Employee(employee_num=employee[0], employee_type=employee_type, schedules=schedules,
                                    nd_am=self.nd_am, nd_pm=self.nd_pm)
            self.employees.append(new_employee)

    def undo_last_change(self):

        if self.last_change_method == 0:
            return

        if self.last_change_method == 1:
            self.undo_employee_type_change()

        if self.last_change_method == 2:
            self.undo_schedule_move()

        if self.last_change_method == 3:
            self.undo_schedule_swap()

    def change_an_employee_type(self, employee_index, new_type):

        if new_type not in ("PTF", "FT", "PSE"):
            print("bad employee type")
            return False

        self.last_change_type = self.employees[employee_index].employee_type
        self.last_change_index = employee_index
        self.last_change_method = 1

        self.employees[employee_index].change_type(new_type)
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def undo_employee_type_change(self):

        self.change_an_employee_type(self.last_change_index, self.last_change_type)

    def calculate_hours_totals(self):

        ft_lists = [x.annual_totals_list for x in self.employees if x.employee_type == "FT"]
        ptf_lists = [x.annual_totals_list for x in self.employees if x.employee_type == "PTF"]
        pse_lists = [x.annual_totals_list for x in self.employees if x.employee_type == "PSE"]

        ft_output = [0, 0, 0, 0, 0, 0, 0, 0]
        ptf_output = [0, 0, 0, 0, 0, 0, 0, 0]
        pse_output = [0, 0, 0, 0, 0, 0, 0, 0]

        if ft_lists:
            for x, _ in enumerate(ft_output):
                ft_output[x] = sum(y[x] for y in ft_lists)
        if ptf_lists:
            for x, _ in enumerate(ptf_output):
                ptf_output[x] = sum(y[x] for y in ptf_lists)
        if pse_lists:
            for x, _ in enumerate(pse_output):
                pse_output[x] = sum(y[x] for y in pse_lists)

        category_list = ["Annual Hours", "Lunch Hours", "Paid Hours", "Regular", "Night Diff.", "Sunday Diff.",
                         "Overtime", "Penalty Overtime"]
        # self.totals = [category_list, ft_output, ptf_output, pse_output]

        labels = ["Category", "FT", "PTF", "PSE"]
        compiled = []
        for x, item in enumerate(category_list):
            compiled.append([item, ft_output[x], ptf_output[x], pse_output[x]])
        self.totals_df = pd.DataFrame(compiled, columns=labels)

    def calculate_other_totals(self):

        ft_count = 0
        ft_pct = 0
        ft_hours = 0
        ft_hours_pct = 0
        ft_avg_week = 0
        ptf_count = 0
        ptf_pct = 0
        ptf_hours = 0
        ptf_hours_pct = 0
        ptf_avg_week = 0
        pse_count = 0
        pse_pct = 0
        pse_hours = 0
        pse_hours_pct = 0
        pse_avg_week = 0

        total_count = len(self.employees)
        total_hours = sum(x.weekly_worked_hours for x in self.employees)

        ft_list = [x for x in self.employees if x.employee_type == "FT"]
        if ft_list:
            ft_count = len(ft_list)
            ft_pct = ft_count / total_count
            ft_hours = sum(x.weekly_worked_hours for x in ft_list)
            ft_hours_pct = ft_hours / total_hours
            ft_avg_week = ft_hours / ft_count

        ptf_list = [x for x in self.employees if x.employee_type == "PTF"]
        if ptf_list:
            ptf_count = len(ptf_list)
            ptf_pct = ptf_count / total_count
            ptf_hours = sum(x.weekly_worked_hours for x in ptf_list)
            ptf_hours_pct = ptf_hours / total_hours
            ptf_avg_week = ptf_hours / ptf_count

        pse_list = [x for x in self.employees if x.employee_type == "PSE"]
        if pse_list:
            pse_count = len(pse_list)
            pse_pct = pse_count / total_count
            pse_hours = sum(x.weekly_worked_hours for x in pse_list)
            pse_hours_pct = pse_hours / total_hours
            pse_avg_week = pse_hours / pse_count

        category_list = ["Drivers", "% Staffed", "Avg. Schedule", "Tot. Hours", "% of Hours"]
        ft_comp = [ft_count, ft_pct, ft_avg_week, ft_hours, ft_hours_pct]
        ptf_comp = [ptf_count, ptf_pct, ptf_avg_week, ptf_hours, ptf_hours_pct]
        pse_comp = [pse_count, pse_pct, pse_avg_week, pse_hours, pse_hours_pct]
        labels = ["Position", "FT", "PTF", "PSE"]

        compiled = []
        for x, item in enumerate(category_list):
            compiled.append([item, ft_comp[x], ptf_comp[x], pse_comp[x]])

        self.others_df = pd.DataFrame(compiled, columns=labels)

    def move_a_schedule(self, employee_index_from, employee_index_to, schedule_index):

        schedule = self.employees[employee_index_from].remove_a_schedule(schedule_index)
        worked = self.employees[employee_index_to].add_a_schedule(schedule)

        if worked:
            print("schedule moved")
        else:
            print("couldn't add schedule")
            self.employees[employee_index_from].add_a_schedule(schedule)

        self.last_from_index = employee_index_from
        self.last_to_index = employee_index_to
        self.last_moved_dow = schedule_index
        self.last_change_method = 2

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def undo_schedule_move(self):

        self.move_a_schedule(self.last_to_index, self.last_from_index, self.last_moved_dow)

    def swap_two_schedules(self, employee_index_a, employee_index_b, schedule_index):

        if employee_index_b == employee_index_a:
            return True

        schedule_a = self.employees[employee_index_a].remove_a_schedule(schedule_index)
        schedule_b = self.employees[employee_index_b].remove_a_schedule(schedule_index)

        a_worked = self.employees[employee_index_a].add_a_schedule(schedule_b)

        if not a_worked:
            print("couldn't swap schedules")
            self.employees[employee_index_a].add_a_schedule(schedule_a)
            self.employees[employee_index_b].add_a_schedule(schedule_b)
            return False

        b_worked = self.employees[employee_index_b].add_a_schedule(schedule_a)

        if not b_worked:
            print("couldn't swap schedules")
            self.employees[employee_index_a].remove_a_schedule(schedule_index)
            self.employees[employee_index_a].add_a_schedule(schedule_a)
            self.employees[employee_index_b].add_a_schedule(schedule_b)
            return False

        print("swapped schedules")

        self.last_a_index = employee_index_a
        self.last_b_index = employee_index_b
        self.last_swapped_dow = schedule_index
        self.last_change_method = 3

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()
        return True

    def undo_schedule_swap(self):

        self.swap_two_schedules(self.last_a_index, self.last_b_index, self.last_swapped_dow)

    def find_eligible_slots(self, schedule):

        if not schedule:
            return None, None

        eligible_indices = []
        swappable_indices = []

        for x, employee in enumerate(self.employees):
            if employee.eligible_schedule_check(schedule):
                eligible_indices.append(x)
            else:
                swappable_indices.append(x)

        return eligible_indices, swappable_indices

    def create_a_new_employee(self):

        new_num = len(self.employees)
        empty_schedules = [None, None, None, None, None, None, None]

        new_employee = Employee(employee_num=new_num, employee_type="FT", schedules=empty_schedules,
                                nd_am=self.nd_am, nd_pm=self.nd_pm)

        self.employees.append(new_employee)
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def remove_blank_employees(self):

        empty_schedules = [None, None, None, None, None, None, None]
        blank_employees = [x for x in self.employees if x.schedules == empty_schedules]

        for employee in blank_employees:
            self.employees.remove(employee)

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def flatten_start_times(self, index):

        self.employees[index].flatten_schedules()
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def round_start_times(self, index):

        self.employees[index].round_schedules()
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def unflatten_start_times(self, index):

        self.employees[index].unflatten_schedules()
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def unround_start_times(self, index):

        self.employees[index].unround_schedules()
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def flatten_all(self):

        for employee in self.employees:
            employee.flatten_schedules()

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def unflatten_all(self):

        for employee in self.employees:
            employee.unflatten_schedules()

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def round_all(self):

        for employee in self.employees:
            employee.round_schedules()

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def unround_all(self):

        for employee in self.employees:
            employee.unround_schedules()

        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def restore_originals(self):

        self.employees = copy.deepcopy(self.original_employees)
        self.calculate_hours_totals()
        self.calculate_other_totals()
        self.set_dfs()

    def print_to_excel(self, new_file_name):

        print("printing")

        if new_file_name[-5:] != ".xlsx":
            new_file_name += ".xlsx"

        output_wb = load_workbook("Optimization Formats/Staffing_Formats_Static.xlsx")
        output_wb.save(new_file_name)

        ws = output_wb["Schedules"]
        row = 7
        for x, employee in enumerate(self.employees):
            employee.excel_print_schedules(ws, row+x)

        ws = output_wb["Staffing"]
        row = 6
        for x, employee in enumerate(self.employees):
            employee.excel_print_staffing(ws, row+x)

        ws = output_wb["By Employee"]
        row = 2
        for x, employee in enumerate(self.employees):
            employee.excel_print_by_employee(ws, row+x)

        ws = output_wb["Data"]
        for x, row in enumerate(self.data_print):
            for y, item in enumerate(row):
                ws.cell(row=x+2, column=y+1).value = item

        # ws = output_wb["Hours for Cost Model"]

        output_wb.save(new_file_name)
        output_wb.close()

        print("printed")

    def set_dfs(self):

        labels = ["Employee ID", "Type", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday",
                  "Sunday", "Valid Week", "Hours", "Days", "Equal Starts", "Round to 15", "O. Sort", "N. Sort"]

        compiled = []
        for employee in self.employees:
            compiled.append(employee.schedules_display_list)

        if len(self.employees) == 0:
            compiled.append([None, None, None, None, None, None, None, None,
                             None, None, None, None, None, None, None, None])

        self.employee_df = pd.DataFrame(compiled, columns=labels)

        labels = ["Employee ID", "Employee Type", "Total Hours", "Lunch Hours", "Worked Hours",
                  "Regular", "Night Diff.", "Sunday Diff.", "Overtime", "Penalty Overtime", "Valid Week"]

        compiled = []

        for employee in self.employees:
            compiled.append(employee.details_display_list)

        if len(self.employees) == 0:
            compiled.append([None, None, None, None, None, None, None, None, None, None, None])

        self.details_df = pd.DataFrame(compiled, columns=labels)

    def set_original_sort_order(self):

        sort_order = {"FT":3, "PTF":2, "PSE":1}

        self.employees.sort(key=lambda x: [sort_order[x.employee_type], x.weekly_worked_hours], reverse=True)
        for x, employee in enumerate(self.employees):
            employee.set_original_sort_order(x)

    def set_new_sort_order(self):

        sort_order = {"FT":3, "PTF":2, "PSE":1}

        self.employees.sort(key=lambda x: [sort_order[x.employee_type], x.weekly_worked_hours], reverse=True)
        for x, employee in enumerate(self.employees):
            employee.set_final_sort_order(x)

        self.set_dfs()

    def modify_existing_schedules(self, existing_schedules_file):

        # Compile a list of all manipulated schedules
        # schedules are of format [day_sched, start, stop, duration, lunch start, lunch stop, lunch duration]
        all_schedules = []
        for employee in self.employees:
            all_schedules += [x for x in employee.schedules if x]

        # split day of week from schedule num
        all_schedules = [x[0].split("_") + x[1:] for x in all_schedules]

        # Read in previous full schedules
        wb = load_workbook(existing_schedules_file)

        # [PVS site, schedule num, freq code, vehicle type, num stops, start datetime.time,
        # stop datetime.time, duration(minutes), optimizer schedule num]
        ws = wb["Schedule Summaries"]
        schedule_summaries = []
        for row in ws.iter_rows(min_row=2):
            schedule_summaries.append([cell.value for cell in row])

        # [schedule num, stop num, stop name, arrive datetime.time, depart datetime.time, original schedule num]
        ws = wb["Schedule Stops"]
        schedule_stops = []
        for row in ws.iter_rows(min_row=2):
            schedule_stops.append([cell.value for cell in row])

        wb.close()

        wb = load_workbook("Optimization Formats/final_schedule_format_to_pull.xlsx")
        new_file_name = "final_schedules" + today() + ".xlsx"
        wb.save(new_file_name)

        # create stop lists by day
        day_list = [[], [], [], [], [], [], []]
        single_list = [[], [], [], [], [], [], []]
        single_counter = 1
        eleven_ton_list = [[], [], [], [], [], [], []]
        eleven_ton_counter = 1

        for day in range(1, 8):
            day_schedules = [x for x in all_schedules if x[0] == str(day)]
            for schedule in day_schedules:
                stops = [x for x in schedule_stops if x[0] == schedule[1]]
                summary = next(x for x in schedule_summaries if x[1] == schedule[1])
                vehicle_type = summary[3]
                freq_code = summary[2]
                duration = schedule[4]
                start_time = cminute_to_tod(schedule[2])
                first_stop = stops[0]
                day_list[day-1].append([first_stop[5], schedule[1], first_stop[1], freq_code, vehicle_type, start_time,
                                        first_stop[2], first_stop[4]])
                for stop in stops[1:]:
                    day_list[day-1].append([stop[5], schedule[1], stop[1], freq_code, vehicle_type, stop[3],
                                            stop[2], stop[4]])

                # Populate the vehicle lists as well
                if vehicle_type in ("SINGLE", "Single", "single"):
                    single_list[day-1].append([schedule[1], single_counter, 0, schedule[2], schedule[3], duration])
                    single_counter += 1
                elif vehicle_type in ("11-TON", "11-Ton", "11-ton"):
                    eleven_ton_list[day-1].append([schedule[1], eleven_ton_counter, 0, schedule[2], schedule[3], duration])
                    eleven_ton_counter += 1

        for day, stop_list in enumerate(day_list):
            day_name = day_from_num(day+1)
            ws = wb[day_name + " Schedules"]
            row = 7
            for stop_row in stop_list:
                ws["A" + str(row)].value = row - 6
                for y, item in enumerate(stop_row):
                    ws.cell(row=row, column=y+2).value = item
                row += 1

        wb.save(new_file_name)
        wb.close()

        # Print out vehicle input
        if single_counter > 1:
            vehicle_wb = load_workbook("Optimization Formats/Vehicle_Formats.xlsx")
            new_file_name = "VehicleOptidata_Single_" + today() + ".xlsx"
            vehicle_wb.save(new_file_name)

            for day, schedule_list in enumerate(single_list):
                day_name = day_from_num(day+1)
                ws = vehicle_wb["OptiData " + day_name]
                day_start_min = ((1 + day) * 1440) - 240
                day_end_min = day_start_min + 2160

                ws["B7"].value = 0
                ws["C7"].value = 0
                ws["D7"].value = 0
                ws["E7"].value = day_start_min
                ws["F7"].value = day_start_min
                ws["G7"].value = 0

                ws["B8"].value = 0
                ws["C8"].value = 999
                ws["D8"].value = 0
                ws["E8"].value = day_end_min
                ws["F8"].value = day_end_min
                ws["G8"].value = 0

                row = 8
                for schedule in schedule_list:
                    row += 1
                    ws["A" + str(row)].value = row - 8
                    for y, item in enumerate(schedule):
                        ws.cell(row=row, column=y+2).value = item

            vehicle_wb.create_named_range("TRIPS" + str(day + 1), ws, "$B$7:$G$" + str(row))

            vehicle_wb.save(new_file_name)
            vehicle_wb.close()

        if eleven_ton_counter > 1:
            vehicle_wb = load_workbook("Optimization Formats/Vehicle_Formats.xlsx")
            new_file_name = "VehicleOptidata_11-Ton_" + today() + ".xlsx"
            vehicle_wb.save(new_file_name)

            for day, schedule_list in enumerate(eleven_ton_list):
                day_name = day_from_num(day+1)
                ws = vehicle_wb["OptiData " + day_name]
                day_start_min = ((1 + day) * 1440) - 240
                day_end_min = day_start_min + 2160

                ws["B7"].value = 0
                ws["C7"].value = 0
                ws["D7"].value = 0
                ws["E7"].value = day_start_min
                ws["F7"].value = day_start_min
                ws["G7"].value = 0

                ws["B8"].value = 0
                ws["C8"].value = 999
                ws["D8"].value = 0
                ws["E8"].value = day_end_min
                ws["F8"].value = day_end_min
                ws["G8"].value = 0

                row = 8
                for schedule in schedule_list:
                    row += 1
                    ws["A" + str(row)].value = row - 8
                    for y, item in enumerate(schedule):
                        ws.cell(row=row, column=y+2).value = item

            vehicle_wb.create_named_range("TRIPS" + str(day + 1), ws, "$B$7:$G$" + str(row))

            vehicle_wb.save(new_file_name)
            vehicle_wb.close()


class EmployeeModel(QAbstractTableModel):

    def __init__(self, data, details, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns
        self.selection_mode = False
        self.dow = None
        self.eligible_indices = None
        self.swappable_indices = None
        self.selected_index = None
        self.details_bool = details

    def rowCount(self, parent=None):
        try:
            if not self._data[0][0]:
                return 0
            return len(self._data)
        except:
            return 0

    def columnCount(self, parent=None):
        try:
            return len(self._data[0])
        except:
            return 0

    def data(self, index, role):

        if not index.isValid():
            return None

        row = index.row()
        col = index.column()
        val = self._data[row][col]

        if role == Qt.TextAlignmentRole:
            return Qt.AlignCenter

        if self.details_bool:
            if role == Qt.DisplayRole:
                if col in (2, 3, 4, 5, 6, 7, 8, 9):
                    return number(val)
                elif col in (14, 15):
                    return int(val)
                else:
                    return str(val)
            else:
                return None

        if col in (12, 13):
            if role == Qt.CheckStateRole:
                if val:
                    return Qt.Checked
                else:
                    return Qt.Unchecked
            else:
                return None

        if role == Qt.DisplayRole:
            if col == 10:
                return number_decimal(val, 1)
            else:
                return str(val)

        if role == Qt.EditRole:
            return val

        if role == Qt.BackgroundRole:
            if self.selection_mode:
                if self.eligible_indices:
                    if row in self.eligible_indices and col == (self.dow + 2):
                        return quick_color("green")
                if self.swappable_indices:
                    if row in self.swappable_indices and col == (self.dow + 2):
                        return quick_color("yellow")
            if not self._data[row][9]:
                return quick_color("red")

        return None

    def setData(self, index, value, role):

        row = index.row()
        col = index.column()

        if col in (12, 13):
            if role == Qt.CheckStateRole:
                self.dataChanged.emit(index, index, [])

        if role == Qt.EditRole:
            self._data[row][col] = value
            self.dataChanged.emit(index, index, [])

        super().setData(index, value, role)
        return True

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]

        return QVariant()

    def flags(self, index):

        col = index.column()
        if col == 1:
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

        if not self.details_bool:
            if col in (3, 4, 5, 6, 7, 8, 9):
                return Qt.ItemIsEnabled | Qt.ItemIsSelectable
            if col in (12, 13):
                return Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable

        return Qt.ItemIsEnabled | Qt.ItemIsSelectable

    def go_to_selection_mode(self, eligible_indices, swappable_indices, dow, selected_index):

        if self.selection_mode:
            self.leave_selection_mode()
        self.selection_mode = True
        self.eligible_indices = eligible_indices
        self.swappable_indices = swappable_indices
        self.dow = dow
        self.selected_index = selected_index
        top_left = self.createIndex(0, dow+2)
        bottom_right = self.createIndex(self.rowCount(), dow+3)
        self.dataChanged.emit(top_left, bottom_right, [])

    def leave_selection_mode(self):

        self.selection_mode = False
        self.eligible_indices = None
        self.selected_index = None

        top_left = self.createIndex(0, self.dow+2)
        bottom_right = self.createIndex(self.rowCount(), self.columnCount())

        self.dow = None

        self.dataChanged.emit(top_left, bottom_right, [])

    def update_table(self):

        top_left = self.createIndex(0, 0)
        bottom_right = self.createIndex(self.rowCount(), self.columnCount())
        self.dataChanged.emit(top_left, bottom_right, [])


class EmployeeView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, data, details):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(False)
        self.sorter.setSortRole(Qt.EditRole)
        self.sorter.setSourceModel(EmployeeModel(data, details))
        self.setModel(self.sorter)
        self.setSelectionMode(QTableView.SingleSelection)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)
        self.setColumnHidden(14, True)
        self.setColumnHidden(15, True)
        self.set_to_original_sort_order()

    def go_to_selection_mode(self, eligible_indices, swappable_indices, dow, selected_index):

        self.sorter.sourceModel().go_to_selection_mode(eligible_indices, swappable_indices, dow, selected_index)
        self.update()
        self.setModel(self.sorter)
        self.update()

    def leave_selection_mode(self):

        self.sorter.sourceModel().leave_selection_mode()
        self.update()
        self.setModel(self.sorter)
        self.update()

    def set_to_original_sort_order(self):

        self.sortByColumn(14, Qt.AscendingOrder)

    def set_to_new_sort_order(self):

        self.sortByColumn(15, Qt.AscendingOrder)

    def update_table(self):

        self.sorter.sourceModel().update_table()
        self.setModel(self.sorter)
        self.update()


class TotalsModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        try:
            if not self._data[0][0]:
                return 0
            return len(self._data)
        except:
            return 0

    def columnCount(self, parent=None):
        try:
            return len(self._data[0])
        except:
            return 0

    def data(self, index, role):
        if not index.isValid():
            return None

        row = index.row()
        col = index.column()
        val = self._data[row][col]

        # if role == Qt.TextAlignmentRole:
        #     return Qt.AlignCenter
        if role == Qt.DisplayRole:
            if col in (1, 2, 3):
                return number(val)
            else:
                return str(val)
        else:
            return None

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]

        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class TotalsView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, data):

        self.sorter = QSortFilterProxyModel()
        # self.sorter.setDynamicSortFilter(True)
        # self.sorter.setSortRole(Qt.EditRole)
        self.sorter.setSourceModel(TotalsModel(data))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        # self.setSortingEnabled(True)
        # self.setSelectionBehavior(QAbstractItemView.SelectRows)


class OthersModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        try:
            if not self._data[0][0]:
                return 0
            return len(self._data)
        except:
            return 0

    def columnCount(self, parent=None):
        try:
            return len(self._data[0])
        except:
            return 0

    def data(self, index, role):
        if not index.isValid():
            return None

        row = index.row()
        col = index.column()
        val = self._data[row][col]

        # if role == Qt.TextAlignmentRole:
        #     return Qt.AlignCenter
        if role == Qt.DisplayRole:
            if col in (1, 2, 3):
                if row in (1, 4):
                    return percent(val)
                if row == 2:
                    return number_decimal(val, 1)
                return number(val)
            else:
                return str(val)
        else:
            return None

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]

        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class OthersView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, data):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setSourceModel(OthersModel(data))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
