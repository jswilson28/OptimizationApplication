# This file contains all methods and classes related to reading in HCR Plates and producing a list of schedules.
from datetime import datetime, timedelta
from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine
import PyPDF2
import re
from ScheduleClasses import Schedule, Stop
from GeneralMethods import duration_between_two_times as dur
from GeneralMethods import time_to_datetime
from PyQt5.QtWidgets import (QGridLayout, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableView,
                             QDialog, QMessageBox, QRadioButton)
from PyQt5.QtCore import Qt, QAbstractTableModel
from openpyxl import load_workbook
import pandas as pd
from LookupReader import FacilityNameEntry, AddFrequencyCode


def hcr_to_time(var):
    return datetime.strptime(var, "%H%M").time()


class HalfStop:

    def __init__(self, time, tag, time_zone, stop_name):
        self.time = time
        self.tag = tag
        self.time_zone = time_zone
        self.stop_name = stop_name

    def push_stop(self, minutes):

        time_dt = datetime(100, 1, 1, self.time.hour, self.time.minute)
        self.time = (time_dt + timedelta(minutes=minutes)).time()

    def adjust_tz(self, to_tz):

        if to_tz in ('UK', "UK", None, ""):
            print("unknown timezones, adjust by hand")
            return

        if to_tz not in ('ET', 'CT', 'MT', 'PT', 'ADJ', 'PR', 'HI'):
            print("uh oh, timezonio")

        to_list = ['ET', 'CT', 'MT', 'PT', 'ADJ', 'PR', 'HI']
        et_list = [0, -1, -2, -3, 0, 0, 0]
        ct_list = [1, 0, -1, -2, 0, 0, 0]
        mt_list = [2, 1, 0, -1, 0, 0, 0]
        pt_list = [3, 2, 1, 0, 0, 0, 0]
        adj_list = [0, 0, 0, 0, 0, 0, 0]
        pr_list = [0, 0, 0, 0, 0, 0, 0]
        hi_list = [0, 0, 0, 0, 0, 0, 0]
        comp_list = [et_list, ct_list, mt_list, pt_list, adj_list, pr_list, hi_list]

        from_tz = self.time_zone

        from_index = to_list.index(from_tz)
        to_index = to_list.index(to_tz)

        minutes = 60 * comp_list[from_index][to_index]
        self.push_stop(minutes)
        if minutes != 0:
            self.time_zone = to_tz


class UPart:

    def __init__(self, part, schedule_num, freq_code, half_stops, vehicle_type, mileage):

        self.part = part
        self.schedule_num = schedule_num
        self.freq_code = freq_code.zfill(4)
        self.half_stops = half_stops
        self.vehicle_type = vehicle_type
        self.mileage = mileage
        self.fix_missing_times()
        self.annual_trips = None
        self.freq_code_desc = None
        self.original_freq_code = str(self.freq_code)

    def fix_missing_times(self):

        last_tag = ""
        stops_to_fix = []
        for stop in self.half_stops:
            if stop.tag == last_tag:
                stops_to_fix.append(stop)
            last_tag = stop.tag

        for stop in stops_to_fix:
            index = self.half_stops.index(stop)
            old_stop = stop
            new_datetime = datetime(100, 1, 1, old_stop.time.hour, old_stop.time.minute)
            new_datetime = new_datetime - timedelta(minutes=1)
            new_time = new_datetime.time()
            tag = "Ar"
            time_zone = old_stop.time_zone
            name = old_stop.stop_name
            new_stop = HalfStop(new_time, tag, time_zone, name)
            self.half_stops.insert(index, new_stop)

        return

    def add_arrival_to_start(self):

        first_stop = self.half_stops[0]
        one_min = timedelta(minutes=1)

        if first_stop.tag == 'Ar':
            return

        start_zone = first_stop.time_zone
        start_stop = first_stop.stop_name

        start_time = first_stop.time
        start_dt = datetime(100, 1, 1, start_time.hour, start_time.minute)
        start_dt -= one_min
        start_time = start_dt.time()

        new_first_stop = HalfStop(start_time, 'Ar', start_zone, start_stop)
        self.half_stops.insert(0, new_first_stop)

    def add_departure_to_end(self):

        last_stop = self.half_stops[-1]
        one_min = timedelta(minutes=1)

        if last_stop.tag == 'Lv':
            return

        start_zone = last_stop.time_zone
        start_stop = last_stop.stop_name

        start_time = last_stop.time
        start_dt = datetime(100, 1, 1, start_time.hour, start_time.minute)
        start_dt += one_min
        start_time = start_dt.time()

        new_last_stop = HalfStop(start_time, 'Lv', start_zone, start_stop)
        self.half_stops.append(new_last_stop)

    def add_hcr_freq_code_info(self, annual_trips, freq_code_desc):

        self.annual_trips = annual_trips
        self.freq_code_desc = freq_code_desc

    def raw_duration(self):

        return dur(self.half_stops[0].time, self.half_stops[-1].time)

    def __str__(self):

        return str(self.schedule_num)


class HCRReader:

    def __init__(self, pdf_name, lb, source_app, check_list, pdc_name=None, pdc_address=None):

        print("reading plate: " + pdf_name)

        self.source_app = source_app
        self.file_name = pdf_name

        self.min_check_time = check_list[0]
        self.max_check_time = check_list[1]
        self.max_combined_time = check_list[2]

        self.custom_pdc_name = pdc_name
        self.custom_pdc_address = pdc_address

        # read in switch codes
        self.from_codes = []
        self.to_codes = []
        ws = load_workbook("Lookups/SwitchCodes.xlsx")['Code to Switch Code']
        max_row = ws.max_row
        x = 2
        while x <= max_row:
            self.from_codes.append(ws["A" + str(x)].value)
            self.to_codes.append(ws["B" + str(x)].value)
            x += 1

        self.lb = lb
        self.source = "HCR"
        self.source_type = "PDF"
        self.all_data = None
        self.is_readable = False
        try:
            self.is_readable = self.read_in_pdf_pdfparser(self.file_name)
        except:
            pass

        if not self.is_readable:
            try:
                self.read_in_pdf_pypdf()
                self.is_readable = self.read_in_pdf_pdfparser("PDF Copy.pdf")
            except:
                QMessageBox.question(None, "Misread Plate", "Could not read plate " + self.file_name + ".",
                                     QMessageBox.Ok)
                self.is_readable = False
                return

        if not self.is_readable:
            print("Couldn't read PDF " + self.file_name)
            return

        self.plate_number = pdf_name.split('/')[-1][:-4]

        self.cargo_van_count = None
        self.tractor_sa_count = None
        self.tractor_ta_count = None
        self.trailer_count = None
        self.vehicle_contract = None
        self.read_in_vehicles()

        self.panels = []
        try:
            self.separate_panels()
        except:
            print("Could not separate panels on plate " + self.plate_number)
            self.is_readable = False
            return

        self.pvs_name = None
        self.pvs_pdc = None
        self.hcr_pdc = None
        self.backup_name = None
        self.likely_name = None
        self.short_name = None
        self.pdc_address = None

        if not self.set_names_by_index():
            self.is_readable = False
            return

        self.stop_addresses = []
        self.potential_pdcs = []
        self.picked_addresses = []

        try:
            self.get_stop_addresses()
        except:
            print("Could not read addresses on plate " + self.plate_number)
            self.is_readable = False
            return

        self.found_codes = []
        self.all_u_parts = []

        try:
            for panel in self.panels:
                self.all_u_parts.append(self.break_down_a_panel(panel))
        except:
            print("Couldn't break down panels on plate " + self.plate_number)
            self.is_readable = False
            return

        if not self.pdc_address:
            self.set_pdc_address()

        self.pdc_tz = self.set_pdc_tz_from_u_parts()
        self.adjust_all_stop_times()

        self.schedules = []
        self.new_codes = []
        self.read_in_frequency_codes()
        self.attach_freq_code_info_to_uparts()

        self.merge_u_parts()

        # self.attach_freq_code_info_to_schedules()
        self.set_day_strings()

    def read_in_pdf_pdfparser(self, file_name):

        all_data = []
        fp = open(file_name, 'rb')
        parser = PDFParser(fp)
        doc = PDFDocument()
        parser.set_document(doc)
        doc.set_parser(parser)
        doc.initialize('')
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        # Process each page contained in the document.
        for page in doc.get_pages():
            interpreter.process_page(page)
            layout = device.get_result()
            for lt_obj in layout:
                if isinstance(lt_obj, LTTextBox) or isinstance(lt_obj, LTTextLine):
                    all_data.append(lt_obj.get_text())

        self.all_data = all_data

        check1 = False
        check2 = False
        check3 = False
        check4 = False

        for line in self.all_data:
            if "SCHEDULE" in line:
                check1 = True
            if "EFFECTIVE" in line:
                check2 = True
            if "HCR:" in line:
                check3 = True
            if "SUPPLIER" in line:
                check4 = True

        print(check1, check2, check3, check4)
        return check1 and check2 and check3 and check4

    def read_in_pdf_pypdf(self):

        input_pdf = PyPDF2.PdfFileReader(open(self.file_name, mode='rb'))
        output_pdf = PyPDF2.PdfFileWriter()

        input_pdf_pages = input_pdf.getNumPages()
        for page_num in range (0, input_pdf_pages):
            output_pdf.addPage(input_pdf.getPage(page_num))

        output_pdf.write(open("PDF Copy.pdf", mode='wb'))

    def separate_panels(self):
        # separate schedule panels from rest of plate

        schedule_panels = []
        max_index = len(self.all_data) - 1
        match_strings = ["FREQUENCY", "IDENTIFICATION"]

        for x, line in enumerate(self.all_data):
            if all(word in line for word in match_strings):
                max_index = x
                break

        start_strings = ["PART", "TRIP", "FREQUENCY"]
        end_string = "MILEAGE"

        for x, line in enumerate(self.all_data[1:max_index]):
            if all(word in line for word in start_strings):
                if end_string not in line:
                    end_line = next(y for y in self.all_data[x:max_index] if end_string in y)
                    line = line + end_line
                    # print(line)
                schedule_panels.append(line)

        self.panels = schedule_panels

    def set_names_by_index(self):

        index = None
        likely_name = None

        if self.custom_pdc_address:
            self.pdc_address = self.custom_pdc_address

        if self.custom_pdc_name:
            if self.custom_pdc_name in self.lb.hcr_pdcs:
                index = self.lb.hcr_pdcs.index(self.custom_pdc_name)
            elif self.custom_pdc_name in self.lb.pvs_pdcs:
                index = self.lb.hcr_pdcs.index(self.custom_pdc_name)
            else:
                likely_name = self.custom_pdc_name

        else:
            info_string = self.all_data[1].splitlines()
            info_string2 = info_string[1]
            comma_index = info_string2.index(',') + 1
            hyphen_index = info_string2.index('-')
            backup_name = info_string2[comma_index:hyphen_index].strip()
            self.backup_name = backup_name

            for line in info_string:
                if any(x in line for x in self.lb.hcr_pdcs):
                    next1 = next(x for x in self.lb.hcr_pdcs if x in line)
                    index = self.lb.hcr_pdcs.index(next1)
                    break
                # if line in self.lb.hcr_pdcs:
                #     index = self.lb.hcr_pdcs.index(line)

            if not index:
                print("here")
                if backup_name in self.lb.hcr_pdcs:
                    # print("here??")
                    index = self.lb.hcr_pdcs.index(backup_name)

        if not index and index != 0:
            QMessageBox.question(None, "Facility Names", "Could not identify P&DC name in lookup table for plate " +
                                 self.plate_number + ".", QMessageBox.Ok)
            if not likely_name:
                likely_name = self.get_likely_name()
            new_names = FacilityNameEntry(self.plate_number, likely_name, self.lb, "InSorceror").exec_()
            if new_names:
                self.hcr_pdc = new_names[0]
                self.pvs_pdc = new_names[1]
                self.pvs_name = new_names[2]
                self.short_name = new_names[3]
                self.lb.hcr_pdcs.append(self.hcr_pdc)
                self.lb.pvs_pdcs.append(self.pvs_pdc)
                self.lb.pvs_names.append(self.pvs_name)
                self.lb.short_names.append(self.short_name)
                return True

            return False

        # if "33030" in self.plate_number:
        #     index = self.lb.hcr_pdcs.index("ROYAL PALM P&DC, FL")

        self.hcr_pdc = self.lb.hcr_pdcs[index]
        self.pvs_pdc = self.lb.pvs_pdcs[index]
        self.pvs_name = self.lb.pvs_names[index]
        self.short_name = self.lb.short_names[index]
        return True

    def get_likely_name(self):
        info_string = self.all_data[1].splitlines()
        match_string = "ADMINISTRATIVE OFFICIAL: "
        if match_string in info_string:
            likely_index = info_string.index(match_string) + 1
            return info_string[likely_index]
        else:
            return "Unknown"

    def get_stop_addresses(self):

        start_index = self.all_data.index('PHYSICAL LOCATION OF POINTS SERVED:\n')
        end_index = self.all_data.index('TRIP PURPOSE AND MAIL CLASS:\n')

        address_table = self.all_data[start_index + 1:end_index]

        # for x, row in enumerate(address_table[0:5]):
        #     print(x, row)

        del_list = []
        for item in address_table:
            if "HCR:" in item:
                del_list.append(item)
            elif "EFFECTIVE:" in item:
                del_list.append(item)
            elif "TRIP " in item:
                del_list.append(item)
            elif len(item) < 4 and "--" not in item:
                del_list.append(item)

        for item in del_list:
            address_table.remove(item)

        # for x, row in enumerate(address_table[0:5]):
        #     print(x, row)

        address_lines = []
        for line in address_table:
            for subline in line.splitlines():
                address_lines.append(subline)

        phone_num_indices = []
        pattern = '\d\d\d-\d\d\d-\d\d\d\d'
        r = re.compile(pattern)
        for x, address in enumerate(address_lines):
            if r.search(address) or "--" in address:
                phone_num_indices.append(x)

        x = 0
        address_string = []
        address_strings = []

        for num in phone_num_indices:
            while x < num:
                address_string.append(address_lines[x].strip())
                x += 1
            address_strings.append(address_string)
            address_string = []
            x += 1

        for address in address_strings:
            address_edit = [x for x in address if x != '']
            address_string = ""
            name = address_edit[0]
            for line in address_edit[1:]:
                address_string = address_string + line + " "
            address_to_add = address_string.strip()
            self.stop_addresses.append([name, address_to_add])

        # for line in self.stop_addresses:
        #     print(line)

        self.potential_pdcs = [x for x in self.stop_addresses if "P&DC" in x[0]]
        # print(self.potential_pdcs)

    def set_pdc_address(self):

        if self.pdc_address:
            return

        possible_names = (self.hcr_pdc, self.backup_name, self.hcr_pdc[:-4], self.backup_name[:-4])
        potential_addresses = [x for x in self.stop_addresses if x[0] in possible_names]

        if len(potential_addresses) < 1:
            print("Error finding P&DC address on plate " + self.plate_number)
            return
        elif len(potential_addresses) > 1:
            potential_addresses = [x for x in self.stop_addresses if x[0] in (self.hcr_pdc, self.hcr_pdc[:-4])]

        self.pdc_address = potential_addresses[0][1]

    def try_harder_stop_address(self, possible_names):

        possible_addresses = [x[1] for x in self.stop_addresses if any(y in x[1] for y in possible_names)]
        new_addresses = []

        if not possible_addresses:
            return []

        # print(possible_addresses)
        # print(possible_names)

        for possible_address in possible_addresses:
            split_up = possible_address.split()
            try:
                name = next(x for x in split_up if x in possible_names)
                name_index = split_up.index(name)
                name_len = 1
            except:
                name_full = next(x for x in possible_names if x in possible_address).split()
                name = name_full[0]
                name_len = len(name_full)
                name_index = split_up.index(name)

            address_string = ""
            name = ""
            print(name_index, name_len)
            for x in split_up[name_index:name_index + name_len]:
                name += " " + x

            name = name.strip()
            # print(name)
            for word in split_up[name_index + name_len:]:
                address_string += word + " "
            address_string.strip()
            new_addresses.append([name, address_string])

        return new_addresses

    def get_stop_address(self, stop_name):

        possible_names = (stop_name, stop_name[:-3], stop_name.split(",")[0], stop_name.split(", ")[0])
        found_addresses = [x for x in self.stop_addresses if x[0] in possible_names]

        if not found_addresses:
            found_addresses = self.try_harder_stop_address(possible_names)

        if len(found_addresses) < 1:
            print("Couldn't find address for " + stop_name + " " + self.plate_number)
            self.stop_addresses.append([stop_name, "No address found"])
            return "No address found"
        if len(found_addresses) > 1:
            print("Multiple addresses found for " + stop_name)
            if stop_name in [x[0] for x in self.picked_addresses]:
                return next(x[1] for x in self.picked_addresses if x[0] == stop_name)
            else:
                picked_address = AddressPicker(self.plate_number, stop_name, found_addresses).exec_()
                self.picked_addresses.append([stop_name, picked_address[1]])
                return picked_address[1]
        else:
            return found_addresses[0][1]

    def break_down_a_panel(self, panel):

        rows = panel.splitlines()
        first_line = rows[0].split()
        mid = first_line.index('PART')
        left_side_cols = mid
        right_side_cols = len(first_line) - (mid + 1)

        schedule_num_line = rows[1].split()
        freq_code_line = rows[2].split()
        mileage_line = rows[-1].split()
        vehicle_line = rows[-2].split()

        match_string = "------------"
        break_lines = [i for i, x in enumerate(rows) if match_string in x]

        break_one = break_lines[1] + 1
        break_two = break_lines[2]

        stop_names = []
        time_zone_indices = []
        name_string = ""

        for line in rows[break_one:break_two]:
            words = line.split()
            for i, word in enumerate(words):
                if word in ('CT', 'MT', 'PT', 'ET', 'PR', 'HI'):
                    time_zone_index = i
            num_words = time_zone_index - (left_side_cols + 1)
            x = 1
            while x <= num_words:
                name_string = name_string + " " + words[x + left_side_cols]
                x += 1
            name_string = name_string.replace(",", ", ")
            stop_names.append(name_string.strip())
            time_zone_indices.append(time_zone_index)
            name_string = ""

        u_parts = []
        x = 0

        while x < left_side_cols:
            part = first_line[x]
            schedule_num = int(schedule_num_line[x])
            freq_code = freq_code_line[x]
            self.found_codes.append(freq_code)
            mileage = float(mileage_line[x])
            vehicle = vehicle_line[x]
            half_stops = []

            for y, line in enumerate(rows[break_one:break_two]):
                words = line.split()
                tag = words[left_side_cols]
                time_zone = words[time_zone_indices[y]]
                stop_name = stop_names[y]
                if words[x] != "--":
                    time = hcr_to_time(words[x])
                    new_stop = HalfStop(time, tag, time_zone, stop_name)
                    half_stops.append(new_stop)

            new_u_part = UPart(part, schedule_num, freq_code, half_stops, vehicle, mileage)
            u_parts.append(new_u_part)
            x += 1

        x = 0

        while x < right_side_cols:
            part = first_line[x - right_side_cols]
            schedule_num = int(schedule_num_line[x - right_side_cols])
            freq_code = freq_code_line[x - right_side_cols]
            self.found_codes.append(freq_code)
            mileage = float(mileage_line[x - right_side_cols])
            vehicle = vehicle_line[x - right_side_cols]
            half_stops = []

            for y, line in enumerate(rows[break_one:break_two]):
                words = line.split()
                tag = ""
                for i, word in enumerate(words[time_zone_indices[y]:]):
                    if word in ('Lv', 'Ar'):
                        tag = word
                time_zone = words[time_zone_indices[y]]
                stop_name = stop_names[y]
                if words[x - right_side_cols] != "--":
                    time = hcr_to_time(words[x - right_side_cols])
                    new_stop = HalfStop(time, tag, time_zone, stop_name)
                    half_stops.append(new_stop)

            half_stops.reverse()
            new_u_part = UPart(part, schedule_num, freq_code, half_stops, vehicle, mileage)
            u_parts.append(new_u_part)
            x += 1

        return u_parts

    def set_pdc_tz_from_u_parts(self):

        stop_tzs = []

        for panel in self.all_u_parts:
            for u_part in panel:
                for stop in u_part.half_stops:
                    stop_tzs.append(stop.time_zone)
                    if stop.stop_name == self.hcr_pdc:
                        return stop.time_zone

        stop_tzs = list(set(stop_tzs))
        if len(stop_tzs) == 1:
            return stop_tzs[0]

        if "LOUISVILLE" in self.hcr_pdc:
            return "ET"

        print ("Time zone not found??")
        return "UK"

    def adjust_all_stop_times(self):

        for panel in self.all_u_parts:
            for u_part in panel:
                for stop in u_part.half_stops:
                    stop.adjust_tz(self.pdc_tz)

    def merge_u_parts(self):

        schedules = []
        compiled_schedules = []
        unmatched_u_parts = []

        for x, panel in enumerate(self.all_u_parts):
            match_1 = False
            match_2 = False
            match_3 = False
            match_4 = False
            match_5 = False

            length = len(panel)

            if length > 0:
                u_part_1 = panel[0]
                schedule_1 = u_part_1.schedule_num
            if length > 1:
                u_part_2 = panel[1]
                schedule_2 = u_part_2.schedule_num
            if length > 2:
                u_part_3 = panel[2]
                schedule_3 = u_part_3.schedule_num
            if length > 3:
                u_part_4 = panel[3]
                schedule_4 = u_part_4.schedule_num
            if length > 4:
                u_part_5 = panel[4]
                schedule_5 = u_part_5.schedule_num

            if length == 1:
                unmatched_u_parts.append(u_part_1)

            if length == 2:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)

            if length == 3:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_1 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_3))
                    match_1 = True
                    match_3 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)

            if length == 4:
                if schedule_1 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_4))
                    match_1 = True
                    match_4 = True

                if schedule_2 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_3))
                    match_2 = True
                    match_3 = True

                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_3 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_4))
                    match_3 = True
                    match_4 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)
                if not match_4:
                    unmatched_u_parts.append(u_part_4)

            if length == 5:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_1 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_3))
                    match_1 = True
                    match_3 = True

                if schedule_1 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_4))
                    match_1 = True
                    match_4 = True

                if schedule_1 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_5))
                    match_1 = True
                    match_5 = True

                if schedule_2 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_3))
                    match_2 = True
                    match_3 = True

                if schedule_2 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_4))
                    match_2 = True
                    match_4 = True

                if schedule_2 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_5))
                    match_2 = True
                    match_5 = True

                if schedule_3 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_4))
                    match_3 = True
                    match_4 = True

                if schedule_3 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_5))
                    match_3 = True
                    match_5 = True

                if schedule_4 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_4, u_part_4))
                    match_4 = True
                    match_5 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)
                if not match_4:
                    unmatched_u_parts.append(u_part_4)
                if not match_5:
                    unmatched_u_parts.append(u_part_5)

        for part in unmatched_u_parts:
            schedules.append([self.u_part_to_schedule(part)])

        for items in schedules:
            for item in items:
                compiled_schedules.append(item)

        self.schedules = compiled_schedules

    def u_part_to_schedule(self, u_part):

        u_part.add_arrival_to_start()
        u_part.add_departure_to_end()

        combined_stops = []
        for x, half_stop in enumerate(u_part.half_stops):
            if half_stop.tag == 'Ar':
                next_stop = u_part.half_stops[x+1]
                new_stop = Stop(arrive_time=half_stop.time, depart_time=next_stop.time,
                                stop_name=half_stop.stop_name)
                new_stop.add_address(self.get_stop_address(new_stop.stop_name))
                combined_stops.append(new_stop)

        vc = u_part.vehicle_type[0]

        if vc == "T":
            vehicle_category = "Single"
        # else vc == "S":
        else:
            vehicle_category = "11-Ton"
        # elif self.cargo_van_count > 0:
        #     vehicle_category = "Cargo Van"
        # else:
        #     vehicle_category = "Unknown"

        annual_trips = float(u_part.annual_trips)

        return_schedule = Schedule(source_type=self.source_type, source=self.source, part=u_part.part,
                                   plate_num=self.plate_number, trip_num=u_part.schedule_num,
                                   freq_code=u_part.freq_code, stops=combined_stops, vehicle_type=u_part.vehicle_type,
                                   hcr_pdc=self.hcr_pdc, pvs_pdc=self.pvs_pdc, pvs_name=self.pvs_name,
                                   pdc_address=self.pdc_address, mileage=u_part.mileage,
                                   vehicle_category=vehicle_category, short_name=self.short_name,
                                   annual_trips=annual_trips)

        return_schedule.add_flag("This schedule is only one column of a U")

        return return_schedule

    def merge_two_u_parts(self, u_part_1, u_part_2):

        merged_stops = []
        # at this stage, HCR schedules should start at a 'Lv' and end at an 'Ar'
        # note this is done here and not in the UPart itself, because of the merge
        # this is where we should check the layover as well

        merge_anyway = True
        stop1 = u_part_1.half_stops[-1].time
        stop2 = u_part_2.half_stops[0].time
        layover = dur(stop1, stop2)

        switch_code_check = False
        if u_part_1.freq_code in self.from_codes:
            if u_part_2.freq_code == self.to_codes[self.from_codes.index(u_part_1.freq_code)]:
                switch_code_check = True

        different_day_check = False
        left_start = u_part_1.half_stops[0].time
        right_start = u_part_2.half_stops[0].time
        if time_to_datetime(left_start) > time_to_datetime(right_start):
            different_day_check = True

        dead_head_check = True
        if len(u_part_1.half_stops) == len(u_part_2.half_stops):
            for x, stop in enumerate(u_part_1.half_stops):
                other_stop = u_part_2.half_stops[x]
                if other_stop.time == stop.time:
                    continue
                else:
                    dead_head_check = False
                    break
        else:
            dead_head_check = False

        if self.source_app == "Optimizer" and not dead_head_check:

            check_time = self.min_check_time
            check_time_2 = self.max_check_time
            check_time_3 = self.max_combined_time
            combined_duration = u_part_1.raw_duration() + u_part_2.raw_duration() + layover

            if not u_part_1.freq_code == u_part_2.freq_code:
                if not different_day_check:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                elif switch_code_check:
                    merge_anyway = True
                else:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                    if merge_anyway:
                        ask_string = "Apply this answer to future occurrences of this frequency code combination?"
                        add_to_list = QMessageBox.question(None, "Remember this answer?", ask_string,
                                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if add_to_list == QMessageBox.Yes:
                            wb = load_workbook("Lookups/SwitchCodes.xlsx")
                            ws = wb['Code to Switch Code']
                            new_row = ws.max_row + 1
                            ws["A" + str(new_row)].value = u_part_1.freq_code
                            ws["B" + str(new_row)].value = u_part_2.freq_code
                            wb.save("Lookups/SwitchCodes.xlsx")
                            wb.close()
                            self.from_codes.append(u_part_1.freq_code)
                            self.to_codes.append(u_part_2.freq_code)
            elif layover < check_time:
                merge_anyway = True
            elif layover > check_time_2:
                merge_anyway = False
            elif combined_duration > check_time_3:
                merge_anyway = False
            elif check_time < layover < check_time_2:
                merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()

        elif not dead_head_check and self.source_app != "Test":
            if not u_part_1.freq_code == u_part_2.freq_code:
                if not different_day_check:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                elif switch_code_check:
                    merge_anyway = True
                else:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                    if merge_anyway:
                        ask_string = "Apply this answer to future occurrences of this frequency code combination?"
                        add_to_list = QMessageBox.question(None, "Remember this answer?", ask_string,
                                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if add_to_list == QMessageBox.Yes:
                            wb = load_workbook("Lookups/SwitchCodes.xlsx")
                            ws = wb['Code to Switch Code']
                            new_row = ws.max_row + 1
                            ws["A" + str(new_row)].value = u_part_1.freq_code
                            ws["B" + str(new_row)].value = u_part_2.freq_code
                            wb.save("Lookups/SwitchCodes.xlsx")
                            wb.close()
                            self.from_codes.append(u_part_1.freq_code)
                            self.to_codes.append(u_part_2.freq_code)
        else:
            merge_anyway = False

        if not merge_anyway:
            return [self.u_part_to_schedule(u_part_1), self.u_part_to_schedule(u_part_2)]

        u_part_1.add_arrival_to_start()
        u_part_2.add_departure_to_end()

        for half_stop in u_part_1.half_stops:
            merged_stops.append(half_stop)
        for half_stop in u_part_2.half_stops:
            merged_stops.append(half_stop)

        combined_stops = []
        for x, half_stop in enumerate(merged_stops):
            if half_stop.tag == 'Ar':
                next_stop = merged_stops[x+1]
                new_stop = Stop(arrive_time=half_stop.time, depart_time=next_stop.time,
                                stop_name=half_stop.stop_name, time_zone=half_stop.time_zone)
                new_stop.add_address(self.get_stop_address(new_stop.stop_name))
                combined_stops.append(new_stop)

        part = u_part_1.part
        schedule_num = u_part_1.schedule_num

        freq_code = u_part_1.freq_code

        mileage = u_part_1.mileage + u_part_2.mileage

        vt_left = u_part_1.vehicle_type
        vt_right = u_part_2.vehicle_type
        vc_left = vt_left[0]
        vc_right = vt_right[0]

        if vt_left == vt_right:
            vehicle_type = vt_left
        else:
            vehicle_type = vt_left + "/" + vt_right + " Unmatched"

        if vc_left == vc_right:
            if vc_left == "T":
                vehicle_category = "Single"
            else:
            # elif vc_left == "S":
                vehicle_category = "11-Ton"
        #     elif self.cargo_van_count > 0:
        #         vehicle_category = "Cargo Van"
        #     else:
        #         vehicle_category = "Matched, Unknown"
        else:
            vehicle_category = "Unmatched, Unknown"

        annual_trips = None
        if float(u_part_1.annual_trips) == float(u_part_2.annual_trips):
            annual_trips = float(u_part_1.annual_trips)

        return_schedule = Schedule(source_type=self.source_type, source=self.source, part=part,
                                   plate_num=self.plate_number, trip_num=schedule_num, freq_code=freq_code,
                                   stops=combined_stops, vehicle_type=vehicle_type, hcr_pdc=self.hcr_pdc,
                                   pvs_pdc=self.pvs_pdc, pvs_name=self.pvs_name, pdc_address=self.pdc_address,
                                   mileage=mileage, vehicle_category=vehicle_category, short_name=self.short_name,
                                   annual_trips=annual_trips)

        return [return_schedule]

    def read_in_frequency_codes(self):

        first_check = ["FREQUENCY", "IDENTIFICATION"]
        first_line = next(x for x in self.all_data if all(word in x for word in first_check))
        if not first_line:
            print("Couldn't find frequency table")
            return
        first_index = self.all_data.index(first_line)

        last_check = ["VEHICLE", "REQUIREMENTS"]
        last_line = next(x for x in self.all_data if all(word in x for word in last_check))
        if not last_line:
            print("Couldn't find end of frequency table")
            return
        last_index = self.all_data.index(last_line)

        frequency_table = self.all_data[first_index:last_index]
        found_codes = list(set(self.found_codes))
        found_codes.sort()

        frequency_table = [x.strip() for x in frequency_table]
        listed_codes = [x for x in frequency_table if x in found_codes]

        # this should only be triggered if there's a page num amidst the freq table that is the same as a freq code

        pattern = '\.\d\d'
        r = re.compile(pattern)
        last_panel_index = self.all_data.index(self.panels[-1])

        annual_trips = [x.strip() for x in self.all_data[last_panel_index:] if r.search(x) and len(x.strip()) < 10]

        if len(annual_trips) < len(listed_codes):
            non_nine_duplicates = [x for x in listed_codes if listed_codes.count(x) > 1 and "9" not in x]
            non_nine_duplicates = list(set(non_nine_duplicates))
            listed_codes.reverse()
            for item in non_nine_duplicates:
                listed_codes.remove(item)
            listed_codes.reverse()

        if len(annual_trips) < len(listed_codes):
            seen = set()
            listed_codes = [x for x in listed_codes if not (x in seen or seen.add(x))]

        if len(annual_trips) != len(listed_codes):
            print(annual_trips)
            print(listed_codes)
            print("Misread frequency code")
            return

        words = ("ANNUAL", "FREQUENCY", "IDENTIFICATION", "PAGE", "EFFECTIVE", "HCR:", "LENGTH", "CUBES", "PAYLOAD",
                 "END", "12/24", "12/23", "12/22", "12/21")

        list_descriptions = [x for x in frequency_table if x not in found_codes and x not in annual_trips and
                             not any(word in x for word in words) and not x.isdigit() and x not in ('', ' ', None)]

        if len(list_descriptions) != len(annual_trips):
            print(list_descriptions)
            print(annual_trips)
            print(listed_codes)
            print("Misread frequency descriptions")
            return

        # for x in range(0,len(list_descriptions)):
        #     print(listed_codes[x], annual_trips[x], list_descriptions[x])

        self.new_codes = []
        for x, code in enumerate(listed_codes):
            code = code.zfill(4)
            self.new_codes.append([code, annual_trips[x], list_descriptions[x]])

        for new_code in self.new_codes:
            if "9" not in new_code[0]:
                if new_code[0] not in [x[0] for x in self.lb.known_codes]:
                        AddFrequencyCode(new_code[0], new_code[1], new_code[2], self.lb).exec_()
                elif not next(x[1] for x in self.lb.known_codes if new_code[0] == x[0]):
                    self.lb.add_hcr_info(new_code)

    def attach_freq_code_info_to_uparts(self):

        for panel in self.all_u_parts:
            for u_part in panel:
                frequency_code = next(x for x in self.new_codes if x[0] == u_part.freq_code)
                if not frequency_code:
                    print("Couldn't find listed frequency code!")
                u_part.add_hcr_freq_code_info(frequency_code[1], frequency_code[2])

    def read_in_vehicles(self):

        match_strings = ["VEHICLE", "REQUIREMENTS"]

        first_line = None
        try:
            first_line = next(x for x in self.all_data if all(y in x for y in match_strings))
        except:
            print("oops")

        if not first_line:
            print("Couldn't find end of frequency table")
            return
        first_index = self.all_data.index(first_line)

        one = "DATES"
        two = "SEASONAL"

        three = "SCHEDULE"
        four = "NOTES"

        five = "PHYSICAL"
        six = "LOCATION"

        last_line = None

        try:
            last_line = next(x for x in self.all_data if (one in x and two in x))
        except:
            try:
                last_line = next(x for x in self.all_data if (three in x and four in x))
            except:
                try:
                    last_line = next(x for x in self.all_data if (five in x and six in x))
                except:
                    print("oops!")

        if not last_line:
            print("Couldn't find end of vehicle table")
            return

        last_index = self.all_data.index(last_line)

        vehicle_table = self.all_data[first_index:last_index]

        vehicle_table = [x.strip() for x in vehicle_table]

        page_list = ["PAGE"]
        page_numbers = [x for x in vehicle_table if any(word in x for word in page_list)]
        page_nums = []

        for page in page_numbers:
            page_nums.append(vehicle_table[vehicle_table.index(page) + 1])

        for num in page_nums:
            vehicle_table.remove(num)

        # print(self.plate_number, vehicle_table)
        vehicle_quantities = [x for x in vehicle_table if x.isdigit() and int(x) < 1000]

        # bad_vehicles = ['Boat', 'Four Wheel Drive', 'Passenger Car', 'Pick-up Truck', 'Station Wagon']
        peak_list = ["PEAK", "Peak", "peak", "SPB", "-SPB"]

        vehicles_listed = []
        full_names = []

        for item in vehicle_table:
            item.replace('\n', '')
            item.replace('/n', '')
            if item == "(PEAK)":
                full_names[-1] = full_names[-1] + " " + item
                if not vehicles_listed[-1] == "Peak Something":
                    vehicles_listed[-1] = "Peak Something"
            # elif not any(word in item for word in bad_vehicles):
            else:
                if any(word in item for word in peak_list):
                    full_names.append(item)
                    vehicles_listed.append("Peak Something")
                elif "Tractor" in item:
                    if any(word in item for word in ("Two Axle", "Tandem")):
                        full_names.append(item)
                        vehicles_listed.append("Tractor (TA)")
                    elif any(word in item for word in ("Single Axle", "Single", "single")):
                        full_names.append(item)
                        vehicles_listed.append("Tractor (SA)")
                    elif any(word in item for word in ("SPOTTER", "Spotter", "spotter")):
                        full_names.append(item)
                        vehicles_listed.append("Spotter Tractor")
                elif "Trailer" in item:
                    full_names.append(item)
                    vehicles_listed.append("Trailer")
                elif any(word in item for word in ("TRUCK", "truck", "Truck", "VAN", "van", "Van",
                                                   "CAR", "Car", "car", "WAGON", "Wagon", "wagon")):
                    full_names.append(item)
                    vehicles_listed.append("Cargo Van")

        vehicles_paired = []

        if len(vehicles_listed) != len(vehicle_quantities):
            # first thing, check if any vehicle names have a length, then remove that from after the name
            removables = []
            for vehicle in full_names:
                for word in vehicle.split():
                    if word.isdigit():
                        removables.append(word)
            removables.reverse()
            vehicle_quantities.reverse()
            for removable in removables:
                if removable in vehicle_quantities:
                    vehicle_quantities.remove(removable)

            if len(vehicles_listed) < len(vehicle_quantities):
                if "Cargo Van" in vehicles_listed and "20" in vehicle_quantities:
                    vehicle_quantities.reverse()
                    vehicle_quantities.remove("20")
                    vehicle_quantities.reverse()
                elif "Cargo Van" in vehicles_listed and "18" in vehicle_quantities:
                    vehicle_quantities.reverse()
                    vehicle_quantities.remove("18")
                    vehicle_quantities.reverse()

            while len(vehicles_listed) < len(vehicle_quantities):
                remove = [x for x in vehicle_quantities if int(x) >= 22]
                if remove:
                    vehicle_quantities.remove(remove[0])
                    # print(vehicles_listed)
                    # print(vehicle_quantities)
                else:
                    print("Vehicles table not read correctly on plate " + self.plate_number)
                    break

            vehicle_quantities.reverse()

            if len(vehicles_listed) != len(vehicle_quantities):
                print(vehicles_listed)
                print(vehicle_quantities)
                print("Vehicles table not read correctly on plate " + self.plate_number)

        for x, vehicle in enumerate(vehicles_listed):
            vehicles_paired.append([vehicle, vehicle_quantities[x]])

        # for row in vehicles_paired:
        #     print(row)

        vehicles_paired = [x for x in vehicles_paired if not x[0] == "Peak Something"]

        cargo_vans = [x for x in vehicles_paired if x[0] == "Cargo Van"]
        self.cargo_van_count = 0
        for pair in cargo_vans:
            self.cargo_van_count += int(pair[1])

        tractor_sa = [x for x in vehicles_paired if x[0] == "Tractor (SA)"]
        self.tractor_sa_count = 0
        for pair in tractor_sa:
            self.tractor_sa_count += int(pair[1])

        tractor_ta = [x for x in vehicles_paired if x[0] == "Tractor (TA)"]
        self.tractor_ta_count = 0
        for pair in tractor_ta:
            self.tractor_ta_count += int(pair[1])

        trailers = [x for x in vehicles_paired if x[0] == "Trailer"]
        self.trailer_count = 0
        for pair in trailers:
            self.trailer_count += int(pair[1])

    def set_day_strings(self):

        for schedule in self.schedules:
            schedule.set_bin_string(self.lb)
            schedule.is_holiday_schedule()

    def get_potential_pdcs_and_addresses(self):

        addresses = []
        potential_pdc_names = [x[0] for x in self.potential_pdcs]
        potential_pdc_addresses = [x[1] for x in self.potential_pdcs]

        return potential_pdc_names, potential_pdc_addresses

    # def change_pdc(self):
    #
    #     for schedule in self.schedules:
    #         schedule.change_pdc()


class MergeTwoUParts(QDialog):

    def __init__(self, u_part_1, u_part_2, layover, plate_number, parent=None):
        super(MergeTwoUParts, self).__init__(parent)

        self.u_part_1 = u_part_1
        self.u_part_2 = u_part_2
        self.layover = layover
        self.combined_duration = u_part_1.raw_duration() + u_part_2.raw_duration() + layover

        self.setWindowTitle(str(plate_number) + " : Verify Schedule Combination")
        self.setGeometry(300, 150, 650, 400)

        self.initUI()

        self.return_bool = None
        self.no_button = None
        self.yes_button = None

    def initUI(self):

        vbox = QVBoxLayout()
        vbox.addWidget(QLabel("Are these one schedule?"))
        vbox.addWidget(QLabel("(Layover: " + str(self.layover) + " minutes)"))
        vbox.addWidget(QLabel("(Combined Duration: " + str(self.combined_duration) + " minutes)"))
        hbox = QHBoxLayout()

        left_vbox = QVBoxLayout()
        left_grid = QGridLayout()
        left_grid.addWidget(QLabel("Schedule Part: "), 0, 0)
        left_grid.addWidget(QLabel("Trip Number: "), 1, 0)
        left_grid.addWidget(QLabel("Frequency Code: "), 2, 0)
        left_grid.addWidget(QLabel("Description: "), 3, 0)

        left_part = self.u_part_1.part
        left_num = str(self.u_part_1.schedule_num)
        left_freq = self.u_part_1.freq_code
        left_desc = self.u_part_1.freq_code_desc

        left_grid.addWidget(QLabel(left_part), 0, 1)
        left_grid.addWidget(QLabel(left_num), 1, 1)
        left_grid.addWidget(QLabel(left_freq), 2, 1)
        left_grid.addWidget(QLabel(left_desc), 3, 1)
        left_vbox.addLayout(left_grid)

        left_table = UPartView()
        left_table.set_model(UPartToPandas(self.u_part_1).df)
        left_vbox.addWidget(left_table)

        hbox.addLayout(left_vbox)

        right_vbox = QVBoxLayout()
        right_grid = QGridLayout()
        right_grid.addWidget(QLabel("Schedule Part: "), 0, 0)
        right_grid.addWidget(QLabel("Trip Number: "), 1, 0)
        right_grid.addWidget(QLabel("Frequency Code: "), 2, 0)
        right_grid.addWidget(QLabel("Description: "), 3, 0)

        right_part = self.u_part_2.part
        right_num = str(self.u_part_2.schedule_num)
        right_freq = self.u_part_2.freq_code
        right_desc = self.u_part_2.freq_code_desc

        right_grid.addWidget(QLabel(right_part), 0, 1)
        right_grid.addWidget(QLabel(right_num), 1, 1)
        right_grid.addWidget(QLabel(right_freq), 2, 1)
        right_grid.addWidget(QLabel(right_desc), 3, 1)
        right_vbox.addLayout(right_grid)

        right_table = UPartView()
        right_table.set_model(UPartToPandas(self.u_part_2).df)
        right_vbox.addWidget(right_table)

        hbox.addLayout(right_vbox)

        hbox2 = QHBoxLayout()
        self.no_button = QPushButton("No", self)
        self.no_button.clicked[bool].connect(self.no)
        hbox2.addWidget(self.no_button)
        self.yes_button = QPushButton("Yes", self)
        self.yes_button.clicked[bool].connect(self.yes)
        hbox2.addWidget(self.yes_button)

        vbox.addLayout(hbox)
        vbox.addLayout(hbox2)
        self.setLayout(vbox)

    def yes(self):
        self.return_bool = True
        self.accept()

    def no(self):
        self.return_bool = False
        self.accept()

    def exec_(self):
        super().exec_()
        self.close()
        return self.return_bool


class UPartModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        if not self._data[0][0]:
            return 0
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role == Qt.DisplayRole:
            value = self._data[index.row()][index.column()]
            return str(value)
        else:
            return None

    def headerData(self, col, orientation, role):

        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]

        return None

    def flags(self, index):

        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class UPartView(QTableView):

    def __init__(self, parent=None):
        super().__init__()

    def set_model(self, data):

        self.setModel(UPartModel(data))
        self.resizeColumnsToContents()


class UPartToPandas:

    def __init__(self, u_part):

        labels = ["Stop", "Stop Name", "Tag", "Time"]

        compiled = []

        for x, stop in enumerate(u_part.half_stops):
            compiled.append([str(x+1), stop.stop_name, stop.tag, str(stop.time)])

        self.df = pd.DataFrame(compiled, columns=labels)


class AddressPicker(QDialog):

    def __init__(self, plate_num, name, addresses, parent=None):
        super(AddressPicker, self).__init__(parent)

        self.setWindowTitle(plate_num)
        self.name = name
        self.addresses = addresses

        self.button_list = []
        self.accept_button = None

        self.initUI()

    def initUI(self):

        vbox = QVBoxLayout()
        grid = QGridLayout()

        header = QLabel("Pick Correct Address for: " + self.name)

        for x, address in enumerate(self.addresses):
            label = QLabel(address[1])
            rb = QRadioButton()
            self.button_list.append(rb)
            grid.addWidget(label, x, 0)
            grid.addWidget(rb, x, 1)

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        vbox.addWidget(header)
        vbox.addLayout(grid)
        vbox.addWidget(self.accept_button)

        self.setLayout(vbox)

    def exec_(self):
        super().exec_()

        return_index = 0

        for x, button in enumerate(self.button_list):
            if button.isChecked():
                return_index = x

        self.close()

        return self.addresses[return_index]