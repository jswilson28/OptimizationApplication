
from openpyxl import load_workbook


class PythonCostModel:

    def __init__(self, input_file):

        self.wb = load_workbook(filename=input_file, read_only=True)

        # hourly wages
        ws = self.wb['Inputs']

        self.full_time_apwu_ps_07_hourly = float(ws['C8'].value)
        self.full_time_apwu_ps_08_hourly = float(ws['C9'].value)
        self.postal_support_employee_apwu_7_hourly = float(ws['C10'].value)
        self.postal_support_employee_apwu_8_hourly = float(ws['C11'].value)

        self.full_time_apwu_ps_07_full_load = float(ws['D8'].value)
        self.full_time_apwu_ps_08_full_load = float(ws['D9'].value)
        self.postal_support_employee_apwu_7_full_load = float(ws['D10'].value)
        self.postal_support_employee_apwu_8_full_load = float(ws['D11'].value)

        self.full_time_apwu_ps_07_night_diff = float(ws['E8'].value)
        self.full_time_apwu_ps_08_night_diff = float(ws['E9'].value)
        self.postal_support_employee_apwu_7_night_diff = float(ws['E10'].value)
        self.postal_support_employee_apwu_8_night_diff = float(ws['E11'].value)

        # hours split
        self.ft = float(ws['C14'].value)
        self.pse = float(ws['C15'].value)
        self.night_diff = float(ws['C16'].value)

        # other inputs
        self.torts_per_mile = float(ws['C19'].value)
        self.hcr_miles_hours_inf = float(ws['C20'].value)
        self.recruiting_per_employee = float(ws['C21'].value)
        self.vacation_sick_inf = float(ws['C22'].value)
        self.attrition = float(ws['C23'].value)
        self.ot_due_to_attrition = float(ws['C24'].value)
        self.ot_base_increase = float(ws['C25'].value)
        self.annual_ft_hours = float(ws['C26'].value)
        self.annual_pse_hours = float(ws['C27'].value)
        self.annual_manager_salary = float(ws['C28'].value)
        self.drivers_per_manager = int(ws['C29'].value)
        self.trailers_per_tractor = int(ws['C30'].value)
        self.annual_maint_per_trailer = None

        # vehicle acquisition costs
        self.cargo_van_acq_cost = float(ws['C33'].value)
        self.tractor_sa_acq_cost = float(ws['C34'].value)
        self.tractor_ta_acq_cost = float(ws['C35'].value)
        self.trailer_acq_cost = float(ws['C36'].value)

        self.cargo_van_yrs_dep = int(ws['D33'].value)
        self.tractor_sa_yrs_dep = int(ws['D34'].value)
        self.tractor_ta_yrs_dep = int(ws['D35'].value)

        self.cargo_van_1_yr_ebuy_monthly_cost = float(ws['C39'].value)
        self.tractor_sa_1_yr_ebuy_monthly_cost = float(ws['C40'].value)
        self.tractor_ta_1_yr_ebuy_monthly_cost = float(ws['C41'].value)
        self.trailer_1_yr_ebuy_monthly_cost = float(ws['C42'].value)

        self.cargo_van_annual_lease_cost = float(ws['D39'].value)
        self.tractor_sa_annual_lease_cost = float(ws['D40'].value)
        self.tractor_ta_annual_lease_cost = float(ws['D41'].value)
        self.trailer_annual_lease_cost = float(ws['D42'].value)

        self.cargo_van_mileage_max = float(ws['E39'].value)
        self.tractor_sa_mileage_max = float(ws['E40'].value)
        self.tractor_ta_mileage_max = float(ws['E41'].value)
        self.trailer_mileage_max = ws['E42'].value

        self.cargo_van_extra_mileage_charge = float(ws['F39'].value)
        self.tractor_sa_extra_mileage_charge = float(ws['F40'].value)
        self.tractor_ta_extra_mileage_charge = float(ws['F41'].value)
        self.trailer_extra_mileage_charge = ws['F42'].value

        # cost per mile
        self.regions = ['NE', 'EA', 'CM', 'WE', 'PA', 'SA', 'GL', 'HQ']

        self.cargo_van_cpm = []
        self.tractor_sa_cpm = []
        self.tractor_ta_cpm = []
        self.trailer_cost_by_region = []

        for y in range(3,11):
            self.cargo_van_cpm.append(float(ws.cell(row=46, column=y).value))
            self.tractor_sa_cpm.append(float(ws.cell(row=47, column=y).value))
            self.tractor_ta_cpm.append(float(ws.cell(row=48, column=y).value))
            self.trailer_cost_by_region.append(float(ws.cell(row=52, column=y).value))

        # fuel inputs
        self.cargo_van_annual_fuel = float(ws['C55'].value)
        self.tractor_sa_annual_fuel = float(ws['C56'].value)
        self.tractor_ta_annual_fuel = float(ws['C57'].value)

        self.van_cpm = None
        self.sa_cpm = None
        self.ta_cpm = None

        self.hcr_contract = None
        self.vehicles = None

        self.wtd_ft_wage_rate = None
        self.wtd_pse_wage_rate = None

    def process_contract(self, hcr_id, hcr_contract, hcr_vehicle_contract):

        self.hcr_contract = hcr_contract
        self.vehicles = hcr_vehicle_contract

        if not self.hcr_contract:
            print("No contract no. ", hcr_id, " found.")
            return

        if not self.vehicles:
            print("No vehicles for contract no. ", hcr_id, " found.")
            return

        # set wage rates
        self.wtd_ft_wage_rate_calc()
        self.wtd_pse_wage_rate_calc()

        # do the calculations for HCR Contract
        self.usps_equivalents()
        self.usps_labor_hours()
        self.usps_employees()
        self.usps_labor_cost()

        # do the calculations for the fleet cost
        self.usps_transportation_costs()
        self.usps_owned_fleet()
        self.usps_leased_fleet()

        self.hcr_contract.total_lease_cost_calc()

    def usps_equivalents(self):

        self.hcr_contract.usps_equivalents_calc(self.hcr_miles_hours_inf, self.vacation_sick_inf)

    def wtd_ft_wage_rate_calc(self):

        van_percent = self.vehicles.percent_vans
        tractor_percent = self.vehicles.percent_tractors

        if van_percent + tractor_percent == 0:
            van_percent = 0.5
            tractor_percent = 0.5

        self.wtd_ft_wage_rate = ((van_percent * self.full_time_apwu_ps_07_full_load) +
                                 (tractor_percent * self.full_time_apwu_ps_08_full_load))
        self.hcr_contract.wtd_ft_wage_rate = self.wtd_ft_wage_rate

    def wtd_pse_wage_rate_calc(self):

        van_percent = self.vehicles.percent_vans
        tractor_percent = self.vehicles.percent_tractors

        if van_percent + tractor_percent == 0:
            van_percent = 0.5
            tractor_percent = 0.5

        self.wtd_pse_wage_rate = ((van_percent * self.postal_support_employee_apwu_7_full_load) +
                                 (tractor_percent * self.postal_support_employee_apwu_8_full_load))
        self.hcr_contract.wtd_pse_wage_rate = self.wtd_pse_wage_rate

    def usps_labor_hours(self):

        self.hcr_contract.usps_labor_hours_calc(self.ft, self.pse, self.night_diff, self.ot_due_to_attrition)

    def usps_employees(self):

        self.hcr_contract.usps_employees_calc(self.annual_ft_hours, self.annual_pse_hours, self.drivers_per_manager)

    def usps_labor_cost(self):

        # FT Base Hours
        vans = self.vehicles.percent_vans
        tractors = self.vehicles.percent_tractors
        van_rate = self.full_time_apwu_ps_07_full_load
        tractor_rate = self.full_time_apwu_ps_08_full_load
        hours = self.annual_ft_hours
        self.hcr_contract.usps_ft_costs_calc(vans, tractors, van_rate, tractor_rate, hours)

        # FT OT Hours
        ot_van_rate = van_rate + (self.ot_base_increase * self.full_time_apwu_ps_07_hourly)
        ot_tractor_rate = tractor_rate + (self.ot_base_increase * self.full_time_apwu_ps_08_hourly)
        self.hcr_contract.usps_ft_ot_costs_calc(vans, tractors, ot_van_rate, ot_tractor_rate)

        # PSE Base Hours
        van_rate = self.postal_support_employee_apwu_7_full_load
        tractor_rate = self.postal_support_employee_apwu_8_full_load
        self.hcr_contract.usps_pse_costs_calc(vans, tractors, van_rate, tractor_rate)

        # PSE OT Hours
        ot_van_rate = van_rate + (self.ot_base_increase * self.postal_support_employee_apwu_7_hourly)
        ot_tractor_rate = tractor_rate + (self.ot_base_increase * self.postal_support_employee_apwu_8_hourly)
        self.hcr_contract.usps_pse_ot_costs_calc(vans, tractors, ot_van_rate, ot_tractor_rate)

        # Manager Costs
        manager_salary = self.annual_manager_salary
        self.hcr_contract.usps_management_costs_calc(manager_salary)

        # Night Differential
        night_diff = self.night_diff
        self.hcr_contract.usps_night_diff_calc(night_diff)

        # Recruitment costs
        rec_cost = self.recruiting_per_employee
        attrition = self.attrition
        self.hcr_contract.usps_attrition_costs_calc(rec_cost, attrition)

        # Total labor costs
        self.hcr_contract.usps_total_labor_cost_calc()

    def usps_transportation_costs(self):

        total_miles = self.hcr_contract.usps_annual_miles
        postalized_miles = self.hcr_contract.total_postalized_mileage

        self.hcr_contract.van_miles = self.vehicles.percent_vans * total_miles
        self.hcr_contract.tractor_sa_miles = self.vehicles.percent_tractors_sa * total_miles
        self.hcr_contract.tractor_ta_miles = self.vehicles.percent_tractors_ta * total_miles

        self.hcr_contract.van_miles_postalized = self.vehicles.percent_vans * postalized_miles
        self.hcr_contract.tractor_sa_miles_postalized = self.vehicles.percent_tractors_sa * postalized_miles
        self.hcr_contract.tractor_ta_miles_postalized = self.vehicles.percent_tractors_ta * postalized_miles

        self.hcr_contract.cargo_van_annual_fuel = self.cargo_van_annual_fuel
        self.hcr_contract.tractor_sa_annual_fuel = self.tractor_sa_annual_fuel
        self.hcr_contract.tractor_ta_annual_fuel = self.tractor_ta_annual_fuel

        self.hcr_contract.van_fuel_cost = self.vehicles.cargo_van_count * self.cargo_van_annual_fuel
        self.hcr_contract.tractor_sa_fuel_cost = self.vehicles.tractor_sa_count * self.tractor_sa_annual_fuel
        self.hcr_contract.tractor_ta_fuel_cost = self.vehicles.tractor_ta_count * self.tractor_ta_annual_fuel
        self.hcr_contract.total_fuel_cost = (self.hcr_contract.van_fuel_cost +
                                             self.hcr_contract.tractor_sa_fuel_cost +
                                             self.hcr_contract.tractor_ta_fuel_cost)

        try:
            region_index = self.regions.index(self.hcr_contract.area)
        except:
            print(self.hcr_contract.hcr_id)

        self.van_cpm = self.cargo_van_cpm[region_index]
        self.hcr_contract.van_cpm = self.van_cpm
        self.sa_cpm = self.tractor_sa_cpm[region_index]
        self.hcr_contract.tractor_sa_cpm = self.sa_cpm
        self.ta_cpm = self.tractor_ta_cpm[region_index]
        self.hcr_contract.tractor_ta_cpm = self.ta_cpm
        self.annual_maint_per_trailer = self.trailer_cost_by_region[region_index]

        self.hcr_contract.cargo_van_ops = self.van_cpm * self.hcr_contract.van_miles
        self.hcr_contract.cargo_van_ops_postalized = self.van_cpm * self.hcr_contract.van_miles_postalized
        self.hcr_contract.tractor_sa_ops = self.sa_cpm * self.hcr_contract.tractor_sa_miles
        self.hcr_contract.tractor_sa_ops_postalized = self.sa_cpm * self.hcr_contract.tractor_sa_miles_postalized
        self.hcr_contract.tractor_ta_ops = self.ta_cpm * self.hcr_contract.tractor_ta_miles
        self.hcr_contract.tractor_ta_ops_postalized = self.ta_cpm * self.hcr_contract.tractor_ta_miles_postalized

        self.hcr_contract.annual_maint_per_trailer = self.annual_maint_per_trailer
        self.hcr_contract.trailer_ops = self.vehicles.trailer_count * self.annual_maint_per_trailer

        self.hcr_contract.total_ops = (self.hcr_contract.cargo_van_ops +
                                       self.hcr_contract.tractor_sa_ops +
                                       self.hcr_contract.tractor_ta_ops +
                                       self.hcr_contract.trailer_ops)

        self.hcr_contract.total_ops_postalized = (self.hcr_contract.cargo_van_ops_postalized +
                                                  self.hcr_contract.tractor_sa_ops_postalized +
                                                  self.hcr_contract.tractor_ta_ops_postalized +
                                                  self.hcr_contract.trailer_ops)

        self.hcr_contract.torts_per_mile = self.torts_per_mile
        self.hcr_contract.torts = total_miles * self.torts_per_mile
        self.hcr_contract.torts_postalized = postalized_miles * self.torts_per_mile

        self.hcr_contract.total_transportation_cost_calc()

    def usps_owned_fleet(self):

        van_count = self.vehicles.cargo_van_count
        tractor_sa_count = self.vehicles.tractor_sa_count
        tractor_ta_count = self.vehicles.tractor_ta_count
        trailer_count = self.vehicles.trailer_count

        self.hcr_contract.van_count = van_count
        self.hcr_contract.tractor_sa_count = tractor_sa_count
        self.hcr_contract.tractor_ta_count = tractor_ta_count
        self.hcr_contract.trailer_count = trailer_count

        van_cost = self.cargo_van_acq_cost
        tractor_sa_cost = self.tractor_sa_acq_cost
        tractor_ta_cost =self.tractor_ta_acq_cost
        trailer_cost = self.trailer_acq_cost

        self.hcr_contract.cargo_van_acq_cost = self.cargo_van_acq_cost
        self.hcr_contract.tractor_sa_acq_cost = self.tractor_sa_acq_cost
        self.hcr_contract.tractor_ta_acq_cost = self.tractor_ta_acq_cost
        self.hcr_contract.trailer_acq_cost = self.trailer_acq_cost

        van_dep_cost = van_cost/self.cargo_van_yrs_dep
        self.hcr_contract.cargo_van_dep_life = self.cargo_van_yrs_dep
        self.hcr_contract.cargo_van_dep_cost = van_dep_cost
        tractor_sa_dep_cost = tractor_sa_cost/self.tractor_sa_yrs_dep
        self.hcr_contract.tractor_sa_dep_life = self.tractor_sa_yrs_dep
        self.hcr_contract.tractor_sa_dep_cost = tractor_sa_cost
        tractor_ta_dep_cost = tractor_ta_cost/self.tractor_ta_yrs_dep
        self.hcr_contract.tractor_ta_dep_life = self.tractor_ta_yrs_dep
        self.hcr_contract.tractor_ta_dep_cost = tractor_ta_dep_cost
        trailer_dep_cost = trailer_cost

        self.hcr_contract.total_acq_cost_wo_dep = ((van_cost * van_count) +
                                                   (tractor_sa_cost * tractor_sa_count) +
                                                   (tractor_ta_cost * tractor_ta_count) +
                                                   (trailer_cost * trailer_count))
        self.hcr_contract.total_acq_cost_w_dep = ((van_dep_cost * van_count) +
                                                  (tractor_sa_dep_cost * tractor_sa_count) +
                                                  (tractor_ta_dep_cost * tractor_ta_count) +
                                                  (trailer_dep_cost * trailer_count))

    def usps_leased_fleet(self):

        self.hcr_contract.van_lease_calc(self.cargo_van_annual_lease_cost, self.cargo_van_mileage_max,
                                         self.cargo_van_extra_mileage_charge)

        self.hcr_contract.tractor_sa_lease_calc(self.tractor_sa_annual_lease_cost, self.tractor_sa_mileage_max,
                                                self.tractor_sa_extra_mileage_charge)

        self.hcr_contract.tractor_ta_lease_calc(self.tractor_ta_annual_lease_cost, self.tractor_ta_mileage_max,
                                                self.tractor_ta_extra_mileage_charge)

        self.hcr_contract.trailer_lease_cost_calc(self.trailer_annual_lease_cost)

    def get_table_outputs(self):

        annual_contract_rate = self.hcr_contract.total_annual_rate

        labor = self.hcr_contract.total_labor_cost
        labor_postalized = self.hcr_contract.total_labor_cost_postalized
        transpo = self.hcr_contract.total_owned_transportation_cost
        transpo_postalized = self.hcr_contract.total_owned_transportation_cost_postalized
        no_dep = self.hcr_contract.total_acq_cost_wo_dep
        dep = self.hcr_contract.total_acq_cost_w_dep

        cost_w_dep = labor + transpo + dep
        cost_w_dep_postalized = labor_postalized + transpo_postalized + dep
        self.hcr_contract.total_cost_w_dep = cost_w_dep
        self.hcr_contract.total_cost_w_dep_postalized = cost_w_dep_postalized

        cost_wo_dep = labor + transpo + no_dep
        cost_wo_dep_postalized = labor_postalized + transpo_postalized + no_dep
        self.hcr_contract.total_cost_wo_dep = cost_wo_dep
        self.hcr_contract.total_cost_wo_dep_postalized = cost_wo_dep_postalized

        leased_transpo = self.hcr_contract.total_leased_transportation_cost
        leased_transpo_postalized = self.hcr_contract.total_leased_transportation_cost_postalized
        leased = self.hcr_contract.total_lease_cost
        leased_postalized = self.hcr_contract.total_lease_cost_postalized

        cost_w_lease = labor + leased_transpo + leased
        cost_w_lease_postalized = labor_postalized + leased_transpo_postalized + leased_postalized
        self.hcr_contract.total_cost_w_lease = cost_w_lease
        self.hcr_contract.total_cost_w_lease_postalized = cost_w_lease_postalized

        return annual_contract_rate, cost_w_dep, cost_wo_dep, cost_w_lease, cost_w_dep_postalized, \
               cost_wo_dep_postalized, cost_w_lease_postalized


class CostModelPrinter:

    def __init__(self, cost_model, outfile):

        if len(outfile) > 1:
            if outfile[-1] not in ("\\", "/"):
                outfile = outfile + "/"

        self.file_name = outfile + "CostModelOutputs.xlsx"
        self.cost_model = cost_model
        self.hcr_contract = cost_model.hcr_contract

        self.wb = load_workbook("OutputFormats.xlsx")
        self.save_workbook()

    def print_all_pages(self):

        self.write_summary_page()
        self.write_summary_with_dep_page()
        self.write_summary_wo_dep_page()
        self.write_summary_with_lease_page()

        self.save_workbook()

    def write_summary_page(self):

        ws = self.wb["HCR PVS Summary"]
        hcr_contract = self.hcr_contract

        origin_terminus = self.hcr_contract.origin_terminus
        origin_string = "HCR Contract being compared: " + origin_terminus
        ws["A1"].value = origin_string
        ws["C1"].value = hcr_contract.hcr_id

        # current HCR hours...
        total_hours = hcr_contract.total_annual_hours
        ws["C3"].value = total_hours
        annual_miles = hcr_contract.total_annual_miles
        ws["C4"].value = annual_miles
        annual_rate = hcr_contract.total_annual_rate
        ws["C5"].value = annual_rate
        cargo_van = hcr_contract.van_count
        ws["C6"].value = cargo_van
        tractor_sa = hcr_contract.tractor_sa_count
        ws["C7"].value = tractor_sa
        tractor_ta = hcr_contract.tractor_ta_count
        ws["C8"] = tractor_ta
        trailer = hcr_contract.trailer_count
        ws["C9"].value = trailer

        # PVS labor
        wtd_ft_wage_rate = self.cost_model.wtd_ft_wage_rate
        ws["C11"].value = wtd_ft_wage_rate
        wtd_pse_wage_rate = self.cost_model.wtd_pse_wage_rate
        ws["C12"].value = wtd_pse_wage_rate
        total_driving_hours = hcr_contract.usps_paid_hours
        ws["C13"].value = total_driving_hours
        non_ot_labor_costs = hcr_contract.ft_costs + hcr_contract.pse_costs
        ws["C14"].value = non_ot_labor_costs
        manager_costs = hcr_contract.manager_costs
        ws["C15"].value = manager_costs
        ot_costs = hcr_contract.ft_ot_costs + hcr_contract.pse_ot_costs
        ws["C16"].value = ot_costs
        recruit_costs = hcr_contract.year_one_recruitment_cost + hcr_contract.recurring_recruitment_cost
        ws["C17"].value = recruit_costs
        total_labor_costs = hcr_contract.total_labor_cost
        ws["C18"].value = total_labor_costs

        # PVS Transportation
        total_pvs_miles = hcr_contract.usps_annual_miles
        ws["C20"].value = total_pvs_miles
        van_cpm = self.cost_model.van_cpm
        ws["C21"].value = van_cpm
        sa_cpm = self.cost_model.sa_cpm
        ws["C22"].value = sa_cpm
        ta_cpm = self.cost_model.ta_cpm
        ws["C23"].value = ta_cpm
        torts = self.cost_model.torts_per_mile
        ws["C24"].value = torts
        trailer_maint = self.cost_model.annual_maint_per_trailer
        ws["C25"].value = trailer_maint
        van_miles = hcr_contract.van_miles
        ws["C26"].value = van_miles
        sa_miles = hcr_contract.tractor_sa_miles
        ws["C27"].value = sa_miles
        ta_miles = hcr_contract.tractor_ta_miles
        ws["C28"].value = ta_miles
        van_ops = hcr_contract.cargo_van_ops
        ws["C29"].value = van_ops
        sa_ops = hcr_contract.tractor_sa_ops
        ws["C30"].value = sa_ops
        ta_ops = hcr_contract.tractor_ta_ops
        ws["C31"].value = ta_ops
        total_trailer_maint = hcr_contract.trailer_ops
        ws["C32"].value = total_trailer_maint
        total_torts = hcr_contract.torts
        ws["C33"].value = total_torts
        total_tolls = hcr_contract.tolls
        ws["C34"].value = total_tolls
        van_fuel = hcr_contract.van_fuel_cost
        ws["C35"].value = van_fuel
        sa_fuel = hcr_contract.tractor_sa_fuel_cost
        ws["C36"].value = sa_fuel
        ta_fuel = hcr_contract.tractor_ta_fuel_cost
        ws["C37"].value = ta_fuel
        total_pvs_transpo_cost = hcr_contract.total_owned_transportation_cost
        ws["C38"].value = total_pvs_transpo_cost

        # Estimate of PVS Fleet Cost
        van_life = self.cost_model.cargo_van_yrs_dep
        ws["C40"].value = van_life
        sa_life = self.cost_model.tractor_sa_yrs_dep
        ws["C41"].value = sa_life
        ta_life = self.cost_model.tractor_ta_yrs_dep
        ws["C42"].value = ta_life
        van_cost = self.cost_model.cargo_van_acq_cost
        ws["C43"].value = van_cost
        sa_cost = self.cost_model.tractor_sa_acq_cost
        ws["C44"].value = sa_cost
        ta_cost = self.cost_model.tractor_ta_acq_cost
        ws["C45"].value = ta_cost
        yr_one_van = van_cost / van_life
        ws["C46"].value = yr_one_van
        yr_one_sa = sa_cost / sa_life
        ws["C47"].value = yr_one_sa
        yr_one_ta = ta_cost / ta_life
        ws["C48"].value = yr_one_ta
        trailer_cost = self.cost_model.trailer_acq_cost
        ws["C49"].value = trailer_cost
        dep_cost = hcr_contract.total_acq_cost_w_dep
        ws["C50"].value = dep_cost
        full_cost = hcr_contract.total_acq_cost_wo_dep
        ws["C51"].value = full_cost

        # Estimate of Leased Vehicle Cost
        van_cost = self.cost_model.cargo_van_annual_lease_cost
        ws["C53"].value = van_cost
        sa_cost = self.cost_model.tractor_sa_annual_lease_cost
        ws["C54"].value = sa_cost
        ta_cost = self.cost_model.tractor_ta_annual_lease_cost
        ws["C55"].value = ta_cost
        trailer_cost = self.cost_model.trailer_annual_lease_cost
        ws["C56"].value = trailer_cost
        base_van = hcr_contract.van_base_lease_cost
        ws["C57"].value = base_van
        base_sa = hcr_contract.tractor_sa_base_lease_cost
        ws["C58"].value = base_sa
        base_ta = hcr_contract.tractor_ta_base_lease_cost
        ws["C59"].value = base_ta
        base_trailer = hcr_contract.trailer_base_lease_cost
        ws["C60"].value = base_trailer
        mileage_extra = hcr_contract.total_extra_lease_cost
        ws["C61"].value = mileage_extra
        total = hcr_contract.total_lease_cost
        ws["C62"].value = total

    def write_summary_with_dep_page(self):

        ws = self.wb["Summary with Depreciation"]
        hcr_contract = self.hcr_contract

        origin_terminus = self.hcr_contract.origin_terminus
        origin_string = "HCR Contract being compared: " + origin_terminus
        ws["A1"].value = origin_string
        ws["C1"].value = hcr_contract.hcr_id

        # current HCR hours...
        total_hours = hcr_contract.total_annual_hours
        ws["C3"].value = total_hours
        annual_miles = hcr_contract.total_annual_miles
        ws["C4"].value = annual_miles
        annual_rate = hcr_contract.total_annual_rate
        ws["C5"].value = annual_rate
        cargo_van = hcr_contract.van_count
        ws["C6"].value = cargo_van
        tractor_sa = hcr_contract.tractor_sa_count
        ws["C7"].value = tractor_sa
        tractor_ta = hcr_contract.tractor_ta_count
        ws["C8"] = tractor_ta
        trailer = hcr_contract.trailer_count
        ws["C9"].value = trailer

        # PVS labor
        wtd_ft_wage_rate = self.cost_model.wtd_ft_wage_rate
        ws["C11"].value = wtd_ft_wage_rate
        wtd_pse_wage_rate = self.cost_model.wtd_pse_wage_rate
        ws["C12"].value = wtd_pse_wage_rate
        total_driving_hours = hcr_contract.usps_paid_hours
        ws["C13"].value = total_driving_hours
        non_ot_labor_costs = hcr_contract.ft_costs + hcr_contract.pse_costs
        ws["C14"].value = non_ot_labor_costs
        manager_costs = hcr_contract.manager_costs
        ws["C15"].value = manager_costs
        ot_costs = hcr_contract.ft_ot_costs + hcr_contract.pse_ot_costs
        ws["C16"].value = ot_costs
        recruit_costs = hcr_contract.year_one_recruitment_cost + hcr_contract.recurring_recruitment_cost
        ws["C17"].value = recruit_costs
        total_labor_costs = hcr_contract.total_labor_cost
        ws["C18"].value = total_labor_costs

        # PVS Transportation
        total_pvs_miles = hcr_contract.usps_annual_miles
        ws["C20"].value = total_pvs_miles
        van_cpm = self.cost_model.van_cpm
        ws["C21"].value = van_cpm
        sa_cpm = self.cost_model.sa_cpm
        ws["C22"].value = sa_cpm
        ta_cpm = self.cost_model.ta_cpm
        ws["C23"].value = ta_cpm
        torts = self.cost_model.torts_per_mile
        ws["C24"].value = torts
        trailer_maint = self.cost_model.annual_maint_per_trailer
        ws["C25"].value = trailer_maint
        van_miles = hcr_contract.van_miles
        ws["C26"].value = van_miles
        sa_miles = hcr_contract.tractor_sa_miles
        ws["C27"].value = sa_miles
        ta_miles = hcr_contract.tractor_ta_miles
        ws["C28"].value = ta_miles
        van_ops = hcr_contract.cargo_van_ops
        ws["C29"].value = van_ops
        sa_ops = hcr_contract.tractor_sa_ops
        ws["C30"].value = sa_ops
        ta_ops = hcr_contract.tractor_ta_ops
        ws["C31"].value = ta_ops
        total_trailer_maint = hcr_contract.trailer_ops
        ws["C32"].value = total_trailer_maint
        total_torts = hcr_contract.torts
        ws["C33"].value = total_torts
        total_tolls = hcr_contract.tolls
        ws["C34"].value = total_tolls
        van_fuel = hcr_contract.van_fuel_cost
        ws["C35"].value = van_fuel
        sa_fuel = hcr_contract.tractor_sa_fuel_cost
        ws["C36"].value = sa_fuel
        ta_fuel = hcr_contract.tractor_ta_fuel_cost
        ws["C37"].value = ta_fuel
        total_pvs_transpo_cost = hcr_contract.total_owned_transportation_cost
        ws["C38"].value = total_pvs_transpo_cost

        # Estimate of PVS Fleet Cost
        van_life = self.cost_model.cargo_van_yrs_dep
        ws["C40"].value = van_life
        sa_life = self.cost_model.tractor_sa_yrs_dep
        ws["C41"].value = sa_life
        ta_life = self.cost_model.tractor_ta_yrs_dep
        ws["C42"].value = ta_life
        van_cost = self.cost_model.cargo_van_acq_cost
        ws["C43"].value = van_cost
        sa_cost = self.cost_model.tractor_sa_acq_cost
        ws["C44"].value = sa_cost
        ta_cost = self.cost_model.tractor_ta_acq_cost
        ws["C45"].value = ta_cost
        yr_one_van = van_cost / van_life
        ws["C46"].value = yr_one_van
        yr_one_sa = sa_cost / sa_life
        ws["C47"].value = yr_one_sa
        yr_one_ta = ta_cost / ta_life
        ws["C48"].value = yr_one_ta
        trailer_cost = self.cost_model.trailer_acq_cost
        ws["C49"].value = trailer_cost
        dep_cost = hcr_contract.total_acq_cost_w_dep
        ws["C50"].value = dep_cost

    def write_summary_wo_dep_page(self):

        ws = self.wb["Summary with No Depreciation"]
        hcr_contract = self.hcr_contract

        origin_terminus = self.hcr_contract.origin_terminus
        origin_string = "HCR Contract being compared: " + origin_terminus
        ws["A1"].value = origin_string
        ws["C1"].value = hcr_contract.hcr_id

        # current HCR hours...
        total_hours = hcr_contract.total_annual_hours
        ws["C3"].value = total_hours
        annual_miles = hcr_contract.total_annual_miles
        ws["C4"].value = annual_miles
        annual_rate = hcr_contract.total_annual_rate
        ws["C5"].value = annual_rate
        cargo_van = hcr_contract.van_count
        ws["C6"].value = cargo_van
        tractor_sa = hcr_contract.tractor_sa_count
        ws["C7"].value = tractor_sa
        tractor_ta = hcr_contract.tractor_ta_count
        ws["C8"] = tractor_ta
        trailer = hcr_contract.trailer_count
        ws["C9"].value = trailer

        # PVS labor
        wtd_ft_wage_rate = self.cost_model.wtd_ft_wage_rate
        ws["C11"].value = wtd_ft_wage_rate
        wtd_pse_wage_rate = self.cost_model.wtd_pse_wage_rate
        ws["C12"].value = wtd_pse_wage_rate
        total_driving_hours = hcr_contract.usps_paid_hours
        ws["C13"].value = total_driving_hours
        non_ot_labor_costs = hcr_contract.ft_costs + hcr_contract.pse_costs
        ws["C14"].value = non_ot_labor_costs
        manager_costs = hcr_contract.manager_costs
        ws["C15"].value = manager_costs
        ot_costs = hcr_contract.ft_ot_costs + hcr_contract.pse_ot_costs
        ws["C16"].value = ot_costs
        recruit_costs = hcr_contract.year_one_recruitment_cost + hcr_contract.recurring_recruitment_cost
        ws["C17"].value = recruit_costs
        total_labor_costs = hcr_contract.total_labor_cost
        ws["C18"].value = total_labor_costs

        # PVS Transportation
        total_pvs_miles = hcr_contract.usps_annual_miles
        ws["C20"].value = total_pvs_miles
        van_cpm = self.cost_model.van_cpm
        ws["C21"].value = van_cpm
        sa_cpm = self.cost_model.sa_cpm
        ws["C22"].value = sa_cpm
        ta_cpm = self.cost_model.ta_cpm
        ws["C23"].value = ta_cpm
        torts = self.cost_model.torts_per_mile
        ws["C24"].value = torts
        trailer_maint = self.cost_model.annual_maint_per_trailer
        ws["C25"].value = trailer_maint
        van_miles = hcr_contract.van_miles
        ws["C26"].value = van_miles
        sa_miles = hcr_contract.tractor_sa_miles
        ws["C27"].value = sa_miles
        ta_miles = hcr_contract.tractor_ta_miles
        ws["C28"].value = ta_miles
        van_ops = hcr_contract.cargo_van_ops
        ws["C29"].value = van_ops
        sa_ops = hcr_contract.tractor_sa_ops
        ws["C30"].value = sa_ops
        ta_ops = hcr_contract.tractor_ta_ops
        ws["C31"].value = ta_ops
        total_trailer_maint = hcr_contract.trailer_ops
        ws["C32"].value = total_trailer_maint
        total_torts = hcr_contract.torts
        ws["C33"].value = total_torts
        total_tolls = hcr_contract.tolls
        ws["C34"].value = total_tolls
        van_fuel = hcr_contract.van_fuel_cost
        ws["C35"].value = van_fuel
        sa_fuel = hcr_contract.tractor_sa_fuel_cost
        ws["C36"].value = sa_fuel
        ta_fuel = hcr_contract.tractor_ta_fuel_cost
        ws["C37"].value = ta_fuel
        total_pvs_transpo_cost = hcr_contract.total_owned_transportation_cost
        ws["C38"].value = total_pvs_transpo_cost

        # Estimate of PVS fleet costs
        van_cost = self.cost_model.cargo_van_acq_cost
        ws["C40"].value = van_cost
        sa_cost = self.cost_model.tractor_sa_acq_cost
        ws["C41"].value = sa_cost
        ta_cost = self.cost_model.tractor_ta_acq_cost
        ws["C42"].value = ta_cost
        trailer_cost = self.cost_model.trailer_acq_cost
        ws["C43"].value = trailer_cost
        full_cost = hcr_contract.total_acq_cost_wo_dep
        ws["C44"].value = full_cost

    def write_summary_with_lease_page(self):

        ws = self.wb["Summary with Leased Fleet"]
        hcr_contract = self.hcr_contract

        origin_terminus = self.hcr_contract.origin_terminus
        origin_string = "HCR Contract being compared: " + origin_terminus
        ws["A1"].value = origin_string
        ws["C1"].value = hcr_contract.hcr_id

        # current HCR hours...
        total_hours = hcr_contract.total_annual_hours
        ws["C3"].value = total_hours
        annual_miles = hcr_contract.total_annual_miles
        ws["C4"].value = annual_miles
        annual_rate = hcr_contract.total_annual_rate
        ws["C5"].value = annual_rate
        cargo_van = hcr_contract.van_count
        ws["C6"].value = cargo_van
        tractor_sa = hcr_contract.tractor_sa_count
        ws["C7"].value = tractor_sa
        tractor_ta = hcr_contract.tractor_ta_count
        ws["C8"] = tractor_ta
        trailer = hcr_contract.trailer_count
        ws["C9"].value = trailer

        # PVS labor
        wtd_ft_wage_rate = self.cost_model.wtd_ft_wage_rate
        ws["C11"].value = wtd_ft_wage_rate
        wtd_pse_wage_rate = self.cost_model.wtd_pse_wage_rate
        ws["C12"].value = wtd_pse_wage_rate
        total_driving_hours = hcr_contract.usps_paid_hours
        ws["C13"].value = total_driving_hours
        non_ot_labor_costs = hcr_contract.ft_costs + hcr_contract.pse_costs
        ws["C14"].value = non_ot_labor_costs
        manager_costs = hcr_contract.manager_costs
        ws["C15"].value = manager_costs
        ot_costs = hcr_contract.ft_ot_costs + hcr_contract.pse_ot_costs
        ws["C16"].value = ot_costs
        recruit_costs = hcr_contract.year_one_recruitment_cost + hcr_contract.recurring_recruitment_cost
        ws["C17"].value = recruit_costs
        total_labor_costs = hcr_contract.total_labor_cost
        ws["C18"].value = total_labor_costs

        # PVS Transportation
        total_pvs_miles = hcr_contract.usps_annual_miles
        ws["C20"].value = total_pvs_miles
        torts = self.cost_model.torts_per_mile
        ws["C21"].value = torts
        van_miles = hcr_contract.van_miles
        ws["C22"].value = van_miles
        sa_miles = hcr_contract.tractor_sa_miles
        ws["C23"].value = sa_miles
        ta_miles = hcr_contract.tractor_ta_miles
        ws["C24"].value = ta_miles
        total_torts = hcr_contract.torts
        ws["C25"].value = total_torts
        total_tolls = hcr_contract.tolls
        ws["C26"].value = total_tolls
        van_fuel = hcr_contract.van_fuel_cost
        ws["C27"].value = van_fuel
        sa_fuel = hcr_contract.tractor_sa_fuel_cost
        ws["C28"].value = sa_fuel
        ta_fuel = hcr_contract.tractor_ta_fuel_cost
        ws["C29"].value = ta_fuel
        total_lease_cost = total_torts + total_tolls + van_fuel + sa_fuel + ta_fuel
        ws["C30"].value = total_lease_cost

        # Estimate of Leased Vehicle Cost
        van_cost = self.cost_model.cargo_van_annual_lease_cost
        ws["C32"].value = van_cost
        sa_cost = self.cost_model.tractor_sa_annual_lease_cost
        ws["C33"].value = sa_cost
        ta_cost = self.cost_model.tractor_ta_annual_lease_cost
        ws["C34"].value = ta_cost
        trailer_cost = self.cost_model.trailer_annual_lease_cost
        ws["C35"].value = trailer_cost
        base_van = hcr_contract.van_base_lease_cost
        ws["C36"].value = base_van
        base_sa = hcr_contract.tractor_sa_base_lease_cost
        ws["C37"].value = base_sa
        base_ta = hcr_contract.tractor_ta_base_lease_cost
        ws["C38"].value = base_ta
        base_trailer = hcr_contract.trailer_base_lease_cost
        ws["C39"].value = base_trailer
        mileage_extra = hcr_contract.total_extra_lease_cost
        ws["C40"].value = mileage_extra
        total = hcr_contract.total_lease_cost
        ws["C41"].value = total

    def save_workbook(self):

        try:
            self.wb.save(self.file_name)
            print("saved!")
        except:
            print("failed to save cost model outputs")