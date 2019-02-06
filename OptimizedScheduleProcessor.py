
import pandas as pd
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant, QSortFilterProxyModel
from PyQt5.QtWidgets import QTableView
from openpyxl import load_workbook
from ScheduleClasses import Stop, Schedule
import operator
import copy
from ScheduleCompilation import get_a_day_string
from GeneralMethods import today


def day_name_from_day_num(day_num):

    day_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    return day_list[day_num]


class Process:

    def __init__(self, file_name, lb, ip):

        self.ip = ip
        self.lb = lb
        self.file_name = file_name

        if "11-ton" in file_name:
            self.vehicle_type = "11-Ton"
        elif "Single" in file_name:
            self.vehicle_type = "Single"
        else:
            self.vehicle_type = "Unknown"

        self.wb = load_workbook(file_name)
        self.output_wb = None
        self.pvs_name = self.wb["Summary"]["B1"].value
        self.pdc_name = self.wb["Summary"]["B2"].value
        self.sun_schedules = self.wb["Summary"]["B3"].value in ("True", True, "TRUE")

        # first read in all stops from the "trips" tab
        self.round_trips = []
        self.read_in_round_trips()
        # then determine which days have schedules
        self.days = []
        self.day_string = ""
        self.set_days()
        # then read in optimal schedules from "solution" tabs
        self.optimized_schedules = []
        self.read_in_optimized_schedules()
        # get new frequency codes from lookup book
        self.day_codes = self.lb.get_day_codes(self.day_string, self.sun_schedules)
        self.schedules = []
        self.generate_schedules()
        self.wb.close()

    def read_in_round_trips(self):

        ws = self.wb["Trips"]

        max_row = ws.max_row
        x = 2

        round_trips = []

        while x <= max_row:
            orig_schedule_name = ws["A" + str(x)].value
            optimizer_trip_number = ws["B" + str(x)].value
            vehicle_type = ws["J" + str(x)].value
            lunch_indicator = ws["I" + str(x)].value
            stops = []
            while ws["B" + str(x)].value == optimizer_trip_number:
                stop_name = ws["D" + str(x)].value
                arrive_time = ws["E" + str(x)].value
                depart_time = ws["F" + str(x)].value
                arrive_cminute = ws["G" + str(x)].value
                depart_cminute = ws["H" + str(x)].value
                stops.append(Stop(stop_name=stop_name, arrive_time=arrive_time, depart_time=depart_time,
                                  arrive_cminute=arrive_cminute, depart_cminute=depart_cminute,
                                  original_schedule_name=orig_schedule_name))
                x += 1
            round_trips.append([optimizer_trip_number, orig_schedule_name, vehicle_type, lunch_indicator, stops])
        self.round_trips = round_trips

    def set_days(self):

        worksheets = self.wb.get_sheet_names()

        mon = "Solution Monday" in worksheets
        tue = "Solution Tuesday" in worksheets
        wed = "Solution Wednesday" in worksheets
        thu = "Solution Thursday" in worksheets
        fri = "Solution Friday" in worksheets
        sat = "Solution Saturday" in worksheets
        sun = "Solution Sunday" in worksheets

        self.days = [mon, tue, wed, thu, fri, sat, sun]
        for day in self.days:
            if day:
                self.day_string += "1"
            else:
                self.day_string += "0"

    def read_in_optimized_schedules(self):

        day_schedules = []

        for day_num, value in enumerate(self.days):
            schedule_rows = []
            if value:
                day_name = day_name_from_day_num(day_num)
                ws = self.wb["Solution " + day_name]
                x = 7
                max_row = ws.max_row
                while x <= max_row:
                    sched_num = ws["B" + str(x)].value
                    if sched_num in ("", None, " "):
                        break
                    trip_num = ws["C" + str(x)].value
                    trip = next(x for x in self.round_trips if x[0] == trip_num)
                    start_minute = trip[4][0].arrive_cminute
                    end_minute = trip[4][-1].depart_cminute
                    ws["F" + str(x)].value = start_minute
                    ws["G" + str(x)].value = end_minute
                    ws["H" + str(x)].value = end_minute - start_minute
                    break_in = ws["D" + str(x)].value
                    break_after = ws["E" + str(x)].value
                    x += 1
                    schedule_rows.append([day_num, sched_num, trip_num, break_in, break_after, start_minute, trip])
                self.wb.save(self.file_name)
                day_schedules.append(sorted(schedule_rows, key=operator.itemgetter(1, 5)))

        for day in day_schedules:
            sched_nums = list(set([x[1] for x in day]))
            for sched_num in sched_nums:
                compiled = [day[0][0], sched_num, [y[2:] for y in day if y[1] == sched_num]]
                self.optimized_schedules.append(compiled)

    def generate_schedules(self):

        for x, schedule in enumerate(self.optimized_schedules):
            sched_num = schedule[1]
            stops = []
            vehicle_type = schedule[2][0][4][2]
            num_trips = len(schedule[2])

            for trip in schedule[2]:
                break_after = trip[2] in (1, "1")
                for stop in trip[4][4]:
                    stops.append(copy.deepcopy(stop))
                if break_after:
                    stops[-1].lunch_after = True

            freq_code = self.day_codes[schedule[0]]
            bin_string = self.lb.get_new_bin_string(freq_code)

            new_schedule = Schedule(source="OPT", stops=stops, schedule_num=sched_num, freq_code=freq_code,
                                    pvs_name=self.pvs_name, pvs_pdc=self.pdc_name, bin_string=bin_string,
                                    num_trips=num_trips, mileage=0, read_in_index=(x+1), vehicle_type=vehicle_type,
                                    vehicle_category=vehicle_type)
            new_schedule.set_days_of_week()
            if new_schedule.adjust:
                new_schedule.adjust_frequency_code()
            new_schedule.clean_optimized_schedule(self.ip)
            new_schedule.set_cminutes(self.ip)
            self.schedules.append(new_schedule)

    def print_all(self):

        self.print_staffing_input()
        self.print_vehicles_input()
        self.print_output_schedules()
        self.wb.save(self.file_name)
        self.wb.close()

    def print_staffing_input(self):

        staffing_wb = load_workbook("Optimization Formats/Staffing_Formats.xlsx")
        new_file_name = "StaffingInputData_" + self.pdc_name + "_" + self.vehicle_type + "_" + today() + ".xlsx"
        staffing_wb.save(new_file_name)

        ws = staffing_wb["Data"]
        ws.sheet_view.showGridLines = False
        ws["A1"].value = "DOW"
        ws["B1"].value = "Schedule Number"
        ws["C1"].value = "Start"
        ws["D1"].value = "Stop"
        ws["E1"].value = "Duration"
        ws["F1"].value = "Lunch Minutes"
        ws["G1"].value = "Lunch start"
        ws["H1"].value = "Lunch End"

        x = 2

        for day in range(1, 8):
            for schedule in [y for y in self.schedules if y.bin_string[day-1] == "1"]:
                schedule.staffing_print(ws, x, day)
                x += 1

        staffing_wb.save(new_file_name)

    def print_vehicles_input(self):

        # reset days to account for shifted schedules
        self.days, _ = get_a_day_string(self.schedules)

        vehicle_wb = load_workbook("Optimization Formats/Vehicle_Formats.xlsx")
        new_file_name = "VehicleOptidata_" + self.pdc_name + "_" + self.vehicle_type + "_" + today() + ".xlsx"
        vehicle_wb.save(new_file_name)

        for day_num, day in enumerate(self.days):
            day_name = day_name_from_day_num(day_num)
            day_string = "OptiData " + day_name
            solution_day_string = "Solution " + day_name
            ws = vehicle_wb[day_string]

            if day:
                row = self.print_a_vehicle_day_sheet(ws, day_num)
                vehicle_wb.create_named_range("TRIPS" + str(day_num + 1), ws, "$B$7:$G$" + str(row))
            else:
                vehicle_wb.remove_sheet(vehicle_wb[day_string])
                vehicle_wb.remove_sheet(vehicle_wb[solution_day_string])

        vehicle_wb.save(new_file_name)
        vehicle_wb.close()

    def print_a_vehicle_day_sheet(self, ws, day_num):

        day_start_min = ((1 + day_num) * 1440) - 240
        day_end_min = day_start_min + 2160

        ws["B7"].value = 0
        ws["C7"].value = 0
        ws["D7"].value = 0
        ws["E7"].value = day_start_min
        ws["F7"].value = day_start_min
        ws["G7"].value = 0

        ws["B8"].value = 0
        ws["C8"].value = 999
        ws["D8"].value = 0
        ws["E8"].value = day_end_min
        ws["F8"].value = day_end_min
        ws["G8"].value = 0

        to_print = []

        for schedule in self.schedules:
            if schedule.bin_string[day_num] == "1":
                to_print.append(schedule)

        x = 9
        for schedule in to_print:
            schedule.vehicle_print(ws, x, day_num)
            x += 1

        return (x-1)

    def print_output_schedules(self):

        ws = self.wb.create_sheet(title="Schedule Summaries")
        ws.sheet_view.showGridLines = False
        row = 1
        ws["A" + str(row)].value = "Short Name"
        ws["B" + str(row)].value = "Schedule Name"
        ws["C" + str(row)].value = "Frequency Code"
        ws["D" + str(row)].value = "Vehicle Type"
        ws["E" + str(row)].value = "Num. Stops"
        ws["F" + str(row)].value = "Schedule Start"
        ws["G" + str(row)].value = "Schedule End"
        ws["H" + str(row)].value = "Duration (Minutes)"
        ws["I" + str(row)].value = "Optimizer Schedule Num"
        row += 1

        for schedule in self.schedules:
            schedule.excel_print_summary(ws, row)
            row += 1

        ws = self.wb.create_sheet(title="Schedule Stops")
        ws.sheet_view.showGridLines = False
        row = 1
        ws["A" + str(row)].value = "Schedule Name"
        ws["B" + str(row)].value = "Stop Number"
        ws["C" + str(row)].value = "Service Point"
        ws["D" + str(row)].value = "Arrive Time"
        ws["E" + str(row)].value = "Depart Time"
        ws["F" + str(row)].value = "Original Schedule"

        row += 1
        for schedule in self.schedules:
            schedule.excel_print_stops(ws, row)
            row += len(schedule.stops)

        ws = self.wb["Summary"]
        day_counts = [0, 0, 0, 0, 0, 0, 0]
        for schedule in self.schedules:
            for x in range(0, 7):
                if schedule.bin_string[x] in (1, "1"):
                    day_counts[x] += 1

        for x, day in enumerate(day_counts):
            ws.cell(row=(x+7), column=3).value = day


class OptimizedScheduleModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self._headers = data.columns

    def rowCount(self, parent=None):
        if not self._data[0][0]:
            return 0
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role == Qt.DisplayRole:
            value = self._data[index.row()][index.column()]
            return str(value)
        else:
            return None

    def headerData(self, col, orientation, role):

        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._headers[col]

        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class OptimizedScheduleView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, data):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(True)
        self.sorter.setSortRole(Qt.EditRole)
        self.sorter.setSourceModel(OptimizedScheduleModel(data))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)


class OptimizedSchedulesToPandas:

    def __init__(self, schedules):

        labels = ["Schedule Number", "Freq Code", "Start Time", "End Time", "Duration", "Trips", "Stops",
                  "Postal Compliant"]

        compiled = []
        for schedule in schedules:
            schedule_number = schedule.schedule_name
            freq_code = schedule.freq_code
            start_time = schedule.stops[0].arrive_time
            end_time = schedule.stops[-1].depart_time
            duration = schedule.raw_duration()
            trips = schedule.num_trips
            stops = len(schedule.stops)
            postalized = schedule.is_postalized
            compiled.append([schedule_number, freq_code, start_time, end_time, duration, trips, stops, postalized])

        self.df = pd.DataFrame(compiled, columns=labels)


class OneOptScheduleModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self._headers = data.columns

    def rowCount(self, parent=None):
        if not self._data[0][0]:
            return 0
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role == Qt.DisplayRole:
            value = self._data[index.row()][index.column()]
            return str(value)
        else:
            return None

    def headerData(self, col, orientation, role):

        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._headers[col]

        return QVariant()


class OneOptScheduleView(QTableView):

    def __init__(self, parent=None):
        super().__init__()

    def set_model(self, data):

        self.setModel(OneOptScheduleModel(data))
        self.resizeColumnsToContents()


class OneOptScheduleToPandas:

    def __init__(self, schedule, case):

        labels = ["#", "Stop Name", "Arrive Time", "Depart Time", "Original Schedule"]

        if case == 1:
            stops = schedule.stops
        else:
            stops = schedule.pre_cleaned_stops

        compiled = []
        for x, stop in enumerate(stops):
            stop_number = x + 1
            stop_name = stop.stop_name
            arrive_time = stop.arrive_time
            depart_time = stop.depart_time
            original_schedule = stop.original_schedule_name
            # lunch_after = stop.lunch_after
            compiled.append([stop_number, stop_name, arrive_time, depart_time, original_schedule])

        self.df = pd.DataFrame(compiled, columns=labels)
