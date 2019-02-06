from datetime import datetime, timedelta
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine
from PyQt5.QtWidgets import (QGridLayout, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableView,
                             QDialog, QMessageBox, QRadioButton, QLineEdit, QComboBox, QCheckBox)
from PyQt5.QtCore import Qt, QAbstractTableModel
import pandas as pd
import PyPDF2
import re
from openpyxl import load_workbook
from LookupReader import FacilityNameEntry, AddFrequencyCode
from ScheduleClasses import Schedule, Stop
from GeneralMethods import duration_between_two_times as dur
from GeneralMethods import time_to_datetime


def hcr_to_time(var):
    return datetime.strptime(var, "%H%M").time()


class HalfStop:

    def __init__(self, time, tag, time_zone, stop_name, nass_code=None):
        self.time = time
        self.tag = tag
        self.time_zone = time_zone
        self.stop_name = stop_name
        self.nass_code = nass_code

    def push_stop(self, minutes):

        time_dt = datetime(100, 1, 1, self.time.hour, self.time.minute)
        self.time = (time_dt + timedelta(minutes=minutes)).time()

    def adjust_tz(self, to_tz):

        if to_tz in ('UK', "UK", None, ""):
            print("unknown timezones, adjust by hand")
            return

        if to_tz not in ('ET', 'CT', 'MT', 'PT', 'ADJ', 'PR', 'HI'):
            print("uh oh, timezonio")

        to_list = ['ET', 'CT', 'MT', 'PT', 'ADJ', 'PR', 'HI']
        et_list = [0, -1, -2, -3, 0, 0, 0]
        ct_list = [1, 0, -1, -2, 0, 0, 0]
        mt_list = [2, 1, 0, -1, 0, 0, 0]
        pt_list = [3, 2, 1, 0, 0, 0, 0]
        adj_list = [0, 0, 0, 0, 0, 0, 0]
        pr_list = [0, 0, 0, 0, 0, 0, 0]
        hi_list = [0, 0, 0, 0, 0, 0, 0]
        comp_list = [et_list, ct_list, mt_list, pt_list, adj_list, pr_list, hi_list]

        from_tz = self.time_zone

        from_index = to_list.index(from_tz)
        to_index = to_list.index(to_tz)

        minutes = 60 * comp_list[from_index][to_index]
        self.push_stop(minutes)
        if minutes != 0:
            self.time_zone = to_tz


class UPart:

    def __init__(self, part, schedule_num, freq_code, half_stops, vehicle_type, mileage, purpose=None, mail_class=None):

        self.part = part
        self.schedule_num = schedule_num
        self.freq_code = freq_code.strip().zfill(4)
        self.half_stops = half_stops
        self.vehicle_type = vehicle_type
        self.mileage = mileage
        self.purpose = purpose
        self.mail_class = mail_class
        self.fix_missing_times()
        self.annual_trips = None
        self.freq_code_desc = None
        self.original_freq_code = str(self.freq_code)

        # print("UPart:", schedule_num, purpose, mail_class)

    def fix_missing_times(self):

        last_tag = ""
        stops_to_fix = []
        for stop in self.half_stops:
            if stop.tag == last_tag:
                stops_to_fix.append(stop)
            last_tag = stop.tag

        for stop in stops_to_fix:
            index = self.half_stops.index(stop)
            old_stop = stop
            new_datetime = datetime(100, 1, 1, old_stop.time.hour, old_stop.time.minute)
            new_datetime = new_datetime - timedelta(minutes=1)
            new_time = new_datetime.time()
            tag = "Ar"
            time_zone = old_stop.time_zone
            name = old_stop.stop_name
            new_stop = HalfStop(new_time, tag, time_zone, name, nass_code=stop.nass_code)
            self.half_stops.insert(index, new_stop)

        return

    def add_arrival_to_start(self):

        first_stop = self.half_stops[0]
        one_min = timedelta(minutes=1)

        if first_stop.tag == 'Ar':
            return

        start_zone = first_stop.time_zone
        start_stop = first_stop.stop_name

        start_time = first_stop.time
        start_dt = datetime(100, 1, 1, start_time.hour, start_time.minute)
        start_dt -= one_min
        start_time = start_dt.time()

        new_first_stop = HalfStop(start_time, 'Ar', start_zone, start_stop, nass_code=first_stop.nass_code)
        self.half_stops.insert(0, new_first_stop)

    def add_departure_to_end(self):

        last_stop = self.half_stops[-1]
        one_min = timedelta(minutes=1)

        if last_stop.tag == 'Lv':
            return

        start_zone = last_stop.time_zone
        start_stop = last_stop.stop_name

        start_time = last_stop.time
        start_dt = datetime(100, 1, 1, start_time.hour, start_time.minute)
        start_dt += one_min
        start_time = start_dt.time()

        new_last_stop = HalfStop(start_time, 'Lv', start_zone, start_stop, nass_code=last_stop.nass_code)
        self.half_stops.append(new_last_stop)

    def add_hcr_freq_code_info(self, annual_trips, freq_code_desc):

        self.annual_trips = annual_trips
        self.freq_code_desc = freq_code_desc

    def raw_duration(self):

        return dur(self.half_stops[0].time, self.half_stops[-1].time)

    def __str__(self):

        return str(self.schedule_num)


class HCRReader:

    def __init__(self, pdf_name, lookups, source_app, check_times, strict_merge,
                 custom_pdc_name=None, custom_pdc_address=None):

        print("Reading plate: " + pdf_name)
        print("Strict merge: ", strict_merge)
        print("Check times: ", check_times)

        self.file_name = pdf_name
        self.source_app = source_app
        self.strict_merge = strict_merge

        self.min_check_time = check_times[0]
        self.max_check_time = check_times[1]
        self.max_combined_time = check_times[2]

        # custom input details
        self.custom_pdc_name = custom_pdc_name
        self.custom_pdc_address = custom_pdc_address

        # switch codes
        self.from_codes = []
        self.to_codes = []
        self.not_from_codes = []
        self.not_to_codes = []
        self.set_switch_codes()

        self.lb = lookups
        self.source = "HCR"
        self.source_type = "PDF"

        self.all_data = None
        self.hcr_line = None
        self.contract_start = None
        self.contract_end = None
        self.effective_date = None
        self.estimated_annual_hours = None
        self.estimated_annual_miles = None
        self.is_readable = False
        self.read_in_pdf_super()

        if not self.is_readable:
            return

        self.plate_number = None
        self.find_plate_number()

        self.cargo_van_count = None
        self.tractor_sa_count = None
        self.tractor_ta_count = None
        self.trailer_count = None
        self.read_in_vehicles()

        self.first_panel_index = None
        self.last_panel_index = None
        self.panels = None
        try:
            self.separate_panels()
        except:
            print("Could not separate panels")
            self.is_readable = False
            return

        # Read in all panels
        self.found_codes = []
        self.found_stops = []
        self.time_zones = []
        self.first_stop = None
        self.all_panels = []
        self.trip_nums = []
        self.trip_purposes = []
        self.get_trip_purposes()
        self.build_u_parts_super()

        if not self.is_readable:
            return

        # Facility Names
        self.hcr_pdc_name = None
        if self.custom_pdc_name:
            self.hcr_pdc_name = self.custom_pdc_name
        else:
            self.find_facility_names()

        self.pvs_pdc = None
        self.pvs_name = None
        self.short_name = None
        self.set_names_from_index()

        # Collect stop addresses
        self.stop_addresses = []
        self.picked_addresses = []
        self.get_stop_addresses()

        # Set P&DC details
        self.pdc_address = None
        if self.custom_pdc_address:
            self.pdc_address = self.custom_pdc_address
        else:
            self.set_pdc_address()

        # Adjust stop times
        self.pdc_tz = self.set_pdc_tz()
        self.adjust_all_stop_times()

        # Collect frequency code information, attach to pre-schedules (u-parts)
        self.new_codes = []
        self.read_in_frequency_codes()
        self.attach_freq_code_info_to_uparts()

        # Build final schedules
        self.schedules = []
        self.merge_u_parts()
        self.set_day_strings()
        self.find_network_schedules()
        self.get_other_details()

    def find_network_schedules(self):

        network_sites = self.lb.network_nass_codes
        for schedule in self.schedules:
            schedule.is_network_schedule(network_sites)
            schedule.crosses_state_lines()

    def set_switch_codes(self):
        wb = load_workbook("Lookups/SwitchCodes.xlsx")
        ws = wb['Code to Switch Code']
        max_row = ws.max_row
        x = 2
        while x <= max_row:
            self.from_codes.append(ws["A" + str(x)].value)
            self.to_codes.append(ws["B" + str(x)].value)
            x += 1

        ws = wb['Not Code to Switch Code']
        max_row = ws.max_row
        x = 2
        while x <= max_row:
            self.not_from_codes.append(ws["A" + str(x)].value)
            self.not_to_codes.append(ws["B" + str(x)].value)
            x += 1

    def read_in_pdf_super(self):

        try:
            self.is_readable = self.read_in_pdf_pdfparser(self.file_name)
        except:
            pass

        if self.is_readable:
            return

        print("First read attempt not successful.")

        try:
            self.read_in_pdf_pypdf()
            self.is_readable = self.read_in_pdf_pdfparser("PDF Copy.pdf")
        except:
            self.is_readable = False
            pass

        if not self.is_readable:
            print("Second read attempt not successful.")
            QMessageBox.question(None, "Misread Plate", "Could not read plate " + self.file_name + ".", QMessageBox.Ok)

        return

    def read_in_pdf_pdfparser(self, file_name):

        all_data = []
        fp = open(file_name, 'rb')
        parser = PDFParser(fp)
        doc = PDFDocument()
        parser.set_document(doc)
        doc.set_parser(parser)
        doc.initialize('')
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        # Process each page contained in the document.
        for page in doc.get_pages():
            interpreter.process_page(page)
            layout = device.get_result()
            for lt_obj in layout:
                if isinstance(lt_obj, LTTextBox) or isinstance(lt_obj, LTTextLine):
                    all_data.append(lt_obj.get_text())

        self.all_data = all_data

        check1 = False
        check2 = False
        check3 = False
        check4 = False

        for line in self.all_data:
            if "SCHEDULE" in line:
                check1 = True
            if "EFFECTIVE" in line:
                check2 = True
            if "HCR:" in line:
                check3 = True
            if "SUPPLIER" in line:
                check4 = True

        self.hcr_line = next(line for line in self.all_data if "HCR:" in line)

        if len(self.hcr_line) < 6:
            index = self.all_data.index(self.hcr_line)
            self.hcr_line += self.all_data[index+1]

        return check1 and check2 and check3 and check4

    def read_in_pdf_pypdf(self):

        input_pdf = PyPDF2.PdfFileReader(open(self.file_name, mode='rb'))
        output_pdf = PyPDF2.PdfFileWriter()

        input_pdf_pages = input_pdf.getNumPages()
        for page_num in range(0, input_pdf_pages):
            output_pdf.addPage(input_pdf.getPage(page_num))

        output_pdf.write(open("PDF Copy.pdf", mode='wb'))

    def find_plate_number(self):

        # first plate number option is in the file name
        first_option = self.file_name.split('/')[-1][:5]

        # second is in the HCR: line of the PDF
        try:
            second_option = self.hcr_line[4:self.hcr_line.index(',')].strip()
        except:
            second_option = None
            print(self.hcr_line)

        if first_option == second_option:
            self.plate_number = first_option
            return

        selected_option = PlateNumberEntry(first_option, second_option, self.file_name).exec_()
        if len(selected_option) != 5:
            print("WTF PLATE NUMBER!!!")

        self.plate_number = selected_option

    def read_in_vehicles(self):

        match_strings = ["VEHICLE", "REQUIREMENTS"]

        first_line = None
        try:
            first_line = next(x for x in self.all_data if all(y in x for y in match_strings))
        except:
            print("oops")

        if not first_line:
            print("Couldn't find end of frequency table")
            return
        first_index = self.all_data.index(first_line)

        one = "DATES"
        two = "SEASONAL"

        three = "SCHEDULE"
        four = "NOTES"

        five = "PHYSICAL"
        six = "LOCATION"

        last_line = None

        try:
            last_line = next(x for x in self.all_data if (one in x and two in x))
        except:
            try:
                last_line = next(x for x in self.all_data if (three in x and four in x))
            except:
                try:
                    last_line = next(x for x in self.all_data if (five in x and six in x))
                except:
                    print("oops!")

        if not last_line:
            print("Couldn't find end of vehicle table")
            return

        last_index = self.all_data.index(last_line)

        vehicle_table = self.all_data[first_index:last_index]

        vehicle_table = [x.strip() for x in vehicle_table]

        page_list = ["PAGE"]
        page_numbers = [x for x in vehicle_table if any(word in x for word in page_list)]
        page_nums = []

        for page in page_numbers:
            page_nums.append(vehicle_table[vehicle_table.index(page) + 1])

        for num in page_nums:
            vehicle_table.remove(num)

        # print("HERE!!!!")
        # print(vehicle_table)
        # print("HERE!!!!!!")

        # print(self.plate_number, vehicle_table)
        vehicle_quantities = [x for x in vehicle_table if x.isdigit() and int(x) < 1000]

        # bad_vehicles = ['Boat', 'Four Wheel Drive', 'Passenger Car', 'Pick-up Truck', 'Station Wagon']
        peak_list = ["PEAK", "Peak", "peak", "SPB", "-SPB"]

        vehicles_listed = []
        full_names = []

        for item in vehicle_table:
            item.replace('\n', '')
            item.replace('/n', '')
            if item == "(PEAK)":
                full_names[-1] = full_names[-1] + " " + item
                if not vehicles_listed[-1] == "Peak Something":
                    vehicles_listed[-1] = "Peak Something"
            # elif not any(word in item for word in bad_vehicles):
            else:
                if any(word in item for word in peak_list):
                    full_names.append(item)
                    vehicles_listed.append("Peak Something")
                elif "Tractor" in item:
                    if any(word in item for word in ("Two Axle", "Tandem")):
                        full_names.append(item)
                        vehicles_listed.append("Tractor (TA)")
                    elif any(word in item for word in ("Single Axle", "Single", "single")):
                        full_names.append(item)
                        vehicles_listed.append("Tractor (SA)")
                    elif any(word in item for word in ("SPOTTER", "Spotter", "spotter")):
                        full_names.append(item)
                        vehicles_listed.append("Spotter Tractor")
                elif "Trailer" in item:
                    full_names.append(item)
                    vehicles_listed.append("Trailer")
                elif any(word in item for word in ("TRUCK", "truck", "Truck", "VAN", "van", "Van",
                                                   "CAR", "Car", "car", "WAGON", "Wagon", "wagon")):
                    if len(item) < 50:
                        full_names.append(item)
                        vehicles_listed.append("Cargo Van")

        vehicles_paired = []

        if len(vehicles_listed) != len(vehicle_quantities):
            # first thing, check if any vehicle names have a length, then remove that from after the name
            removables = []
            for vehicle in full_names:
                for word in vehicle.split():
                    if word.isdigit():
                        removables.append(word)
            removables.reverse()
            vehicle_quantities.reverse()
            for removable in removables:
                if removable in vehicle_quantities:
                    vehicle_quantities.remove(removable)

            if len(vehicles_listed) < len(vehicle_quantities):
                if "Cargo Van" in vehicles_listed and "20" in vehicle_quantities:
                    vehicle_quantities.reverse()
                    vehicle_quantities.remove("20")
                    vehicle_quantities.reverse()
                elif "Cargo Van" in vehicles_listed and "18" in vehicle_quantities:
                    vehicle_quantities.reverse()
                    vehicle_quantities.remove("18")
                    vehicle_quantities.reverse()

            while len(vehicles_listed) < len(vehicle_quantities):
                remove = [x for x in vehicle_quantities if int(x) >= 22]
                if remove:
                    vehicle_quantities.remove(remove[0])
                    # print(vehicles_listed)
                    # print(vehicle_quantities)
                else:
                    print("Vehicles table not read correctly on plate " + self.plate_number)
                    break

            vehicle_quantities.reverse()

            if len(vehicles_listed) != len(vehicle_quantities):
                print(vehicles_listed)
                print(vehicle_quantities)
                print("Vehicles table not read correctly on plate " + self.plate_number)

        for x, vehicle in enumerate(vehicles_listed):
            vehicles_paired.append([vehicle, vehicle_quantities[x]])

        # for row in vehicles_paired:
        #     print(row)

        vehicles_paired = [x for x in vehicles_paired if not x[0] == "Peak Something"]

        cargo_vans = [x for x in vehicles_paired if x[0] == "Cargo Van"]
        self.cargo_van_count = 0
        for pair in cargo_vans:
            self.cargo_van_count += int(pair[1])

        tractor_sa = [x for x in vehicles_paired if x[0] == "Tractor (SA)"]
        self.tractor_sa_count = 0
        for pair in tractor_sa:
            self.tractor_sa_count += int(pair[1])

        tractor_ta = [x for x in vehicles_paired if x[0] == "Tractor (TA)"]
        self.tractor_ta_count = 0
        for pair in tractor_ta:
            self.tractor_ta_count += int(pair[1])

        trailers = [x for x in vehicles_paired if x[0] == "Trailer"]
        self.trailer_count = 0
        for pair in trailers:
            self.trailer_count += int(pair[1])

    def separate_panels(self):
        # separate schedule panels from rest of plate
        start_strings = ["PART", "TRIP", "FREQUENCY"]
        end_string = "MILEAGE"

        potential_panels = [x for x in self.all_data if all(word in x for word in start_strings) or end_string in x]
        self.first_panel_index = self.all_data.index(potential_panels[0])
        self.last_panel_index = self.all_data.index(potential_panels[-1])

        full_panels = []
        start_panels = []
        end_panels = []

        for panel in potential_panels:
            if all(word in panel for word in start_strings) and end_string in panel:
                full_panels.append(panel)
            elif all(word in panel for word in start_strings):
                start_panels.append(panel)
            else:
                end_panels.append(panel)

        if len(start_panels) != len(end_panels):
            print("PROBLEM WITH PANELS!!!!!!!")

        for x, panel in enumerate(start_panels):
            panel = panel + end_panels[x]
            full_panels.append(panel)

        self.panels = full_panels

    def separate_panels_old(self):
        # separate schedule panels from rest of plate

        schedule_panels = []
        max_index = len(self.all_data) - 1
        match_strings = ["FREQUENCY", "IDENTIFICATION"]

        for x, line in enumerate(self.all_data):
            if all(word in line for word in match_strings):
                max_index = x
                break

        start_strings = ["PART", "TRIP", "FREQUENCY"]
        end_string = "MILEAGE"

        for x, line in enumerate(self.all_data[1:max_index]):
            if all(word in line for word in start_strings):
                if end_string not in line:
                    end_line = next(y for y in self.all_data[x:max_index] if end_string in y)
                    line = line + end_line
                    # print(line)
                schedule_panels.append(line)

        self.panels = schedule_panels

    def build_u_parts_super(self):

        try:
            for panel in self.panels:
                self.all_panels.append(self.break_down_a_panel(panel))
        except:
            print("Couldn't break down panels on plate " + self.plate_number)
            self.is_readable = False
            return

    def break_down_a_panel(self, panel):

        rows = panel.splitlines()
        first_line = rows[0].split()
        mid = first_line.index('PART')
        left_side_cols = mid
        right_side_cols = len(first_line) - (mid + 1)

        schedule_num_line = rows[1].split()
        # because frequency codes can have a space in them (i.e., "9 1"), a simple split can't be used
        # freq_code_line = rows[2].split()
        # print(freq_code_line)
        freq_code_line = re.split(r'\s{2,}', rows[2])
        if freq_code_line[0] == "":
            del freq_code_line[0]
        # print(freq_code_line)
        mileage_line = rows[-1].split()
        vehicle_line = rows[-2].split()

        match_string = "------------"
        break_lines = [i for i, x in enumerate(rows) if match_string in x]

        break_one = break_lines[1] + 1
        break_two = break_lines[2]

        stop_names = []
        time_zone_indices = []
        name_string = ""

        for line in rows[break_one:break_two]:
            words = line.split()
            for i, word in enumerate(words):
                if word in ('CT', 'MT', 'PT', 'ET', 'PR', 'HI'):
                    time_zone_index = i
                    self.time_zones.append(word)
            num_words = time_zone_index - (left_side_cols + 1)
            x = 1
            while x <= num_words:
                name_string = name_string + " " + words[x + left_side_cols]
                x += 1
            name_string = name_string.replace(",", ", ")
            stop_names.append(name_string.strip())
            if not self.first_stop:
                self.first_stop = name_string.strip()
            self.found_stops.append(name_string.strip())
            time_zone_indices.append(time_zone_index)
            name_string = ""

        u_parts = []
        x = 0

        while x < left_side_cols:
            part = first_line[x]
            schedule_num = int(schedule_num_line[x])
            freq_code = freq_code_line[x]
            self.found_codes.append(freq_code)
            mileage = float(mileage_line[x])
            vehicle = vehicle_line[x]
            half_stops = []

            for y, line in enumerate(rows[break_one:break_two]):
                words = line.split()
                tag = words[left_side_cols]
                time_zone = words[time_zone_indices[y]]
                nass_code = words[-(right_side_cols + 1)]
                if nass_code in ("AR", "Ar", "ar", "LV", "Lv", "lv"):
                    nass_code = None
                stop_name = stop_names[y]
                if words[x] != "--":
                    time = hcr_to_time(words[x])
                    new_stop = HalfStop(time, tag, time_zone, stop_name, nass_code=nass_code)
                    half_stops.append(new_stop)

            purpose_row = [x for x in self.trip_purposes if x[0] == schedule_num]
            purpose, mail_class = None, None
            if purpose_row:
                purpose, mail_class = purpose_row[0][1], purpose_row[0][2]
                purpose_row[0][3] = True
            new_u_part = UPart(part, schedule_num, freq_code, half_stops, vehicle, mileage, purpose, mail_class)
            u_parts.append(new_u_part)
            self.trip_nums.append(schedule_num)
            x += 1

        x = 0

        while x < right_side_cols:
            part = first_line[x - right_side_cols]
            schedule_num = int(schedule_num_line[x - right_side_cols])
            freq_code = freq_code_line[x - right_side_cols]
            self.found_codes.append(freq_code)
            mileage = float(mileage_line[x - right_side_cols])
            vehicle = vehicle_line[x - right_side_cols]
            half_stops = []

            for y, line in enumerate(rows[break_one:break_two]):
                words = line.split()
                tag = ""
                for i, word in enumerate(words[time_zone_indices[y]:]):
                    if word in ('Lv', 'Ar'):
                        tag = word
                time_zone = words[time_zone_indices[y]]
                nass_code = words[-(right_side_cols + 1)]
                stop_name = stop_names[y]
                if words[x - right_side_cols] != "--":
                    time = hcr_to_time(words[x - right_side_cols])
                    new_stop = HalfStop(time, tag, time_zone, stop_name, nass_code=nass_code)
                    half_stops.append(new_stop)

            half_stops.reverse()
            purpose_row = [x for x in self.trip_purposes if x[0] == schedule_num]
            purpose, mail_class = None, None
            if purpose_row:
                purpose, mail_class = purpose_row[0][1], purpose_row[0][2]
                purpose_row[0][3] = True
            new_u_part = UPart(part, schedule_num, freq_code, half_stops, vehicle, mileage, purpose, mail_class)
            self.trip_nums.append(schedule_num)
            u_parts.append(new_u_part)
            x += 1

        return u_parts

    def find_facility_names(self):

        # Check if plate is in persistent plate list
        if self.plate_number in self.lb.persistent_plates:
            short_name = self.lb.persistent_sites[self.lb.persistent_plates.index(self.plate_number)]
            self.hcr_pdc_name = self.lb.hcr_pdcs[self.lb.short_names.index(short_name)]
            return

        # first option is in "HCR:" row
        hcr_line_name = self.hcr_line[self.hcr_line.index(',')+1:self.hcr_line.index("-")].strip()

        # second option is admin official
        start_index = self.hcr_line.index('OFFICIAL:') + 10
        end_index = self.hcr_line.index("ESTIMATED")
        admin_official_name = self.hcr_line[start_index:end_index].strip()

        # third option is first stop name
        first_stop_name = self.first_stop

        if hcr_line_name == admin_official_name == first_stop_name:
            self.hcr_pdc_name = hcr_line_name
            return

        # next check if any of the three are in the lookup book
        check1 = hcr_line_name in self.lb.hcr_pdcs
        check2 = admin_official_name in self.lb.hcr_pdcs
        check3 = first_stop_name in self.lb.hcr_pdcs

        self.hcr_pdc_name, add = HCRPDCEntry(hcr_line_name, check1, admin_official_name, check2, first_stop_name, check3,
                                             self.file_name).exec_()

        if add:
            short_name = self.lb.short_names[self.lb.hcr_pdcs.index(self.hcr_pdc_name)]
            self.lb.add_persistent_info(self.plate_number, short_name)

    def set_names_from_index(self):

        # if name not in Lookup book use facility name entry tool
        for name in self.lb.hcr_pdcs:
            if name:
                if self.hcr_pdc_name in name:
                    self.hcr_pdc_name = name

        if self.hcr_pdc_name not in self.lb.hcr_pdcs:
            FacilityNameEntry(source_type="hcr_pdc", lb=self.lb,
                              file_name=self.file_name, hcr_pdc=self.hcr_pdc_name).exec_()

        print(self.hcr_pdc_name)
        index = self.lb.hcr_pdcs.index(self.hcr_pdc_name)
        self.pvs_pdc = self.lb.pvs_pdcs[index]
        self.pvs_name = self.lb.pvs_names[index]
        self.short_name = self.lb.short_names[index]
        return

    def get_stop_addresses(self):

        start_index = self.all_data.index('PHYSICAL LOCATION OF POINTS SERVED:\n')
        end_index = self.all_data.index('TRIP PURPOSE AND MAIL CLASS:\n')

        address_table = self.all_data[start_index + 1:end_index]

        del_list = []
        for item in address_table:
            if "HCR:" in item:
                del_list.append(item)
            elif "EFFECTIVE:" in item:
                del_list.append(item)
            elif "TRIP " in item:
                del_list.append(item)
            elif len(item) < 4 and "--" not in item:
                del_list.append(item)

        for item in del_list:
            address_table.remove(item)

        address_lines = []
        for line in address_table:
            for subline in line.splitlines():
                address_lines.append(subline)

        phone_num_indices = []
        pattern = '\d\d\d-\d\d\d-\d\d\d\d'
        r = re.compile(pattern)
        for x, address in enumerate(address_lines):
            if r.search(address) or "--" in address:
                phone_num_indices.append(x)

        x = 0
        address_string = []
        address_strings = []

        for num in phone_num_indices:
            while x < num:
                address_string.append(address_lines[x].strip())
                x += 1
            address_strings.append(address_string)
            address_string = []
            x += 1

        for address in address_strings:
            address_edit = [x for x in address if x != '']
            address_string = ""
            name = address_edit[0]
            for line in address_edit[1:]:
                address_string = address_string + line + " "
            address_to_add = address_string.strip()
            self.stop_addresses.append([name, address_to_add])

    def set_pdc_address(self):

        less_state = self.hcr_pdc_name[:self.hcr_pdc_name.rfind(',')]
        potential_addresses = [x for x in self.stop_addresses if x[0] == less_state]

        if not potential_addresses:
            print("Error finding P&DC address on plate", self.plate_number)
            return
        elif len(potential_addresses) > 1:
            self.pdc_address = AddressSelector(self.plate_number, self.hcr_pdc_name, potential_addresses).exec_()[1]
            return

        self.pdc_address = potential_addresses[0][1]

    def set_pdc_tz(self):

        try:
            return self.time_zones[self.found_stops.index(self.hcr_pdc_name)]
        except:
            pass

        time_zones = list(set(self.time_zones))
        if len(time_zones) == 1:
            return time_zones[0]

        print("AUGH TIME ZONES!!")
        return "UK"

    def adjust_all_stop_times(self):

        for panel in self.all_panels:
            for u_part in panel:
                for stop in u_part.half_stops:
                    stop.adjust_tz(self.pdc_tz)

    def read_in_frequency_codes(self):

        first_check = ["FREQUENCY", "IDENTIFICATION"]
        first_line = next(x for x in self.all_data if all(word in x for word in first_check))
        if not first_line:
            print("Couldn't find frequency table")
            return
        first_index = self.all_data.index(first_line)

        last_check = ["VEHICLE", "REQUIREMENTS"]
        last_line = next(x for x in self.all_data if all(word in x for word in last_check))
        if not last_line:
            print("Couldn't find end of frequency table")
            return
        last_index = self.all_data.index(last_line)

        frequency_table = self.all_data[first_index:last_index]
        found_codes = list(set(self.found_codes))
        found_codes = [x.strip() for x in found_codes]
        found_codes.sort()

        frequency_table = [x.strip() for x in frequency_table]
        listed_codes = [x for x in frequency_table if x in found_codes]

        # this should only be triggered if there's a page num amidst the freq table that is the same as a freq code

        pattern = '\.\d\d'
        r = re.compile(pattern)
        last_panel_index = self.last_panel_index

        annual_trips = [x.strip() for x in self.all_data[last_panel_index:] if r.search(x) and len(x.strip()) < 10]

        # print(len(annual_trips), annual_trips)
        # print(len(listed_codes), listed_codes)

        if len(annual_trips) == len(listed_codes) - 1:
            if "9" in listed_codes[-1]:
                del listed_codes[-1]

        if len(listed_codes) < len(annual_trips):
            listed_codes = [x for x in frequency_table if x in found_codes or
                            ("9" in x and x.strip() not in annual_trips)]

        if len(annual_trips) < len(listed_codes):
            non_nine_duplicates = [x for x in listed_codes if listed_codes.count(x) > 1 and "9" not in x]
            non_nine_duplicates = list(set(non_nine_duplicates))
            listed_codes.reverse()
            for item in non_nine_duplicates:
                listed_codes.remove(item)
            listed_codes.reverse()

        if len(annual_trips) < len(listed_codes):
            seen = set()
            listed_codes = [x for x in listed_codes if not (x in seen or seen.add(x))]

        if len(annual_trips) != len(listed_codes):
            print(len(annual_trips), annual_trips)
            print(len(listed_codes), listed_codes)
            print("Misread frequency code")
            return

        words = ("ANNUAL", "FREQUENCY", "IDENTIFICATION", "PAGE", "EFFECTIVE", "HCR:", "LENGTH", "CUBES", "PAYLOAD",
                 "END", "12/24", "12/23", "12/22", "12/21")

        list_descriptions = [x for x in frequency_table if x not in found_codes and x not in annual_trips and
                             not any(word in x for word in words) and not x.isdigit() and x not in ('', ' ', None)]

        while len(list_descriptions) > len(annual_trips):
            del list_descriptions[-1]

        # while len(annual_trips) > len(list_descriptions):
        #     del annual_trips[-1]

        if len(list_descriptions) != len(annual_trips):
            print(len(list_descriptions), list_descriptions)
            print(len(annual_trips), annual_trips)
            print(len(listed_codes), listed_codes)
            print("Misread frequency descriptions")
            return

        # for x in range(0,len(list_descriptions)):
        #     print(listed_codes[x], annual_trips[x], list_descriptions[x])

        self.new_codes = []
        for x, code in enumerate(listed_codes):
            code = code.strip().zfill(4)
            self.new_codes.append([code, annual_trips[x], list_descriptions[x]])

        for new_code in self.new_codes:
            # if "9" not in new_code[0]:
            if float(new_code[1]) < 40:
              pass
            elif new_code[0] not in [x[0] for x in self.lb.known_codes]:
                AddFrequencyCode(new_code[0], new_code[1], new_code[2], self.lb, plate_num=self.plate_number).exec_()
            elif not next(x[1] for x in self.lb.known_codes if new_code[0] == x[0]):
                self.lb.add_hcr_info(new_code)

    def attach_freq_code_info_to_uparts(self):

        for panel in self.all_panels:
            for u_part in panel:
                try:
                    frequency_code = next(x for x in self.new_codes if x[0] == u_part.freq_code)
                except:
                    print("Couldn't find frequency code for ", u_part.freq_code, u_part.schedule_num)
                if not frequency_code:
                    print("Couldn't find listed frequency code!")
                u_part.add_hcr_freq_code_info(frequency_code[1], frequency_code[2])

    def get_stop_address(self, stop_name):

        possible_names = (stop_name, stop_name[:-3], stop_name.split(",")[0], stop_name.split(", ")[0])
        found_addresses = [x for x in self.stop_addresses if x[0] in possible_names]

        if not found_addresses:
            found_addresses = self.try_harder_stop_address(possible_names)

        if len(found_addresses) < 1:
            print("Couldn't find address for " + stop_name + " " + self.plate_number)
            self.stop_addresses.append([stop_name, "No address found"])
            return "No address found"
        if len(found_addresses) > 1:
            print("Multiple addresses found for " + stop_name)
            if stop_name in [x[0] for x in self.picked_addresses]:
                return next(x[1] for x in self.picked_addresses if x[0] == stop_name)
            else:
                picked_address = AddressSelector(self.plate_number, stop_name, found_addresses).exec_()
                self.picked_addresses.append([stop_name, picked_address[1]])
                return picked_address[1]
        else:
            return found_addresses[0][1]

    def try_harder_stop_address(self, possible_names):

        possible_addresses = [x[1] for x in self.stop_addresses if any(y in x[1] for y in possible_names)]
        new_addresses = []

        if not possible_addresses:
            return []

        # print(possible_addresses)
        # print(possible_names)

        for possible_address in possible_addresses:
            split_up = possible_address.split()
            try:
                name = next(x for x in split_up if x in possible_names)
                name_index = split_up.index(name)
                name_len = 1
            except:
                name_full = next(x for x in possible_names if x in possible_address).split()
                name = name_full[0]
                name_len = len(name_full)
                name_index = split_up.index(name)

            address_string = ""
            name = ""
            # print("line 942:", name_index, name_len)
            for x in split_up[name_index:name_index + name_len]:
                name += " " + x

            name = name.strip()
            # print(name)
            for word in split_up[name_index + name_len:]:
                address_string += word + " "
            address_string.strip()
            new_addresses.append([name, address_string])

        return new_addresses

    def merge_u_parts(self):

        schedules = []
        compiled_schedules = []
        unmatched_u_parts = []

        for x, panel in enumerate(self.all_panels):
            match_1 = False
            match_2 = False
            match_3 = False
            match_4 = False
            match_5 = False

            length = len(panel)

            if length > 0:
                u_part_1 = panel[0]
                schedule_1 = u_part_1.schedule_num
            if length > 1:
                u_part_2 = panel[1]
                schedule_2 = u_part_2.schedule_num
            if length > 2:
                u_part_3 = panel[2]
                schedule_3 = u_part_3.schedule_num
            if length > 3:
                u_part_4 = panel[3]
                schedule_4 = u_part_4.schedule_num
            if length > 4:
                u_part_5 = panel[4]
                schedule_5 = u_part_5.schedule_num

            if length == 1:
                unmatched_u_parts.append(u_part_1)

            if length == 2:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)

            if length == 3:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_1 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_3))
                    match_1 = True
                    match_3 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)

            if length == 4:
                if schedule_1 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_4))
                    match_1 = True
                    match_4 = True

                if schedule_2 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_3))
                    match_2 = True
                    match_3 = True

                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_3 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_4))
                    match_3 = True
                    match_4 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)
                if not match_4:
                    unmatched_u_parts.append(u_part_4)

            if length == 5:
                if schedule_1 == schedule_2 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_2))
                    match_1 = True
                    match_2 = True

                if schedule_1 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_3))
                    match_1 = True
                    match_3 = True

                if schedule_1 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_4))
                    match_1 = True
                    match_4 = True

                if schedule_1 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_1, u_part_5))
                    match_1 = True
                    match_5 = True

                if schedule_2 == schedule_3 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_3))
                    match_2 = True
                    match_3 = True

                if schedule_2 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_4))
                    match_2 = True
                    match_4 = True

                if schedule_2 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_2, u_part_5))
                    match_2 = True
                    match_5 = True

                if schedule_3 == schedule_4 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_4))
                    match_3 = True
                    match_4 = True

                if schedule_3 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_3, u_part_5))
                    match_3 = True
                    match_5 = True

                if schedule_4 == schedule_5 - 1:
                    schedules.append(self.merge_two_u_parts(u_part_4, u_part_4))
                    match_4 = True
                    match_5 = True

                if not match_1:
                    unmatched_u_parts.append(u_part_1)
                if not match_2:
                    unmatched_u_parts.append(u_part_2)
                if not match_3:
                    unmatched_u_parts.append(u_part_3)
                if not match_4:
                    unmatched_u_parts.append(u_part_4)
                if not match_5:
                    unmatched_u_parts.append(u_part_5)

        for part in unmatched_u_parts:
            schedules.append([self.u_part_to_schedule(part)])

        for items in schedules:
            for item in items:
                compiled_schedules.append(item)

        self.schedules = compiled_schedules

    def u_part_to_schedule(self, u_part):

        u_part.add_arrival_to_start()
        u_part.add_departure_to_end()

        combined_stops = []
        for x, half_stop in enumerate(u_part.half_stops):
            if half_stop.tag == 'Ar':
                next_stop = u_part.half_stops[x + 1]
                new_stop = Stop(arrive_time=half_stop.time, depart_time=next_stop.time,
                                stop_name=half_stop.stop_name, nass_code=half_stop.nass_code)
                new_stop.add_address(self.get_stop_address(new_stop.stop_name))
                combined_stops.append(new_stop)

        vc = u_part.vehicle_type[0]

        if vc == "T":
            vehicle_category = "Single"
        # else vc == "S":
        else:
            vehicle_category = "11-Ton"
        # elif self.cargo_van_count > 0:
        #     vehicle_category = "Cargo Van"
        # else:
        #     vehicle_category = "Unknown"

        annual_trips = float(u_part.annual_trips)
        purpose = u_part.purpose
        mail_class = u_part.mail_class

        return_schedule = Schedule(source_type=self.source_type, source=self.source, part=u_part.part,
                                   plate_num=self.plate_number, trip_num=u_part.schedule_num,
                                   freq_code=u_part.freq_code, stops=combined_stops, vehicle_type=u_part.vehicle_type,
                                   hcr_pdc=self.hcr_pdc_name, pvs_pdc=self.pvs_pdc, pvs_name=self.pvs_name,
                                   pdc_address=self.pdc_address, mileage=u_part.mileage,
                                   vehicle_category=vehicle_category, short_name=self.short_name,
                                   annual_trips=annual_trips, purpose=purpose, mail_class=mail_class)

        return_schedule.add_flag("This schedule is only one column of a U")

        return return_schedule

    def merge_two_u_parts(self, u_part_1, u_part_2):

        # at this stage, HCR schedules should start at a 'Lv' and end at an 'Ar'
        # note this is done here and not in the UPart itself, because of the merge
        # this is where we should check the layover as well

        merge_anyway = True  # this will be the final decider of whether the merge the U parts
        stop1 = u_part_1.half_stops[-1].time
        stop2 = u_part_2.half_stops[0].time
        layover = dur(stop1, stop2)

        # are the frequency codes different?
        diff_freq_codes = u_part_1.freq_code != u_part_2.freq_code

        # have we previously linked these two codes?
        switch_code_check = False
        if u_part_1.freq_code in self.from_codes:
            if u_part_2.freq_code == self.to_codes[self.from_codes.index(u_part_1.freq_code)]:
                switch_code_check = True

        # have we previously said not to link these two codes?
        previous_no = False
        if u_part_1.freq_code in self.not_from_codes:
            if u_part_2.freq_code == self.not_to_codes[self.not_from_codes.index(u_part_1.freq_code)]:
                previous_no = True

        # are the two u_parts on different days?
        different_day_check = False
        left_start = u_part_1.half_stops[0].time
        right_start = u_part_2.half_stops[0].time
        if time_to_datetime(left_start) > time_to_datetime(right_start):
            different_day_check = True

        # is this a deadhead schedule?
        dead_head_check = True
        if len(u_part_1.half_stops) == len(u_part_2.half_stops):
            for x, stop in enumerate(u_part_1.half_stops):
                other_stop = u_part_2.half_stops[x]
                if other_stop.time == stop.time:
                    continue
                else:
                    dead_head_check = False
                    break
        else:
            dead_head_check = False

        if dead_head_check or previous_no:  # never merge deadheads?
            return [self.u_part_to_schedule(u_part_1), self.u_part_to_schedule(u_part_2)]

        if self.strict_merge:  # don't bother with check times in this case
            if diff_freq_codes:
                if not different_day_check:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                elif switch_code_check:
                    merge_anyway = True
                else:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                    if merge_anyway:
                        ask_string = "Apply this answer to future occurrences of this frequency code combination?"
                        add_to_list = QMessageBox.question(None, "Remember this answer?", ask_string,
                                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if add_to_list == QMessageBox.Yes:
                            self.add_switch_code(u_part_1.freq_code, u_part_2.freq_code)
                    else:
                        ask_string = "Apply this answer to future occurrences of this frequency code combination?"
                        add_to_list = QMessageBox.question(None, "Remember this answer?", ask_string,
                                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if add_to_list == QMessageBox.Yes:
                            self.add_not_switch_code(u_part_1.freq_code, u_part_2.freq_code)

            else:
                merge_anyway = True

        else:  # now we need to use the check times

            check_time = self.min_check_time
            check_time_2 = self.max_check_time
            check_time_3 = self.max_combined_time
            combined_duration = u_part_1.raw_duration() + u_part_2.raw_duration() + layover

            if combined_duration > check_time_3 or layover > check_time_2:
                merge_anyway = False
            elif diff_freq_codes:
                if not different_day_check:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                elif switch_code_check:
                    merge_anyway = True
                else:
                    merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()
                    if merge_anyway:
                        ask_string = "Apply this answer to future occurrences of this frequency code combination?"
                        add_to_list = QMessageBox.question(None, "Remember this answer?", ask_string,
                                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if add_to_list == QMessageBox.Yes:
                            self.add_switch_code(u_part_1.freq_code, u_part_2.freq_code)
            elif layover < check_time:
                merge_anyway = True
            else:
                merge_anyway = MergeTwoUParts(u_part_1, u_part_2, layover, self.plate_number).exec_()

        if not merge_anyway:
            return [self.u_part_to_schedule(u_part_1), self.u_part_to_schedule(u_part_2)]

        u_part_1.add_arrival_to_start()
        u_part_2.add_departure_to_end()

        merged_stops = u_part_1.half_stops + u_part_2.half_stops

        combined_stops = []
        for x, half_stop in enumerate(merged_stops):
            if half_stop.tag == 'Ar':
                next_stop = merged_stops[x + 1]
                new_stop = Stop(arrive_time=half_stop.time, depart_time=next_stop.time,
                                stop_name=half_stop.stop_name, time_zone=half_stop.time_zone,
                                nass_code=half_stop.nass_code)
                new_stop.add_address(self.get_stop_address(new_stop.stop_name))
                combined_stops.append(new_stop)

        part = u_part_1.part
        schedule_num = u_part_1.schedule_num
        freq_code = u_part_1.freq_code
        mileage = u_part_1.mileage + u_part_2.mileage

        vt_left = u_part_1.vehicle_type
        vt_right = u_part_2.vehicle_type
        vc_left = vt_left[0]
        vc_right = vt_right[0]

        if vt_left == vt_right:
            vehicle_type = vt_left
        else:
            vehicle_type = vt_left + "/" + vt_right + " Unmatched"

        if vc_left == vc_right:
            if vc_left == "T":
                vehicle_category = "Single"
            else:
                # elif vc_left == "S":
                vehicle_category = "11-Ton"
        #     elif self.cargo_van_count > 0:
        #         vehicle_category = "Cargo Van"
        #     else:
        #         vehicle_category = "Matched, Unknown"
        else:
            vehicle_category = "Unmatched, Unknown"

        annual_trips = None
        if float(u_part_1.annual_trips) == float(u_part_2.annual_trips):
            annual_trips = float(u_part_1.annual_trips)

        if u_part_1.purpose == u_part_2.purpose:
            purpose = u_part_1.purpose
        else:
            purpose = str(u_part_1.purpose) + "/" + str(u_part_2.purpose)

        if u_part_1.mail_class == u_part_2.mail_class:
            mail_class = u_part_1.mail_class
        else:
            mail_class = str(u_part_1.mail_class) + "/" + str(u_part_2.mail_class)

        return_schedule = Schedule(source_type=self.source_type, source=self.source, part=part,
                                   plate_num=self.plate_number, trip_num=schedule_num, freq_code=freq_code,
                                   stops=combined_stops, vehicle_type=vehicle_type, hcr_pdc=self.hcr_pdc_name,
                                   pvs_pdc=self.pvs_pdc, pvs_name=self.pvs_name, pdc_address=self.pdc_address,
                                   mileage=mileage, vehicle_category=vehicle_category, short_name=self.short_name,
                                   wannual_trips=annual_trips, purpose=purpose, mail_class=mail_class)

        return [return_schedule]

    def add_switch_code(self, freq_code1, freq_code2):

        wb = load_workbook("Lookups/SwitchCodes.xlsx")
        ws = wb['Code to Switch Code']
        new_row = ws.max_row + 1
        ws["A" + str(new_row)].value = freq_code1
        ws["B" + str(new_row)].value = freq_code2
        wb.save("Lookups/SwitchCodes.xlsx")
        wb.close()
        self.from_codes.append(freq_code1)
        self.to_codes.append(freq_code2)

    def add_not_switch_code(self, freq_code1, freq_code2):

        wb = load_workbook("Lookups/SwitchCodes.xlsx")
        ws = wb['Not Code to Switch Code']
        new_row = ws.max_row + 1
        ws["A" + str(new_row)].value = freq_code1
        ws["B" + str(new_row)].value = freq_code2
        wb.save("Lookups/SwitchCodes.xlsx")
        wb.close()
        self.not_from_codes.append(freq_code1)
        self.not_to_codes.append(freq_code2)

    def set_day_strings(self):

        for schedule in self.schedules:
            schedule.set_bin_string(self.lb)
            schedule.is_holiday_schedule()

    def get_potential_pdcs_and_addresses(self):

        potential_pdc_names = [x[0] for x in self.stop_addresses]
        potential_pdc_addresses = [x[1] for x in self.stop_addresses]

        return potential_pdc_names, potential_pdc_addresses

    def get_trip_purposes(self):

        first_check = "TRIP PURPOSE"
        first_index = next(i for i, x in enumerate(self.all_data) if first_check in x)
        potential_table = self.all_data[first_index:]
        print(len(potential_table))
        last_item = "blank"
        del_list = []
        for i, item in enumerate(potential_table):
            if "PAGE" in last_item:
                del_list.append(item)
            last_item = item

        check_list = ["TRIP", "PURPOSE", "MAIL CLASS", "HCR: ", "PAGE ", "EFFECTIVE:", "PHYSICAL LOCATION OF"]
        potential_table = [x for x in potential_table if not any(word in x for word in check_list)]
        potential_table = [x for x in potential_table if x not in del_list]
        # for i, item in enumerate(potential_table):
        #     print(i, len(item.splitlines()))
        if len(potential_table)%3 != 0:
            print("PROBLEM WITH PURPOSE TABLE")
            return

        potential_table = [x.splitlines() for x in potential_table]
        trip_nums = []
        purposes = []
        mail_classes = []
        for x in range(0, len(potential_table), 3):
            trip_nums += potential_table[x]
        for x in range(1, len(potential_table), 3):
            purposes += potential_table[x]
        for x in range(2, len(potential_table), 3):
            mail_classes += potential_table[x]

        if len(trip_nums) != len(purposes) or len(trip_nums) != len(mail_classes):
            print("PROBLEM WITH PURPOSE TABLE!!")
            return

        for x, item in enumerate(trip_nums):
            self.trip_purposes.append([int(item), purposes[x], mail_classes[x], False])

    def get_other_details(self):

        possible = self.all_data[:self.first_panel_index]

        mileage_row = next(x for x in possible if "SCHEDULE MILES" in x).split()
        mileage_index = mileage_row.index("MILES:")
        mileage = float(mileage_row[mileage_index+1].strip().replace(',',''))

        print("mileage: ", mileage)

        term_row = next(x for x in possible if "CONTRACT TERM" in x).split()
        term_index = term_row.index("TERM:")
        start_date = term_row[term_index + 1]
        end_date = term_row[term_index + 3]
        print("start:", start_date)
        print("end:", end_date)

        effective_row = next(x for x in possible if "EFFECTIVE DATE" in x).split()
        effective_index = effective_row.index("DATE:")
        effective_date = effective_row[effective_index + 1]

        print("effective:", effective_date)


class PlateNumberEntry(QDialog):

    def __init__(self, first_option, second_option, file_name, parent=None):
        super(PlateNumberEntry, self).__init__(parent)

        self.setWindowTitle("Enter Plate Number")

        self.file_name = file_name.split("/")[-1]

        self.first_option = first_option
        self.second_option = second_option
        self.third_option = QLineEdit()

        self.radio_one = QRadioButton()
        self.radio_two = QRadioButton()
        self.radio_three = QRadioButton()

        self.case = 1

        self.radio_one.case = 1
        self.radio_two.case = 2
        self.radio_three.case = 3

        self.radio_one.toggled.connect(self.change_case)
        self.radio_two.toggled.connect(self.change_case)
        self.radio_three.toggled.connect(self.change_case)

        self.radio_one.setChecked(True)

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        # self.cancel_button = QPushButton("Cancel", self)
        # self.cancel_button.clicked[bool].connect(self.cancel)

        self.initUI()

    def initUI(self):

        main_layout = QVBoxLayout()
        inner_grid = QGridLayout()

        header_label = QLabel("Please input plate number for: " + self.file_name)

        inner_grid.addWidget(QLabel(self.first_option), 0, 0)
        inner_grid.addWidget(QLabel(self.second_option), 1, 0)
        inner_grid.addWidget(self.third_option, 2, 0)

        inner_grid.addWidget(self.radio_one, 0, 1)
        inner_grid.addWidget(self.radio_two, 1, 1)
        inner_grid.addWidget(self.radio_three, 2, 1)

        main_layout.addWidget(header_label)
        main_layout.addLayout(inner_grid)
        main_layout.addWidget(self.accept_button)
        self.setLayout(main_layout)

    def change_case(self):

        radiobutton = self.sender()

        if radiobutton.isChecked():
            self.case = radiobutton.case

    def exec_(self):
        super().exec_()

        return_string = None

        if self.case == 1:
            return_string = self.first_option
        if self.case == 2:
            return_string = self.second_option
        if self.case == 3:
            return_string = self.third_option.text()

        self.close()
        return return_string


class HCRPDCEntry(QDialog):

    def __init__(self, first_option, check1, second_option, check2, third_option, check3, file_name, parent=None):
        super(HCRPDCEntry, self).__init__(parent)

        self.setWindowTitle("Enter HCR P&DC Name")

        self.file_name = file_name.split("/")[-1]

        self.first_option = first_option
        self.second_option = second_option
        self.third_option = third_option
        self.fourth_option = QLineEdit()

        self.check1 = check1
        self.check2 = check2
        self.check3 = check3

        self.radio_one = QRadioButton()
        self.radio_two = QRadioButton()
        self.radio_three = QRadioButton()
        self.radio_four = QRadioButton()

        self.case = 1

        self.radio_one.case = 1
        self.radio_two.case = 2
        self.radio_three.case = 3
        self.radio_four.case = 4

        self.radio_one.toggled.connect(self.change_case)
        self.radio_two.toggled.connect(self.change_case)
        self.radio_three.toggled.connect(self.change_case)
        self.radio_four.toggled.connect(self.change_case)

        self.radio_one.setChecked(True)

        self.add_to_lookups_check = QCheckBox("Add To Lookups?")

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        # self.cancel_button = QPushButton("Cancel", self)
        # self.cancel_button.clicked[bool].connect(self.cancel)

        self.initUI()

    def initUI(self):

        main_layout = QVBoxLayout()
        inner_grid = QGridLayout()

        header_label = QLabel("Please input HCR P&DC name for: " + self.file_name)

        if self.first_option != self.second_option:
            inner_grid.addWidget(QLabel(self.first_option), 0, 0)
            inner_grid.addWidget(self.radio_one, 0, 1)
            inner_grid.addWidget(self.quick_label(self.check1), 0, 2)

        inner_grid.addWidget(QLabel(self.second_option), 1, 0)
        inner_grid.addWidget(QLabel(self.third_option), 2, 0)
        inner_grid.addWidget(self.fourth_option, 3, 0)

        inner_grid.addWidget(self.radio_two, 1, 1)
        inner_grid.addWidget(self.radio_three, 2, 1)
        inner_grid.addWidget(self.radio_four, 3, 1)

        inner_grid.addWidget(self.quick_label(self.check2), 1, 2)
        inner_grid.addWidget(self.quick_label(self.check3), 2, 2)

        main_layout.addWidget(header_label)
        main_layout.addLayout(inner_grid)
        main_layout.addWidget(self.add_to_lookups_check)
        main_layout.addWidget(self.accept_button)
        self.setLayout(main_layout)

    @staticmethod
    def quick_label(check):

        if check:
            return QLabel("In Lookups")
        else:
            return QLabel("")

    def change_case(self):

        radiobutton = self.sender()

        if radiobutton.isChecked():
            self.case = radiobutton.case

    def exec_(self):
        super().exec_()

        return_string = None

        if self.case == 1:
            return_string = self.first_option
        if self.case == 2:
            return_string = self.second_option
        if self.case == 3:
            return_string = self.third_option
        if self.case == 4:
            return_string = self.fourth_option.text()

        self.close()
        return return_string, self.add_to_lookups_check.isChecked()


class AddressSelector(QDialog):

    def __init__(self, plate_num, stop_name, addresses, parent=None):
        super(AddressSelector, self).__init__(parent)

        self.setWindowTitle(plate_num)

        self.stop_name = stop_name
        self.addresses = addresses

        self.button_list = []

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        self.initUI()

    def initUI(self):

        vbox = QVBoxLayout()
        grid = QGridLayout()

        header = QLabel("Pick Correct Address for: " + self.stop_name)

        for x, address in enumerate(self.addresses):
            label = QLabel(address[1])
            rb = QRadioButton()
            self.button_list.append(rb)
            grid.addWidget(label, x, 0)
            grid.addWidget(rb, x, 1)

        vbox.addWidget(header)
        vbox.addLayout(grid)
        vbox.addWidget(self.accept_button)

        self.setLayout(vbox)

    def exec_(self):
        super().exec_()

        return_index = 0
        for x, button in enumerate(self.button_list):
            if button.isChecked():
                return_index = x
        self.close()

        return self.addresses[return_index]


class MergeTwoUParts(QDialog):

    def __init__(self, u_part_1, u_part_2, layover, plate_number, parent=None):
        super(MergeTwoUParts, self).__init__(parent)

        self.u_part_1 = u_part_1
        self.u_part_2 = u_part_2
        self.layover = layover
        self.combined_duration = u_part_1.raw_duration() + u_part_2.raw_duration() + layover

        self.setWindowTitle(str(plate_number) + " : Verify Schedule Combination")
        self.setGeometry(300, 150, 650, 400)

        self.initUI()

        self.return_bool = None
        self.no_button = None
        self.yes_button = None

    def initUI(self):

        vbox = QVBoxLayout()
        vbox.addWidget(QLabel("Are these one schedule?"))
        vbox.addWidget(QLabel("(Layover: " + str(self.layover) + " minutes)"))
        vbox.addWidget(QLabel("(Combined Duration: " + str(self.combined_duration) + " minutes)"))
        hbox = QHBoxLayout()

        left_vbox = QVBoxLayout()
        left_grid = QGridLayout()
        left_grid.addWidget(QLabel("Schedule Part: "), 0, 0)
        left_grid.addWidget(QLabel("Trip Number: "), 1, 0)
        left_grid.addWidget(QLabel("Frequency Code: "), 2, 0)
        left_grid.addWidget(QLabel("Description: "), 3, 0)

        left_part = self.u_part_1.part
        left_num = str(self.u_part_1.schedule_num)
        left_freq = self.u_part_1.freq_code
        left_desc = self.u_part_1.freq_code_desc

        left_grid.addWidget(QLabel(left_part), 0, 1)
        left_grid.addWidget(QLabel(left_num), 1, 1)
        left_grid.addWidget(QLabel(left_freq), 2, 1)
        left_grid.addWidget(QLabel(left_desc), 3, 1)
        left_vbox.addLayout(left_grid)

        left_table = UPartView()
        left_table.set_model(UPartToPandas(self.u_part_1).df)
        left_vbox.addWidget(left_table)

        hbox.addLayout(left_vbox)

        right_vbox = QVBoxLayout()
        right_grid = QGridLayout()
        right_grid.addWidget(QLabel("Schedule Part: "), 0, 0)
        right_grid.addWidget(QLabel("Trip Number: "), 1, 0)
        right_grid.addWidget(QLabel("Frequency Code: "), 2, 0)
        right_grid.addWidget(QLabel("Description: "), 3, 0)

        right_part = self.u_part_2.part
        right_num = str(self.u_part_2.schedule_num)
        right_freq = self.u_part_2.freq_code
        right_desc = self.u_part_2.freq_code_desc

        right_grid.addWidget(QLabel(right_part), 0, 1)
        right_grid.addWidget(QLabel(right_num), 1, 1)
        right_grid.addWidget(QLabel(right_freq), 2, 1)
        right_grid.addWidget(QLabel(right_desc), 3, 1)
        right_vbox.addLayout(right_grid)

        right_table = UPartView()
        right_table.set_model(UPartToPandas(self.u_part_2).df)
        right_vbox.addWidget(right_table)

        hbox.addLayout(right_vbox)

        hbox2 = QHBoxLayout()
        self.no_button = QPushButton("No", self)
        self.no_button.clicked[bool].connect(self.no)
        hbox2.addWidget(self.no_button)
        self.yes_button = QPushButton("Yes", self)
        self.yes_button.clicked[bool].connect(self.yes)
        hbox2.addWidget(self.yes_button)

        vbox.addLayout(hbox)
        vbox.addLayout(hbox2)
        self.setLayout(vbox)

    def yes(self):
        self.return_bool = True
        self.accept()

    def no(self):
        self.return_bool = False
        self.accept()

    def exec_(self):
        super().exec_()
        self.close()
        return self.return_bool


class UPartModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        if not self._data[0][0]:
            return 0
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role == Qt.DisplayRole:
            value = self._data[index.row()][index.column()]
            return str(value)
        else:
            return None

    def headerData(self, col, orientation, role):

        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]

        return None

    def flags(self, index):

        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class UPartView(QTableView):

    def __init__(self, parent=None):
        super().__init__()

    def set_model(self, data):

        self.setModel(UPartModel(data))
        self.resizeColumnsToContents()


class UPartToPandas:

    def __init__(self, u_part):

        labels = ["Stop", "Stop Name", "Tag", "Time"]

        compiled = []

        for x, stop in enumerate(u_part.half_stops):
            compiled.append([str(x+1), stop.stop_name, stop.tag, str(stop.time)])

        self.df = pd.DataFrame(compiled, columns=labels)


class CustomHCRReaderPopUp(QDialog):

    def __init__(self, possible_names, possible_addresses, lb, parent=None):
        super(CustomHCRReaderPopUp, self).__init__(parent)

        self.setWindowTitle("Select P&DC For Plate")
        # self.setGeometry(300, 200, 800, 400)

        self.possible_names = possible_names
        self.possible_addresses = possible_addresses
        self.hcr_pdcs = lb.hcr_pdcs
        self.lb = lb

        self.return_string = None

        self.found_on_plate_selection = QComboBox()
        self.found_on_plate_selection.addItem("Input Other Location")
        for item in self.possible_names:
            self.found_on_plate_selection.addItem(item)

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        self.initUI()

    def initUI(self):

        outer_layout = QVBoxLayout()
        header = QHBoxLayout()
        main_bar = QHBoxLayout()

        header.addWidget(QLabel("Select P&DC for HCR Plate"))

        main_bar.addWidget(QLabel("Found on plate:"))
        main_bar.addWidget(self.found_on_plate_selection)

        outer_layout.addLayout(header)
        outer_layout.addLayout(main_bar)
        outer_layout.addWidget(self.accept_button)

        self.setLayout(outer_layout)

    def exec_(self):

        super().exec_()
        self.close()

        index = self.found_on_plate_selection.currentIndex()
        if index == 0:
            return "New", ""

        name = self.found_on_plate_selection.currentText()
        address = self.possible_addresses[self.possible_names.index(name)]

        return name, address