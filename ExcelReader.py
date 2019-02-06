
from openpyxl import load_workbook
from ScheduleClasses import Schedule, Stop
from AddressCompilation import NewAddressBook, ExistingAddressBook
import os


def excel_time_to_time(var):

    pass


class JDAExcelReader:

    def __init__(self, file_name, lb):

        self.file_name = file_name
        self.lb = lb

        self.source = "JDA"
        self.source_type = "EXCEL"
        self.source_file = file_name.split("/")[-1][:-5]

        self.wb = load_workbook(self.file_name)
        ws = self.wb["Summary"]

        self.pvs_name = ws["C7"].value

        self.pdc_name = ws["C12"].value
        self.pdc_address = None
        self.new_pvs_name = None
        self.new_pdc_name = None
        self.short_name = None
        self.jda_facility_name = None
        self.jda_nass_code = None
        self.address_rows = []
        self.find_names()
        self.read_in_locations()

        self.schedules = []
        self.read_in_schedules()
        self.add_addresses()

    def find_names(self):

        index = None
        if self.pdc_name in self.lb.pvs_pdcs:
            index = self.lb.pvs_pdcs.index(self.pdc_name)
        elif self.pdc_name in self.lb.hcr_pdcs:
            index = self.lb.hcr_pdcs.index(self.pdc_name)
        elif self.pdc_name in self.lb.alternate_pdc_names:
            index = self.lb.alternate_pdc_names.index(self.pdc_name)
        elif self.pvs_name != "PVS":
            if self.pvs_name in self.lb.alternate_pvs_names:
                index = self.lb.alternate_pvs_names.index(self.pvs_name)
            elif self.pvs_name in self.lb.pvs_names:
                index = self.lb.pvs_names.index(self.pvs_name)

        if not index:
            print("Couldn't find facility in lookup book!")
            return

        else:
            self.new_pvs_name = self.lb.pvs_names[index]
            self.new_pdc_name = self.lb.pvs_pdcs[index]
            self.short_name = self.lb.short_names[index]
            self.jda_facility_name = self.lb.jda_facility_names[index]

    def read_in_schedules(self):

        ws = self.wb["Service Point Summary"]

        x = 7
        max_row = ws.max_row

        while x <= max_row:
            schedule_num = ws["C" + str(x)].value
            if schedule_num in (None, "", " "):
                break
            freq_code = str(ws["F" + str(x)].value)
            vehicle_type = ws["H" + str(x)].value
            stops = []
            mileage = 0
            while ws['C' + str(x)].value == schedule_num:
                vehicle_type = ws['H' + str(x)].value
                arrive_time = ws['L' + str(x)].value
                depart_time = ws['N' + str(x)].value
                stop_name = ws['M' + str(x)].value
                nass_code = str(ws['S' + str(x)].value)
                mileage += float(ws['T' + str(x)].value)
                if stop_name == self.pvs_name:
                    stop_name = self.new_pvs_name
                if stop_name in (self.jda_facility_name, self.pdc_name):
                    stop_name = self.new_pdc_name
                load_num = ws["U" + str(x)].value
                stops.append(Stop(arrive_time=arrive_time, depart_time=depart_time,
                                  stop_name=stop_name, vehicle_type=vehicle_type, nass_code=nass_code,
                                  load_num=load_num))
                x += 1

            stops = self.check_stops(stops)

            new_schedule = Schedule(stops=stops, vehicle_category=vehicle_type, vehicle_type=vehicle_type,
                                    schedule_num=schedule_num, freq_code=freq_code, source=self.source,
                                    source_type=self.source_type, source_file=self.source_file,
                                    pvs_name=self.new_pvs_name, short_name=self.short_name,
                                    pvs_pdc=self.new_pdc_name, mileage=mileage)

            new_schedule.set_bin_string(self.lb)
            new_schedule.is_holiday_schedule()

            self.schedules.append(new_schedule)

    def read_in_locations(self):

        try:
            ws = self.wb["Location"]
        except:
            print("No Location sheet found")
            return

        max_row = ws.max_row
        row = 2
        temp_address_rows = []

        while row <= max_row:
            new_row = []
            if ws["A" + str(row)].value in (None, "", " "):
                break
            for col in range(1, 11):
                new_row.append(ws.cell(row=row, column=col).value)
            temp_address_rows.append(new_row)
            row += 1

        self.address_rows = temp_address_rows

    def find_address(self, nass_code):

        if nass_code == self.jda_facility_name:
            possible_rows = [x for x in self.address_rows if x[1] == nass_code]
        else:
            possible_rows = [x for x in self.address_rows if str(x[0]) == nass_code]

        if not possible_rows:
            return None
        elif len(possible_rows) > 1:
            print("multiple addresses found for " + nass_code)
            to_use = possible_rows[0]
        else:
            to_use = possible_rows[0]

        address_string = str(to_use[7]) + " " + str(to_use[8]) + " " + str(to_use[2]) + " " + str(to_use[3])

        if to_use[7] in (None, "", " "):
            address_string = str(to_use[5])

        return address_string

    def add_addresses(self):

        pdc_address = self.find_address(self.jda_facility_name)
        self.pdc_address = pdc_address

        for schedule in self.schedules:
            schedule.pdc_address = pdc_address

        for schedule in self.schedules:
            for stop in schedule.stops:
                if stop.stop_name in (self.new_pdc_name, self.pdc_name, self.pvs_name):
                    address = self.pdc_address
                else:
                    address = self.find_address(stop.nass_code)

                if address:
                    stop.add_address(address)
                else:
                    print("Couldn't find address for " + stop.stop_name)

    @staticmethod
    def check_stops(stops):

        if stops[0].stop_name == stops[1].stop_name:
            if stops[0].layover() < 1:
                return stops[1:]

        return stops