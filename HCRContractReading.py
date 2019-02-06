from openpyxl import load_workbook, styles
import math
import pandas as pd
import csv
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant, QSortFilterProxyModel
from PyQt5.QtWidgets import QTableView
from GeneralMethods import number
from datetime import datetime


class HCRContractReader:

    def __init__(self, file_name):
        print("reading HCR contracts from " + file_name)
        self.file_name = file_name
        self.segments = None
        self.contracts_df = None
        self.contracts = []
        if file_name[-3:] != 'csv':
            print("converting file to csv")
            self.convert_to_csv()
        print("reading in file")
        self.as_of_date = None
        self.pandas_in()
        print("populating contracts list")
        self.populate_contracts()

    def convert_to_csv(self):

        wb = load_workbook(filename=self.file_name, read_only=True)
        sheet_names = wb.get_sheet_names()
        check_string = "Raw Data"
        ws = None

        for sheet_name in sheet_names:
            if check_string in sheet_name:
                ws = wb[sheet_name]

        if not ws:
            print("Couldn't find Raw Data sheet")
            return

        with open('hcr_contracts.csv', 'w', newline="") as f:
            c = csv.writer(f)
            for r in ws.rows:
                c.writerow([cell.value for cell in r])

    def pandas_in(self):

        self.segments = pd.read_csv('hcr_contracts.csv', encoding="ISO-8859-1")

        self.as_of_date = datetime.strptime(self.segments.values[0][0], "%Y-%m-%d %H:%M:%S").date()
        print(self.as_of_date)

        group_by = ['DN NAME', 'CONTRACT ID', 'BEGIN SERVICE POINT', 'BEGIN DATE', 'END DATE']

        df1 = self.segments.groupby(group_by)['ANNUAL HOURS', 'ANNUAL MILES', 'ANNUAL COST'].sum().reset_index()

        df1['Contract ID'] = df1['CONTRACT ID']
        df1['Begin Service Point'] = df1['BEGIN SERVICE POINT']
        df1['BEGIN DATE'] = pd.to_datetime(df1['BEGIN DATE'], infer_datetime_format=True)
        df1['Begin Date'] = df1['BEGIN DATE'].dt.strftime('%d/%m/%Y')
        df1['END DATE'] = pd.to_datetime(df1['END DATE'], infer_datetime_format=True)
        df1['End Date'] = df1['END DATE'].dt.strftime('%d/%m/%Y')

        df1['Ann. Hours'] = df1['ANNUAL HOURS'].map('{:,.0f}'.format)
        df1['Ann. Miles'] = df1['ANNUAL MILES'].map('{:,.0f}'.format)
        df1['Ann. Rate'] = df1['ANNUAL COST'].map('{:,.0f}'.format)

        to_drop = ['ANNUAL HOURS', 'ANNUAL MILES', 'ANNUAL COST', 'BEGIN SERVICE POINT', 'BEGIN DATE', 'END DATE',
                   'CONTRACT ID']

        df1 = df1.drop(to_drop, axis=1)
        df2 = self.loc_cd_to_area_df()

        self.contracts_df = pd.merge(df1, df2, on='DN NAME', how='left')
        self.contracts_df = self.contracts_df.drop(['DN NAME'], axis=1)

    @staticmethod
    def loc_cd_to_area_df():

        # fin_dns = ['1B', '1C', '1E', '1F', '1G', '1J', '1K', '7R', '7Y', '2E', '1H']
        # areas = ['NE', 'EA', 'WE', 'PA', 'SA', 'GL', 'CM', 'HQ', 'HQ', 'WE', 'SA']
        DN_NAMES = ['Northeast', 'Eastern', 'Cap Metro', 'Southern', 'Great Lakes', 'Western', 'Pacific']
        areas = ['NE', 'EA', 'CM', 'SA', 'GL', 'WE', 'PA']

        df = pd.DataFrame({'DN NAME': DN_NAMES, 'Area': areas})

        return df

    def get_contract(self, hcr_id):

        row = [x for x in self.contracts if x.hcr_id.lower() == hcr_id.lower()]
        if len(row) > 1:
            print("more than one contract found??")
        elif len(row) < 1:
            print("contract not found")
            return None
        else:
            print("contract found")
            return row[0]

    def populate_contracts(self):

        data = self.contracts_df.values.tolist()
        for contract in data:
            hcr_id = contract[0]
            area = contract[-1]
            origin_terminus = contract[1]
            begin_date = contract[2]
            end_date = contract[3]
            hours = float(contract[4].replace(',',''))
            miles = float(contract[5].replace(',',''))
            rate = float(contract[6].replace(',',''))
            self.contracts.append(HCRContract(hcr_id, area, origin_terminus, begin_date, end_date, hours, miles, rate))
        print(len(self.contracts))


class HCRContract:

    def __init__(self, hcr_id, area, origin_terminus, begin_date, end_date, total_annual_hours,
                 total_annual_miles, total_annual_rate):

        while len(hcr_id)<5:
            hcr_id = "0" + hcr_id

        self.hcr_id = hcr_id
        self.segments = None
        self.area = area
        self.origin_terminus = origin_terminus
        self.begin_date = begin_date
        self.end_date = end_date
        self.total_annual_hours = total_annual_hours
        self.total_annual_miles = total_annual_miles
        self.total_annual_rate = total_annual_rate

        self.usps_annual_hours = None
        self.usps_annual_miles = None
        self.usps_paid_hours = None

        self.van_miles = 0
        self.van_miles_postalized = 0
        self.tractor_sa_miles = 0
        self.tractor_sa_miles_postalized = 0
        self.tractor_ta_miles = 0
        self.tractor_ta_miles_postalized = 0

        self.ft_hours = None
        self.ft_hours_postalized = None
        self.pse_hours = None
        self.pse_hours_postalized = None
        self.ft_night_diff_hours = None
        self.pse_night_diff_hours = None
        self.total_night_diff_hours_calculated = None
        self.total_night_diff_hours_postalized = None
        self.ft_night_diff_hours_postalized = None
        self.pse_night_diff_hours_postalized = None

        self.ft_ot_hours = None
        self.ft_ot_hours_postalized = None
        self.pse_ot_hours = None
        self.pse_ot_hours_postalized = None

        self.ft_emps = None
        self.pse_emps = None
        self.req_managers = None
        self.total_emps = None

        self.ft_emps_postalized = None
        self.pse_emps_postalized = None
        self.req_managers_postalized = None
        self.total_emps_postalized = None

        self.wtd_ft_wage_rate = None
        self.wtd_pse_wage_rate = None

        self.ft_costs = None
        self.ft_costs_postalized = None
        self.ft_night_diff_costs = None
        self.ft_night_diff_costs_postalized = None
        self.ft_ot_costs = None
        self.ft_ot_costs_postalized = None
        self.pse_costs = None
        self.pse_costs_postalized = None
        self.pse_night_diff_costs = None
        self.pse_night_diff_costs_postalized = None
        self.pse_ot_costs = None
        self.pse_ot_costs_postalized = None

        self.manager_costs = None
        self.manager_costs_postalized = None

        self.year_one_recruitment_cost = None
        self.year_one_recruitment_cost_postalized = None
        self.recurring_recruitment_cost = None
        self.recurring_recruitment_cost_postalized = None

        self.total_labor_cost = None
        self.total_labor_cost_postalized = None

        self.van_count = None
        self.tractor_sa_count = None
        self.tractor_ta_count = None
        self.trailer_count = None

        self.one_van_lease_cost = None
        self.one_tractor_sa_lease_cost = None
        self.one_tractor_ta_lease_cost = None
        self.one_trailer_lease_cost = None

        self.van_fuel_cost = None
        self.tractor_sa_fuel_cost = None
        self.tractor_ta_fuel_cost = None
        self.total_fuel_cost = None

        self.van_cpm = None
        self.tractor_sa_cpm = None
        self.tractor_ta_cpm = None

        self.cargo_van_ops = None
        self.cargo_van_ops_postalized = None
        self.tractor_sa_ops = None
        self.tractor_sa_ops_postalized = None
        self.tractor_ta_ops = None
        self.tractor_ta_ops_postalized = None
        self.trailer_ops = None
        self.total_ops = None
        self.total_ops_postalized = None

        self.torts_per_mile = None
        self.torts = None
        self.torts_postalized = None
        self.tolls = 0
        self.tolls_postalized = 0

        self.total_owned_transportation_cost = None
        self.total_owned_transportation_cost_postalized = None
        self.total_leased_transportation_cost = None
        self.total_leased_transportation_cost_postalized = None

        self.total_acq_cost_wo_dep = None
        self.total_acq_cost_w_dep = None

        self.van_lease_cost = None
        self.van_lease_cost_postalized = None
        self.van_base_lease_cost = None
        self.van_extra_lease_cost = None
        self.van_extra_lease_cost_postalized = None
        self.tractor_sa_lease_cost = None
        self.tractor_sa_lease_cost_postalized = None
        self.tractor_sa_base_lease_cost = None
        self.tractor_sa_extra_lease_cost = None
        self.tractor_sa_extra_lease_cost_postalized = None
        self.tractor_ta_lease_cost = None
        self.tractor_ta_lease_cost_postalized = None
        self.tractor_ta_base_lease_cost = None
        self.tractor_ta_extra_lease_cost = None
        self.tractor_ta_extra_lease_cost_postalized = None
        self.trailer_base_lease_cost = None
        self.trailer_lease_cost = None

        self.total_base_lease_cost = None
        self.total_extra_lease_cost = None
        self.total_extra_lease_cost_postalized = None
        self.total_lease_cost = None
        self.total_lease_cost_postalized = None

        self.annual_maint_per_trailer = None

        self.cargo_van_annual_fuel = None
        self.tractor_sa_annual_fuel = None
        self.tractor_ta_annual_fuel = None

        self.cargo_van_acq_cost = None
        self.cargo_van_dep_life = None
        self.cargo_van_dep_cost = None
        self.tractor_sa_acq_cost = None
        self.tractor_sa_dep_life = None
        self.tractor_sa_dep_cost = None
        self.tractor_ta_acq_cost = None
        self.tractor_ta_dep_life = None
        self.tractor_ta_dep_cost = None
        self.trailer_acq_cost = None

        self.total_cost_w_dep = None
        self.total_cost_w_dep_postalized = None
        self.total_cost_wo_dep = None
        self.total_cost_wo_dep_postalized = None
        self.total_cost_w_lease = None
        self.total_cost_w_lease_postalized = None

        self.vehicle_contract = None
        self.total_calculated_duration = None
        self.total_postalized_duration = None
        self.paid_postalized_hours = None
        self.total_calculated_mileage = None
        self.total_postalized_mileage = None

    def print_contract(self, ws, row):

        ws["A" + str(row)].value = str(row - 7)
        ws["B" + str(row)].value = self.hcr_id
        ws["C" + str(row)].value = self.origin_terminus
        ws["D" + str(row)].value = self.area
        ws["E" + str(row)].value = self.begin_date
        ws["F" + str(row)].value = self.end_date
        ws["G" + str(row)].value = self.total_annual_miles
        ws["H" + str(row)].value = self.total_annual_hours
        ws["I" + str(row)].value = self.total_annual_rate

    def add_night_diff(self, postalized_night_diff, calculated_night_diff):

        self.total_night_diff_hours_postalized = postalized_night_diff
        self.total_night_diff_hours_calculated = calculated_night_diff

    def usps_equivalents_calc(self, main_inflation, vac_sick_inflation):

        main_inflation += 1
        vac_sick_inflation += 1

        self.usps_annual_hours = self.total_annual_hours*main_inflation
        self.usps_annual_miles = self.total_annual_miles*main_inflation
        self.usps_paid_hours = self.usps_annual_hours*vac_sick_inflation
        self.paid_postalized_hours = self.total_postalized_duration * vac_sick_inflation

    def usps_labor_hours_calc(self, ft_rate, pse_rate, night_diff_rate, ot_rate):

        self.ft_hours = self.usps_paid_hours * ft_rate
        self.ft_hours_postalized = self.paid_postalized_hours * ft_rate

        self.pse_hours = self.usps_paid_hours * pse_rate
        self.pse_hours_postalized = self.paid_postalized_hours * pse_rate

        self.ft_night_diff_hours = self.ft_hours * night_diff_rate
        self.ft_night_diff_hours_postalized = self.total_night_diff_hours_postalized * ft_rate

        self.pse_night_diff_hours = self.pse_hours * night_diff_rate
        self.pse_night_diff_hours_postalized = self.total_night_diff_hours_postalized * pse_rate

        self.ft_ot_hours = self.ft_hours * ot_rate
        self.ft_ot_hours_postalized = self.ft_hours_postalized * ot_rate

        self.pse_ot_hours = self.pse_hours * ot_rate
        self.pse_ot_hours_postalized = self.pse_hours_postalized * ot_rate

    def usps_employees_calc(self, ft_hours, pse_hours, req_man):

        self.ft_emps = math.ceil(self.ft_hours/ft_hours)
        self.pse_emps = math.ceil(self.pse_hours/pse_hours)
        self.req_managers = (self.ft_emps + self.pse_emps)/req_man
        self.total_emps = self.ft_emps + self.pse_emps + self.req_managers

        self.ft_emps_postalized = math.ceil(self.ft_hours_postalized/ft_hours)
        self.pse_emps_postalized = math.ceil(self.pse_hours_postalized/pse_hours)
        self.req_managers_postalized = (self.ft_emps_postalized + self.pse_emps_postalized)/req_man
        self.total_emps_postalized = self.ft_emps_postalized + self.pse_emps_postalized + self.req_managers_postalized

    def usps_ft_costs_calc(self, vans, tractors, van_rate, tractor_rate, hours):

        self.ft_costs = (self.ft_emps * hours * vans * van_rate) + (self.ft_emps * hours * tractors * tractor_rate)
        self.ft_costs_postalized = (self.ft_emps_postalized * hours * vans * van_rate) + \
                                   (self.ft_emps_postalized * hours * tractors * tractor_rate)

    def usps_ft_ot_costs_calc(self, vans, tractors, ot_van_rate, ot_tractor_rate):

        self.ft_ot_costs = (self.ft_ot_hours * vans * ot_van_rate) + (self.ft_ot_hours * tractors * ot_tractor_rate)
        self.ft_ot_costs_postalized = (self.ft_ot_hours_postalized * vans * ot_van_rate) + \
                                      (self.ft_ot_hours_postalized * tractors * ot_tractor_rate)

    def usps_pse_costs_calc(self, vans, tractors, van_rate, tractor_rate):

        self.pse_costs = (self.pse_hours * vans * van_rate) + (self.pse_hours * tractors * tractor_rate)
        self.pse_costs_postalized = (self.pse_hours_postalized * vans * van_rate) + \
                                    (self.pse_hours_postalized * tractors * tractor_rate)

    def usps_pse_ot_costs_calc(self, vans, tractors, ot_van_rate, ot_tractor_rate):

        self.pse_ot_costs = (self.pse_ot_hours * vans * ot_van_rate) + (self.pse_ot_hours * tractors * ot_tractor_rate)
        self.pse_ot_costs_postalized = (self.pse_ot_hours_postalized * vans * ot_van_rate) + \
                                       (self.pse_ot_hours_postalized * tractors * ot_tractor_rate)

    def usps_night_diff_calc(self, night_diff):

        self.ft_night_diff_costs = self.ft_night_diff_hours * night_diff
        self.ft_night_diff_costs_postalized = self.ft_night_diff_hours_postalized * night_diff
        self.pse_night_diff_costs = self.pse_night_diff_hours * night_diff
        self.pse_night_diff_costs_postalized = self.pse_night_diff_hours_postalized * night_diff

    def usps_management_costs_calc(self, manager_salary):

        self.manager_costs = self.req_managers * manager_salary
        self.manager_costs_postalized = self.req_managers_postalized * manager_salary

    def usps_attrition_costs_calc(self, rec_cost, attrition):

        self.year_one_recruitment_cost = math.ceil(self.total_emps)*rec_cost
        self.recurring_recruitment_cost = math.ceil(self.pse_emps * attrition)*rec_cost

        self.year_one_recruitment_cost_postalized = math.ceil(self.total_emps_postalized) * rec_cost
        self.recurring_recruitment_cost_postalized = math.ceil(self.pse_emps_postalized * attrition) * rec_cost

    def usps_total_labor_cost_calc(self):

        self.total_labor_cost = (self.ft_costs + self.ft_ot_costs + self.ft_night_diff_costs +
                                 self.pse_costs + self.pse_ot_costs + self.pse_night_diff_costs +
                                 self.manager_costs + self.year_one_recruitment_cost +
                                 self.recurring_recruitment_cost)
        self.total_labor_cost_postalized = (self.ft_costs_postalized + self.ft_ot_costs_postalized +
                                            self.ft_night_diff_costs_postalized + self.pse_costs_postalized +
                                            self.pse_ot_costs_postalized + self.pse_night_diff_costs_postalized +
                                            self.manager_costs_postalized + self.year_one_recruitment_cost_postalized +
                                            self.recurring_recruitment_cost_postalized)

    def total_transportation_cost_calc(self):

        self.total_owned_transportation_cost = self.total_fuel_cost + self.total_ops + self.torts + self.tolls
        self.total_leased_transportation_cost = self.total_fuel_cost + self.torts + self.tolls

        self.total_owned_transportation_cost_postalized = self.total_fuel_cost + self.total_ops_postalized + \
                                                          self.torts_postalized + self.tolls_postalized
        self.total_leased_transportation_cost_postalized = self.total_fuel_cost + self.torts_postalized + \
                                                           self.tolls_postalized

    def van_lease_calc(self, base, mileage_max, mileage_extra):

        self.one_van_lease_cost = base

        van_count = self.van_count
        van_miles = self.van_miles
        van_miles_postalized = self.van_miles_postalized

        van_extra_miles = max(0 , van_miles - (mileage_max * van_count))
        van_extra_miles_postalized = max(0, van_miles_postalized - (mileage_max * van_count))

        self.van_base_lease_cost = (van_count * base)
        self.van_extra_lease_cost = (van_extra_miles * mileage_extra)
        self.van_extra_lease_cost_postalized = (van_extra_miles_postalized * mileage_extra)

        self.van_lease_cost = self.van_base_lease_cost + self.van_extra_lease_cost
        self.van_lease_cost_postalized = self.van_base_lease_cost + self.van_extra_lease_cost_postalized

    def tractor_sa_lease_calc(self, base, mileage_max, mileage_extra):

        self.one_tractor_sa_lease_cost = base

        tractor_sa_count = self.tractor_sa_count
        tractor_sa_miles = self.tractor_sa_miles
        tractor_sa_miles_postalized = self.tractor_sa_miles_postalized

        tractor_sa_extra_miles = max(0, tractor_sa_miles - (mileage_max * tractor_sa_count))
        tractor_sa_extra_miles_postalized = max(0, tractor_sa_miles_postalized - (mileage_max * tractor_sa_count))

        self.tractor_sa_base_lease_cost = (tractor_sa_count * base)
        self.tractor_sa_extra_lease_cost = (tractor_sa_extra_miles * mileage_extra)
        self.tractor_sa_extra_lease_cost_postalized = (tractor_sa_extra_miles_postalized * mileage_extra)

        self.tractor_sa_lease_cost = self.tractor_sa_base_lease_cost + self.tractor_sa_extra_lease_cost
        self.tractor_sa_lease_cost_postalized = self.tractor_sa_base_lease_cost + \
                                                self.tractor_sa_extra_lease_cost_postalized

    def tractor_ta_lease_calc(self, base, mileage_max, mileage_extra):

        self.one_tractor_ta_lease_cost = base

        tractor_ta_count = self.tractor_ta_count
        tractor_ta_miles = self.tractor_ta_miles
        tractor_ta_miles_postalized = self.tractor_ta_miles_postalized

        tractor_ta_extra_miles = max(0, tractor_ta_miles - (mileage_max * tractor_ta_count))
        tractor_ta_extra_miles_postalized = max(0, tractor_ta_miles_postalized - (mileage_max * tractor_ta_count))

        self.tractor_ta_base_lease_cost = (tractor_ta_count * base)
        self.tractor_ta_extra_lease_cost = (tractor_ta_extra_miles * mileage_extra)
        self.tractor_ta_extra_lease_cost_postalized = (tractor_ta_extra_miles_postalized * mileage_extra)

        self.tractor_ta_lease_cost = self.tractor_ta_base_lease_cost + self.tractor_ta_extra_lease_cost
        self.tractor_ta_lease_cost_postalized = self.tractor_ta_base_lease_cost + \
                                                self.tractor_ta_extra_lease_cost_postalized

    def trailer_lease_cost_calc(self, base):

        self.trailer_base_lease_cost = (self.trailer_count * base)
        self.trailer_lease_cost = self.trailer_base_lease_cost

    def total_lease_cost_calc(self):

        self.total_lease_cost = (self.van_lease_cost + self.tractor_sa_lease_cost + self.tractor_ta_lease_cost +
                                 self.trailer_lease_cost)

        self.total_lease_cost_postalized = (self.van_lease_cost_postalized + self.tractor_sa_lease_cost_postalized +
                                            self.tractor_ta_lease_cost_postalized + self.trailer_lease_cost)

        self.total_base_lease_cost = (self.van_base_lease_cost + self.tractor_sa_base_lease_cost +
                                      self.tractor_ta_base_lease_cost + self.trailer_base_lease_cost)

        self.total_extra_lease_cost = (self.van_extra_lease_cost + self.tractor_sa_extra_lease_cost +
                                       self.tractor_ta_extra_lease_cost)
        self.total_extra_lease_cost_postalized = (self.van_extra_lease_cost_postalized +
                                                  self.tractor_sa_extra_lease_cost_postalized +
                                                  self.tractor_ta_extra_lease_cost_postalized)

    def print_one_summary_estimated(self, ws, scenario):

        ws['A1'].value = "HCR Contract Being Compared: " + self.origin_terminus
        ws['B1'].value = self.hcr_id
        ws['D1'].value = self.area
        ws['B3'].value = self.total_annual_hours
        ws['B4'].value = self.total_annual_miles
        ws['B5'].value = self.total_annual_rate
        ws['B6'].value = self.van_count
        ws['B7'].value = self.tractor_sa_count
        ws['B8'].value = self.tractor_ta_count
        ws['B9'].value = self.trailer_count

        ws['A12'].value = "Total Driving Hours: " + str(number(self.usps_paid_hours))
        ws['B13'].value = self.wtd_ft_wage_rate
        ws['B14'].value = self.wtd_pse_wage_rate
        ws['D13'].value = self.ft_costs
        ws['D14'].value = self.pse_costs
        ws['D15'].value = self.manager_costs
        ws['D16'].value = self.ft_ot_costs + self.pse_ot_costs
        ws['D17'].value = self.year_one_recruitment_cost + self.recurring_recruitment_cost
        ws['D18'].value = self.total_labor_cost

        ws['A20'].value = "Total PVS Miles: " + str(number(self.usps_annual_miles))

        if scenario in (1, 2):
            ws['B21'].value = self.van_cpm
            ws['B22'].value = self.tractor_sa_cpm
            ws['B23'].value = self.tractor_ta_cpm
            ws['B24'].value = self.torts_per_mile
            ws['B25'].value = self.annual_maint_per_trailer
            ws['B26'].value = self.tolls
            ws['B27'].value = self.cargo_van_annual_fuel
            ws['B28'].value = self.tractor_sa_annual_fuel
            ws['B29'].value = self.tractor_ta_annual_fuel

            ws['C21'].value = self.van_miles
            ws['C22'].value = self.tractor_sa_miles
            ws['C23'].value = self.tractor_ta_miles
            ws['C24'].value = self.usps_annual_miles
            ws['C25'].value = self.trailer_count
            ws['C26'].value = self.tolls
            ws['C27'].value = self.van_count
            ws['C28'].value = self.tractor_sa_count
            ws['C29'].value = self.tractor_ta_count

            ws['D21'].value = self.cargo_van_ops
            ws['D22'].value = self.tractor_sa_ops
            ws['D23'].value = self.tractor_ta_ops
            ws['D24'].value = self.torts
            ws['D25'].value = self.trailer_ops
            ws['D26'].value = self.tolls
            ws['D27'].value = self.van_fuel_cost
            ws['D28'].value = self.tractor_sa_fuel_cost
            ws['D29'].value = self.tractor_ta_fuel_cost
            ws['D30'].value = self.total_owned_transportation_cost

        if scenario == 3:
            ws['B21'].value = self.torts_per_mile
            ws['B22'].value = self.tolls
            ws['B23'].value = self.cargo_van_annual_fuel
            ws['B24'].value = self.tractor_sa_annual_fuel
            ws['B25'].value = self.tractor_ta_annual_fuel

            ws['C21'].value = self.usps_annual_miles
            ws['C22'].value = self.tolls
            ws['C23'].value = self.van_count
            ws['C24'].value = self.tractor_sa_count
            ws['C25'].value = self.tractor_ta_count

            ws['D21'].value = self.torts
            ws['D22'].value = self.tolls
            ws['D23'].value = self.van_fuel_cost
            ws['D24'].value = self.tractor_sa_fuel_cost
            ws['D25'].value = self.tractor_ta_fuel_cost
            ws['D26'].value = self.total_leased_transportation_cost

        if scenario == 1:
            ws['B32'].value = self.cargo_van_acq_cost
            ws['B33'].value = self.tractor_sa_acq_cost
            ws['B34'].value = self.tractor_ta_acq_cost
            ws['B35'].value = self.trailer_acq_cost
        if scenario == 2:
            ws['B32'].value = self.cargo_van_dep_cost
            ws['B33'].value = self.tractor_sa_dep_cost
            ws['B34'].value = self.tractor_ta_dep_cost
            ws['B35'].value = self.trailer_acq_cost
        if scenario in (1, 2):
            ws['C32'].value = self.van_count
            ws['C33'].value = self.tractor_sa_count
            ws['C34'].value = self.tractor_ta_count
            ws['C35'].value = self.trailer_count

        if scenario == 3:
            ws['B28'].value = self.one_van_lease_cost
            ws['B29'].value = self.one_tractor_sa_lease_cost
            ws['B30'].value = self.one_tractor_ta_lease_cost
            ws['B31'].value = self.one_trailer_lease_cost
            ws['B32'].value = self.trailer_acq_cost

            ws['C28'].value = self.van_count
            ws['C29'].value = self.tractor_sa_count
            ws['C30'].value = self.tractor_ta_count
            ws['C31'].value = self.trailer_count

        if scenario == 1:
            ws['D32'].value = self.cargo_van_acq_cost * self.van_count
            ws['D33'].value = self.tractor_sa_acq_cost * self.tractor_sa_count
            ws['D34'].value = self.tractor_ta_acq_cost * self.tractor_ta_count
            ws['D35'].value = self.trailer_acq_cost * self.trailer_count
            ws['D36'].value = self.total_acq_cost_wo_dep
        if scenario == 2:
            ws['D32'].value = self.cargo_van_dep_cost * self.van_count
            ws['D33'].value = self.tractor_sa_dep_cost * self.tractor_sa_count
            ws['D34'].value = self.tractor_ta_dep_cost * self.tractor_ta_count
            ws['D35'].value = self.trailer_acq_cost * self.trailer_count
            ws['D36'].value = self.total_acq_cost_w_dep
        if scenario == 3:
            ws['D28'].value = self.van_base_lease_cost
            ws['D29'].value = self.tractor_sa_base_lease_cost
            ws['D30'].value = self.tractor_ta_base_lease_cost
            ws['D31'].value = self.trailer_base_lease_cost
            ws['D32'].value = self.total_extra_lease_cost
            ws['D33'].value = self.total_lease_cost

        if scenario == 1:
            ws['B38'].value = self.total_cost_wo_dep
            value_delta = self.total_annual_rate - self.total_cost_wo_dep
            ws['B39'].value = value_delta
        if scenario == 2:
            ws['B38'].value = self.total_cost_w_dep
            value_delta = self.total_annual_rate - self.total_cost_w_dep
            ws['B39'].value = value_delta
        if scenario == 3:
            ws['B35'].value = self.total_cost_w_lease
            value_delta = self.total_annual_rate - self.total_cost_w_lease
            ws['B36'].value = value_delta

        red_fill = styles.fills.PatternFill(patternType='solid', fgColor=styles.colors.RED)
        green_fill = styles.fills.PatternFill(patternType='solid', fgColor=styles.colors.GREEN)

        if value_delta < 0:
            if scenario in (1, 2):
                ws['B39'].fill = red_fill
            elif scenario == 3:
                ws['B36'].fill = red_fill
        if value_delta > 0:
            if scenario in (1, 2):
                ws['B39'].fill = green_fill
            elif scenario == 3:
                ws['B36'].fill = green_fill

    def print_one_summary_postalized(self, ws, scenario):

        ws['A1'].value = "HCR Contract Being Compared: " + self.origin_terminus
        ws['B1'].value = self.hcr_id
        ws['D1'].value = self.area
        ws['B3'].value = self.total_annual_hours
        ws['B4'].value = self.total_annual_miles
        ws['B5'].value = self.total_annual_rate
        ws['B6'].value = self.van_count
        ws['B7'].value = self.tractor_sa_count
        ws['B8'].value = self.tractor_ta_count
        ws['B9'].value = self.trailer_count

        ws['A12'].value = "Total Driving Hours: " + str(number(self.paid_postalized_hours))
        ws['B13'].value = self.wtd_ft_wage_rate
        ws['B14'].value = self.wtd_pse_wage_rate
        ws['D13'].value = self.ft_costs_postalized
        ws['D14'].value = self.pse_costs_postalized
        ws['D15'].value = self.manager_costs_postalized
        ws['D16'].value = self.ft_ot_costs_postalized + self.pse_ot_costs_postalized
        ws['D17'].value = self.year_one_recruitment_cost_postalized + self.recurring_recruitment_cost_postalized
        ws['D18'].value = self.total_labor_cost_postalized

        ws['A20'].value = "Total PVS Miles: " + str(number(self.total_postalized_mileage))

        if scenario in (1, 2):
            ws['B21'].value = self.van_cpm
            ws['B22'].value = self.tractor_sa_cpm
            ws['B23'].value = self.tractor_ta_cpm
            ws['B24'].value = self.torts_per_mile
            ws['B25'].value = self.annual_maint_per_trailer
            ws['B26'].value = self.tolls
            ws['B27'].value = self.cargo_van_annual_fuel
            ws['B28'].value = self.tractor_sa_annual_fuel
            ws['B29'].value = self.tractor_ta_annual_fuel

            ws['C21'].value = self.van_miles_postalized
            ws['C22'].value = self.tractor_sa_miles_postalized
            ws['C23'].value = self.tractor_ta_miles_postalized
            ws['C24'].value = self.total_postalized_mileage
            ws['C25'].value = self.trailer_count
            ws['C26'].value = self.tolls
            ws['C27'].value = self.van_count
            ws['C28'].value = self.tractor_sa_count
            ws['C29'].value = self.tractor_ta_count

            ws['D21'].value = self.cargo_van_ops_postalized
            ws['D22'].value = self.tractor_sa_ops_postalized
            ws['D23'].value = self.tractor_ta_ops_postalized
            ws['D24'].value = self.torts_postalized
            ws['D25'].value = self.trailer_ops
            ws['D26'].value = self.tolls_postalized
            ws['D27'].value = self.van_fuel_cost
            ws['D28'].value = self.tractor_sa_fuel_cost
            ws['D29'].value = self.tractor_ta_fuel_cost
            ws['D30'].value = self.total_owned_transportation_cost_postalized

        if scenario == 3:
            ws['B21'].value = self.torts_per_mile
            ws['B22'].value = self.tolls
            ws['B23'].value = self.cargo_van_annual_fuel
            ws['B24'].value = self.tractor_sa_annual_fuel
            ws['B25'].value = self.tractor_ta_annual_fuel

            ws['C21'].value = self.total_postalized_mileage
            ws['C22'].value = self.tolls
            ws['C23'].value = self.van_count
            ws['C24'].value = self.tractor_sa_count
            ws['C25'].value = self.tractor_ta_count

            ws['D21'].value = self.torts_postalized
            ws['D22'].value = self.tolls_postalized
            ws['D23'].value = self.van_fuel_cost
            ws['D24'].value = self.tractor_sa_fuel_cost
            ws['D25'].value = self.tractor_ta_fuel_cost
            ws['D26'].value = self.total_leased_transportation_cost_postalized

        if scenario == 1:
            ws['B32'].value = self.cargo_van_acq_cost
            ws['B33'].value = self.tractor_sa_acq_cost
            ws['B34'].value = self.tractor_ta_acq_cost
            ws['B35'].value = self.trailer_acq_cost
        if scenario == 2:
            ws['B32'].value = self.cargo_van_dep_cost
            ws['B33'].value = self.tractor_sa_dep_cost
            ws['B34'].value = self.tractor_ta_dep_cost
            ws['B35'].value = self.trailer_acq_cost
        if scenario in (1, 2):
            ws['C32'].value = self.van_count
            ws['C33'].value = self.tractor_sa_count
            ws['C34'].value = self.tractor_ta_count
            ws['C35'].value = self.trailer_count

        if scenario == 3:
            ws['B28'].value = self.one_van_lease_cost
            ws['B29'].value = self.one_tractor_sa_lease_cost
            ws['B30'].value = self.one_tractor_ta_lease_cost
            ws['B31'].value = self.one_trailer_lease_cost

            ws['C28'].value = self.van_count
            ws['C29'].value = self.tractor_sa_count
            ws['C30'].value = self.tractor_ta_count
            ws['C31'].value = self.trailer_count

        if scenario == 1:
            ws['D32'].value = self.cargo_van_acq_cost * self.van_count
            ws['D33'].value = self.tractor_sa_acq_cost * self.tractor_sa_count
            ws['D34'].value = self.tractor_ta_acq_cost * self.tractor_ta_count
            ws['D35'].value = self.trailer_acq_cost * self.trailer_count
            ws['D36'].value = self.total_acq_cost_wo_dep
        if scenario == 2:
            ws['D32'].value = self.cargo_van_dep_cost * self.van_count
            ws['D33'].value = self.tractor_sa_dep_cost * self.tractor_sa_count
            ws['D34'].value = self.tractor_ta_dep_cost * self.tractor_ta_count
            ws['D35'].value = self.trailer_acq_cost * self.trailer_count
            ws['D36'].value = self.total_acq_cost_w_dep
        if scenario == 3:
            ws['D28'].value = self.van_base_lease_cost
            ws['D29'].value = self.tractor_sa_base_lease_cost
            ws['D30'].value = self.tractor_ta_base_lease_cost
            ws['D31'].value = self.trailer_base_lease_cost
            ws['D32'].value = self.total_extra_lease_cost_postalized
            ws['D33'].value = self.total_lease_cost_postalized

        if scenario == 1:
            ws['B38'].value = self.total_cost_wo_dep_postalized
            value_delta = self.total_annual_rate - self.total_cost_wo_dep_postalized
            ws['B39'].value = value_delta
        if scenario == 2:
            ws['B38'].value = self.total_cost_w_dep_postalized
            value_delta = self.total_annual_rate - self.total_cost_w_dep_postalized
            ws['B39'].value = value_delta
        if scenario == 3:
            ws['B35'].value = self.total_cost_w_lease_postalized
            value_delta = self.total_annual_rate - self.total_cost_w_lease_postalized
            ws['B36'].value = value_delta

        red_fill = styles.fills.PatternFill(patternType='solid', fgColor=styles.colors.RED)
        green_fill = styles.fills.PatternFill(patternType='solid', fgColor=styles.colors.GREEN)

        if value_delta < 0:
            if scenario in (1, 2):
                ws['B39'].fill = red_fill
            elif scenario == 3:
                ws['B36'].fill = red_fill
        if value_delta > 0:
            if scenario in (1, 2):
                ws['B39'].fill = green_fill
            elif scenario == 3:
                ws['B36'].fill = green_fill


class HCRContractModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role != Qt.DisplayRole:
            return None

        value = str(self._data[index.row()][index.column()])

        return value

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]
        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable


class HCRContractView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, contracts):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(True)
        self.sorter.setSourceModel(HCRContractModel(contracts))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)


class HCRVehicleReader:

    def __init__(self, file_name):
        print("Reading HCR vehicle contracts from " + file_name)
        self.file_name = file_name
        self.vehicles_df = None
        self.vehicles = []
        if file_name[-3:] != 'csv':
            print("Converting to csv")
            self.convert_to_csv()
        print("Reading in data")
        self.pandas_in()
        print("Populating vehicle list")
        self.populate_vehicles()

    def convert_to_csv(self):

        wb = load_workbook(filename=self.file_name, read_only=True)
        ws = wb['Vehicle_Data']
        with open('vehicle_contracts.csv', 'w', newline="") as f:
            c = csv.writer(f)
            for r in ws.rows:
                c.writerow([cell.value for cell in r])

    def pandas_in(self):

        df = pd.read_csv('vehicle_contracts.csv')
        df = df[df['VEHICLE DESC'].str.contains('PEAK') == False]
        self.vehicles_df = df.groupby(['CONTRACT NO', 'VEHICLE TYPE'])['VEH QTY'].sum().reset_index()

        v_types = self.vehicle_type_df()
        self.vehicles_df = pd.merge(self.vehicles_df, v_types, on='VEHICLE TYPE', how='left')
        self.vehicles_df = self.vehicles_df.pivot_table(index='CONTRACT NO', columns='VEH CAT',
                                                        values='VEH QTY', aggfunc='sum').reset_index()
        self.vehicles_df = self.vehicles_df.fillna(0)

    @staticmethod
    def vehicle_type_df():
        vehicle_types = ['28 Foot Single Axle Trailer', '28 Foot Tandem Axle Trailer',
                         '40 Foot Single Axle Trailer', '40 Foot Tandem Axle Trailer',
                         '45 Foot Single Axle Trailer', '45 Foot Tandem Axle Trailer',
                         '48 Foot Tandem Axle Trailer', '53 Foot Rollerbed Trailer',
                         '53 Foot Tandem Axle Trailer', 'Boat', 'Four Wheel Drive',
                         'Mini Van', 'Passenger Car', 'Pick-up Truck',
                         'Single Axle Tractor', 'Station Wagon',
                         'Tandem Axle Converter Gear', 'Two Axle Tractor (Double Drive)',
                         'Two Axle Tractor (Single Drive)', 'Van']
        vehicle_categories = ['Trailer', 'Trailer', 'Trailer', 'Trailer', 'Trailer', 'Trailer', 'Trailer',
                              'Trailer', 'Trailer', 'n/a', 'n/a', 'n/a', 'n/a', 'n/a', 'Tractor (SA)',
                              'n/a', 'Tractor (TA)', 'Tractor (TA)', 'Tractor (TA)', 'Cargo Van']

        df = pd.DataFrame({'VEHICLE TYPE': vehicle_types, 'VEH CAT': vehicle_categories})

        return df

    def get_contract(self, hcr_id):

        row = [x for x in self.vehicles if x.hcr_id == hcr_id]

        if len(row) > 1:
            print("more than one contract found??")
        elif len(row) < 1:
            print("vehicle contract not found")
            return None
        else:
            print("vehicle contract found")
            return row[0]

    def populate_vehicles(self):

        data = self.vehicles_df.values.tolist()
        for row in data:
            hcr_id = row[0]
            van = int(row[1])
            tractor_sa = int(row[2])
            tractor_ta = int(row[3])
            trailer = int(row[4])
            n_a = int(row[5])
            self.vehicles.append(HCRVehicleContract(hcr_id, van, tractor_sa, tractor_ta, trailer))
        print(len(self.vehicles))


class HCRVehicleContract:

    def __init__(self, hcr_id, cargo_van_count, tractor_sa_count, tractor_ta_count, trailer_count):

        self.hcr_id = hcr_id
        self.cargo_van_count = cargo_van_count
        self.tractor_sa_count = tractor_sa_count
        self.tractor_ta_count = tractor_ta_count
        self.trailer_count = trailer_count

        self.total_count = cargo_van_count + tractor_sa_count + tractor_ta_count
        if self.total_count != 0:
            self.percent_vans = cargo_van_count/self.total_count
            self.percent_tractors = (tractor_sa_count + tractor_ta_count)/self.total_count
            self.percent_tractors_sa = tractor_sa_count/self.total_count
            self.percent_tractors_ta = tractor_ta_count/self.total_count
        else:
            self.percent_vans = 0
            self.percent_tractors = 0
            self.percent_tractors_ta = 0
            self.percent_tractors_sa = 0

    def print_vehicle_contract(self, ws, row):

        ws["A" + str(row)].value = str(row - 7)
        ws["B" + str(row)].value = self.hcr_id
        ws["C" + str(row)].value = self.cargo_van_count
        ws["D" + str(row)].value = self.tractor_sa_count
        ws["E" + str(row)].value = self.tractor_ta_count
        ws["F" + str(row)].value = self.trailer_count
        ws["G" + str(row)].value = self.total_count
        ws["H" + str(row)].value = self.percent_vans
        ws["I" + str(row)].value = self.percent_tractors


class HCRVehicleModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role != Qt.DisplayRole:
            return None

        value = str(self._data[index.row()][index.column()])

        return value

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]
        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable


class HCRVehicleView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, contracts):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(True)
        self.sorter.setSourceModel(HCRContractModel(contracts))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)


class HCRPlateModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self.header_data = data.columns

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role != Qt.DisplayRole:
            return None

        value = str(self._data[index.row()][index.column()])

        return value

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header_data[col]
        return QVariant()

    def flags(self, index):
        return Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable


class HCRPlateView(QTableView):

    def __init__(self, parent=None):
        super().__init__()
        self.sorter = None

    def set_model(self, data):

        self.sorter = QSortFilterProxyModel()
        self.sorter.setDynamicSortFilter(True)
        self.sorter.setSourceModel(HCRPlateModel(data))
        self.setModel(self.sorter)
        self.resizeColumnsToContents()
        self.setSortingEnabled(True)


class HCRPlatesToPandas:

    def __init__(self, plate_nums, read_in_correctly):

        self.plate_nums = plate_nums
        self.read_in_correctly = read_in_correctly

        labels = ["Plate Num", "Read Correctly"]
        compiled = [self.plate_nums, self.read_in_correctly]

        self.df = pd.DataFrame(list(zip(self.plate_nums, self.read_in_correctly)), columns=labels)