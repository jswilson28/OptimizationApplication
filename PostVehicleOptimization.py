# import pandas as pd
# from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant, QSortFilterProxyModel
# from PyQt5.QtWidgets import QTableView
from openpyxl import load_workbook
from GeneralMethods import cminute_to_tod, today


class VehicleSummarizer:

    def __init__(self, eleven_ton_file, single_file, site_name):
        print("Printing vehicles")

        self.site_name = site_name
        self.output_file_name = "FinalVehicles_" + site_name + "_" + today() + ".xlsx"

        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

        self.eleven_ton = []
        self.num_eleven_ton = 0
        self.single = []
        self.num_single = 0

        if eleven_ton_file:
            self.eleven_ton = self.read_in_file(eleven_ton_file)

        if single_file:
            self.single = self.read_in_file(single_file)

        self.output_wb = load_workbook("Optimization Formats/Vehicle Summary Template.xlsx")
        self.output_wb.save(self.output_file_name)

        self.print_days()
        self.print_summary()

        self.output_wb.save(self.output_file_name)
        self.output_wb.close()
        print("Printed Vehicles")

    def read_in_file(self, file_name):

        input_wb = load_workbook(filename=file_name, read_only=True)

        schedule_days = []
        solution_days = []

        for day in self.days:
            schedule_days.append(self.read_in_schedule_day(input_wb, day))
            solution_days.append(self.read_in_solution_day(input_wb, day))

        combined = self.combine_lists(schedule_days, solution_days)
        self.set_time_strings(combined)

        input_wb.close()

        return combined

    @staticmethod
    def read_in_schedule_day(input_wb, day):

        optimized_schedules = []

        day_string = "OptiData " + day

        try:
            ws = input_wb[day_string]
        except:
            return []

        max_row = ws.max_row
        x = 9

        while x <= max_row:

            schedule_num = ws["B" + str(x)].value
            trip_num = ws["C" + str(x)].value
            break_vt = ws["D" + str(x)].value
            pdc_start_time = ws["E" + str(x)].value
            pdc_end_time = ws["F" + str(x)].value
            duration = ws["G" + str(x)].value

            if schedule_num not in (None, "", " "):
                temp_list = [day, schedule_num, trip_num, break_vt, pdc_start_time, pdc_end_time, duration]
                optimized_schedules.append(temp_list)
            else:
                break

            x += 1

        return optimized_schedules

    @staticmethod
    def read_in_solution_day(input_wb, day):

        day_string = "Solution " + day

        solutions = []

        try:
            ws = input_wb[day_string]
        except:
            return []

        max_row = ws.max_row

        x = 7

        while x <= max_row:

            vehicle_num = ws["B" + str(x)].value
            trip = ws["C" + str(x)].value
            break_in = ws["D" + str(x)].value
            break_after = ws["E" + str(x)].value

            if vehicle_num not in (None, " ", ""):
                temp_list = [day, vehicle_num, trip, break_in, break_after]
                solutions.append(temp_list)
            else:
                break

            x += 1

        return solutions

    def combine_lists(self, schedule_days, solution_days):

        combined = []

        for x, day in enumerate(self.days):
            optimized_schedules = schedule_days[x]
            solutions = solution_days[x]
            for schedule in optimized_schedules:
                solution = [x for x in solutions if x[2] == schedule[2]]
                if len(solution) != 1:
                    print("error matching solutions to schedules")
                    return
                else:
                    solution = solution[0]
                solution.append(schedule[4])
                solution.append(schedule[5])
                solution.append(schedule[6])
                solution.append(schedule[1])
                combined.append(solution)

        return combined

    @staticmethod
    def set_time_strings(combined):

        for x, schedule in enumerate(combined):
            start_time = cminute_to_tod(schedule[5])
            end_time = cminute_to_tod(schedule[6])
            time_string = start_time.strftime("%H%M") + "-" + end_time.strftime("%H%M")
            combined[x].append(time_string)

    def print_days(self):

        for day in self.days:

            ws = self.output_wb[day + " Vehicles"]

            eleven_ton = [x for x in self.eleven_ton if x[0] == day]
            single = [x for x in self.single if x[0] == day]

            # print right table
            x = 4
            for schedule in eleven_ton:
                ws["M" + str(x)].value = schedule[8]
                ws["N" + str(x)].value = "M"
                ws["O" + str(x)].value = schedule[9]
                x += 1
            for schedule in single:
                ws["M" + str(x)].value = schedule[8]
                ws["N" + str(x)].value = "T"
                ws["O" + str(x)].value = schedule[9]
                x += 1

            # print left table, eleven ton
            vehicle_nums = list(set([x[1] for x in eleven_ton]))
            self.num_eleven_ton = max(len(vehicle_nums), self.num_eleven_ton)

            x = 4
            y = 1
            col = 1

            for z, vehicle in enumerate(vehicle_nums):
                assignments = [x for x in eleven_ton if x[1] == vehicle]
                ws.cell(row=x, column=col).value = z + 1
                for assignment in assignments:
                    ws.cell(row=x, column=(col + y)).value = assignment[8]
                    y += 1
                x += 1
                y = 1

            # print left_table, single
            vehicle_nums = list(set([x[1] for x in single]))
            self.num_single = max(len(vehicle_nums), self.num_single)

            x = 4
            y = 1
            col = 7

            for z, vehicle in enumerate(vehicle_nums):
                assignments = [x for x in single if x[1] == vehicle]
                ws.cell(row=x, column=col).value = z + 1
                for assignment in assignments:
                    ws.cell(row=x, column=(col + y)).value = assignment[8]
                    y += 1
                x += 1
                y = 1

    def print_summary(self):

        ws = self.output_wb["Additional Vehicles Required"]

        ws["I7"].value = self.num_single
        ws["E7"].value = self.num_eleven_ton
        ws["B2"].value = self.site_name + " Additional Vehicles Required"
