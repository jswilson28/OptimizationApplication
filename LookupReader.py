# Contains references to all the static excel files used to preserve data session to session
from openpyxl import load_workbook
from PyQt5.QtWidgets import (QGridLayout, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QDialog, QLineEdit,
                             QCheckBox, QComboBox)


def create_adjust_code(freq_code):

    freq_code = str(freq_code)
    freq_code = freq_code.zfill(4)
    if "9" in freq_code:
        return freq_code

    if len(freq_code) == 5:
        return freq_code

    days = []
    day_indices = []
    for x, c in enumerate(freq_code):
        if c.isnumeric():
            if c != "0":
                if c == "7":
                    days.append(1)
                    day_indices.append(x)
                elif int(c) < 7:
                    days.append(int(c)+1)
                    day_indices.append(x)

    if not days:
        return freq_code

    days.sort()
    new_code = ""
    y = 0
    for x in range(0, 4):
        if x in day_indices:
            new_code += str(days[y])
            y += 1
        else:
            new_code += freq_code[x]

    return new_code


class LookupBook:

    def __init__(self):

        print("initializing")

        self.wb = load_workbook("Lookups/lookups.xlsx")
        self.save_workbook()

        self.hcr_pdcs = []
        self.pvs_pdcs = []
        self.pvs_names = []
        self.postal_facility_names = []
        self.alternate_pvs_names = []
        self.alternate_pdc_names = []
        self.short_names = []
        self.regions = []
        self.jda_facility_names = []
        self.pdc_addresses = []

        self.single_list = []
        self.eleven_ton_list = []
        self.spotter_list = []
        self.all_vehicle_list = []

        self.day_strings = []
        self.new_monday_codes = []
        self.new_tuesday_codes = []
        self.new_wednesday_codes = []
        self.new_thursday_codes = []
        self.new_friday_codes = []
        self.new_saturday_codes = []
        self.new_sunday_codes = []

        self.day_strings_ns = []
        self.new_monday_codes_ns = []
        self.new_tuesday_codes_ns = []
        self.new_wednesday_codes_ns = []
        self.new_thursday_codes_ns = []
        self.new_friday_codes_ns = []
        self.new_saturday_codes_ns = []
        self.new_sunday_codes_ns = []

        self.bin_strings = []
        self.new_codes = []

        self.known_codes = []

        self.persistent_plates = []
        self.persistent_sites = []

        self.network_nass_codes = []

        self.set_all_lookups()
        print("initialized")

    def set_all_lookups(self):
        # read pvs site names
        ws = self.wb["PVS Lookup"]
        max_row = ws.max_row
        row = 2

        while row <= max_row:
            self.hcr_pdcs.append(ws["A" + str(row)].value)
            self.pvs_pdcs.append(ws["B" + str(row)].value)
            self.pvs_names.append(ws["C" + str(row)].value)
            self.postal_facility_names.append(ws["D" + str(row)].value)
            self.alternate_pvs_names.append(ws["E" + str(row)].value)
            self.alternate_pdc_names.append(ws["F" + str(row)].value)
            self.short_names.append(ws["G" + str(row)].value)
            self.regions.append(ws["H" + str(row)].value)
            self.jda_facility_names.append(ws["I" + str(row)].value)
            self.pdc_addresses.append(ws["J" + str(row)].value)
            row += 1

        # read in vehicle types
        ws = self.wb["Vehicle Types"]
        max_row = ws.max_row
        row = 1

        while row <= max_row:
            if ws["A" + str(row)].value:
                self.single_list.append(ws["A" + str(row)].value)
            if ws["B" + str(row)].value:
                self.eleven_ton_list.append(ws["B" + str(row)].value)
            if ws["C" + str(row)].value:
                self.spotter_list.append(ws["C" + str(row)].value)
            row += 1

        self.spotter_list.append(None)
        self.all_vehicle_list = self.single_list + self.eleven_ton_list + self.spotter_list

        ws = self.wb["New Code Assignment"]
        row = 4
        while row <= 131:
            self.day_strings.append(ws["A" + str(row)].value)
            self.new_monday_codes.append(ws["J" + str(row)].value)
            self.new_tuesday_codes.append(ws["K" + str(row)].value)
            self.new_wednesday_codes.append(ws["L" + str(row)].value)
            self.new_thursday_codes.append(ws["M" + str(row)].value)
            self.new_friday_codes.append(ws["N" + str(row)].value)
            self.new_saturday_codes.append(ws["O" + str(row)].value)
            self.new_sunday_codes.append(ws["P" + str(row)].value)
            row += 1

        ws = self.wb["New Code Assignment No Sun"]
        row = 4
        while row <= 131:
            self.day_strings_ns.append(ws["A" + str(row)].value)
            self.new_monday_codes_ns.append(ws["J" + str(row)].value)
            self.new_tuesday_codes_ns.append(ws["K" + str(row)].value)
            self.new_wednesday_codes_ns.append(ws["L" + str(row)].value)
            self.new_thursday_codes_ns.append(ws["M" + str(row)].value)
            self.new_friday_codes_ns.append(ws["N" + str(row)].value)
            self.new_saturday_codes_ns.append(ws["O" + str(row)].value)
            self.new_sunday_codes_ns.append(ws["P" + str(row)].value)
            row += 1

        ws = self.wb['New Codes']
        row = 4
        while row < 132:
            self.bin_strings.append(ws["A" + str(row)].value)
            self.new_codes.append(ws["I" + str(row)].value)
            row += 1

        ws = self.wb["All Codes Corrected"]
        row = 2
        max_row = ws.max_row
        while row <= max_row:
            freq_code = ws["A" + str(row)].value
            annual_trips = ws["B" + str(row)].value
            code_desc = ws["C" + str(row)].value
            bin_string = ws["L" + str(row)].value
            adjust_code = ws["M" + str(row)].value
            self.known_codes.append([freq_code, annual_trips, code_desc, bin_string, adjust_code])
            row += 1

        ws = self.wb["PlatesToSites"]
        for row in ws.iter_rows(min_row=2):
            self.persistent_plates.append(row[0].value)
            self.persistent_sites.append(row[1].value)

        ws = self.wb["Network Sites"]
        for row in ws.iter_rows(min_row=2):
            self.network_nass_codes.append([cell.value for cell in row])

        self.wb.close()

    def find_code_info(self, freq_code):

        found_codes = [x for x in self.known_codes if x[0] == freq_code]

        if len(found_codes) < 1:
            print("Code not found")
            return
        elif len(found_codes) > 1:
            print("multiple codes found?!?!?")
            return
        else:
            return found_codes[0]

    def get_day_codes(self, day_string, sunday_schedules):

        index = self.day_strings.index(day_string)

        if sunday_schedules:
            mon_code = self.new_monday_codes[index]
            tue_code = self.new_tuesday_codes[index]
            wed_code = self.new_wednesday_codes[index]
            thu_code = self.new_thursday_codes[index]
            fri_code = self.new_friday_codes[index]
            sat_code = self.new_saturday_codes[index]
            sun_code = self.new_sunday_codes[index]
        else:
            mon_code = self.new_monday_codes_ns[index]
            tue_code = self.new_tuesday_codes_ns[index]
            wed_code = self.new_wednesday_codes_ns[index]
            thu_code = self.new_thursday_codes_ns[index]
            fri_code = self.new_friday_codes_ns[index]
            sat_code = self.new_saturday_codes_ns[index]
            sun_code = self.new_sunday_codes_ns[index]

        return [mon_code, tue_code, wed_code, thu_code, fri_code, sat_code, sun_code]

    def get_new_bin_string(self, code):

        return self.bin_strings[self.new_codes.index(code)]

    def add_facility_names(self, **kwargs):

        self.load_wb()
        ws = self.wb["PVS Lookup"]
        max_row = ws.max_row

        hcr_pdc = kwargs.get('hcr_pdc')
        pvs_pdc = kwargs.get('pvs_pdc')
        pvs_name = kwargs.get('pvs_name')
        short_name = kwargs.get('short_name')
        postal_facility_name = kwargs.get('pf_name')
        alternate_pvs = kwargs.get('alternate_pvs')
        alternate_pdc = kwargs.get('alternate_pdc')
        # region = kwargs.get('region')
        address = kwargs.get('address')

        ws["A" + str(max_row)].value = hcr_pdc
        ws["B" + str(max_row)].value = pvs_pdc
        ws["C" + str(max_row)].value = pvs_name
        ws["G" + str(max_row)].value = short_name

        self.save_workbook()
        self.wb.close()

    def add_frequency_code(self, **kwargs):

        freq_code = kwargs.get("freq_code")
        annual_trips = kwargs.get("annual_trips")
        description = kwargs.get("description")
        bin_string = kwargs.get("bin_string")
        adjust_code = kwargs.get("adjust_code")

        if freq_code in [x[0] for x in self.known_codes]:
            return

        self.known_codes.append([freq_code, annual_trips, description, bin_string, adjust_code])

        self.load_wb()
        ws = self.wb["All Codes Corrected"]
        row = str(ws.max_row + 1)
        ws["A" + row].value = freq_code
        ws["B" + row].value = annual_trips
        ws["C" + row].value = description
        ws["D" + row].value = bin_string[0]
        ws["E" + row].value = bin_string[1]
        ws["F" + row].value = bin_string[2]
        ws["G" + row].value = bin_string[3]
        ws["H" + row].value = bin_string[4]
        ws["I" + row].value = bin_string[5]
        ws["J" + row].value = bin_string[6]
        if len(bin_string) == 8:
            ws["K" + row].value = bin_string[7]
        ws["L" + row].value = bin_string
        ws["M" + row].value = adjust_code

        self.save_workbook()
        self.wb.close()

    def add_hcr_info(self, code_list):

        freq_code = code_list[0]
        annual_trips = code_list[1]
        description = code_list[2]

        code_row = next(x for x in self.known_codes if x[0] == freq_code)
        index = self.known_codes.index(code_row)
        self.known_codes[index][1] = annual_trips
        self.known_codes[index][2] = description

        self.load_wb()
        ws = self.wb["All Codes (Final)"]

        x = 2
        while ws["A" + str(x)].value != freq_code:
            if ws["A" + str(x)].value in (" ", "", None):
                print("This should never have happened, check lookupreader.")
                return
            x += 1

        ws["B" + str(x)].value = annual_trips
        ws["C" + str(x)].value = description

        self.save_workbook()
        self.wb.close()

    def add_adjust_code(self, freq_code, adjust_code):
        # this function might be obsolete
        code_row = next(x for x in self.known_codes if x[0] == freq_code)
        index = self.known_codes.index(code_row)
        self.known_codes[index][4] = adjust_code

        self.load_wb()
        ws = self.wb["All Codes (Final)"]

        x = 2
        while ws["A" + str(x)].value != freq_code:
            if ws["A" + str(x)].value in (" ", "", None):
                print("This should never have happened, check lookupreader.")
                return
            x += 1

        ws["M" + str(x)].value = adjust_code

        self.save_workbook()
        self.wb.close()

    def load_wb(self):

        self.wb = load_workbook("Lookups/lookups.xlsx")

    def save_workbook(self):

        try:
            self.wb.save("Lookups/lookups.xlsx")
            return True
        except:
            print("Close Lookup Book!")
            return False

    def get_facility_items(self, name):

        index = None

        if name in self.hcr_pdcs:
            index = self.hcr_pdcs.index(name)
        elif name in self.pvs_pdcs:
            index = self.pvs_pdcs.index(name)
        elif name in self.alternate_pdc_names:
            index = self.alternate_pdc_names.index(name)
        elif name in self.postal_facility_names:
            index = self.postal_facility_names.index(name)

        return_list = [None, None, None, None, None, None, None, None]
        if index:
            return_list = [self.hcr_pdcs[index], self.pvs_pdcs[index], self.pvs_names[index],
                           self.postal_facility_names[index], self.alternate_pdc_names[index],
                           self.alternate_pvs_names[index], self.short_names[index], self.pdc_addresses[index]]

        return return_list

    def update_from_list(self, index):

        self.load_wb()
        ws = self.wb["PVS Lookup"]
        max_row = ws.max_row + 1
        row = 2
        while row <= max_row and ws["G" + str(row)].value != self.short_names[index]:
            row += 1

        ws["A" + str(row)].value = self.hcr_pdcs[index]
        ws["B" + str(row)].value = self.pvs_pdcs[index]
        ws["C" + str(row)].value = self.pvs_names[index]
        ws["D" + str(row)].value = self.postal_facility_names[index]
        ws["E" + str(row)].value = self.alternate_pvs_names[index]
        ws["F" + str(row)].value = self.alternate_pdc_names[index]
        ws["G" + str(row)].value = self.short_names[index]
        ws["H" + str(row)].value = self.regions[index]
        ws["I" + str(row)].value = self.jda_facility_names[index]
        ws["J" + str(row)].value = self.pdc_addresses[index]

        self.save_workbook()
        self.wb.close()

    def add_persistent_info(self, plate_number, short_name):

        self.persistent_plates.append(plate_number)
        self.persistent_sites.append(short_name)

        self.load_wb()
        ws = self.wb["PlatesToSites"]
        max_row = ws.max_row + 1

        ws["A" + str(max_row)].value = str(plate_number)
        ws["B" + str(max_row)].value = str(short_name)

        self.save_workbook()
        self.wb.close()


class FacilityNameEntry(QDialog):

    def __init__(self, parent=None, **kwargs):
        super(FacilityNameEntry, self).__init__(parent)

        # Get any initial names
        self.initial_name = None
        self.pvs_pdc = kwargs.get("pvs_pdc")
        self.pvs_name = kwargs.get("pvs_name")
        self.address = kwargs.get("address")
        self.short_name = kwargs.get("short_name")
        self.region = kwargs.get("region")
        self.hcr_pdc = kwargs.get("hcr_pdc")
        self.html_postal_facility = kwargs.get("html_postal_facility")
        self.html_pvs = kwargs.get("html_pvs")
        self.html_pdc = kwargs.get("html_pdc")
        self.jda_name = kwargs.get("jda_name")

        # source type can be hcr_pdc, pvs_pdc, jda_name, or html_postal_facility
        self.source_type = kwargs.get("source_type")
        self.file_name = kwargs.get("file_name")
        self.lb = kwargs.get("lb")
        self.set_initial_name_from_source()

        # define line edits
        self.pvs_pdc_le = QLineEdit()
        self.pvs_name_le = QLineEdit()
        self.address_le = QLineEdit()
        self.short_name_le = QLineEdit()
        self.region_le = QLineEdit()
        self.hcr_pdc_le = QLineEdit()
        self.html_pf_le = QLineEdit()
        self.html_pvs_le = QLineEdit()
        self.html_pdc_le = QLineEdit()
        self.jda_name_le = QLineEdit()
        self.set_les_to_names()

        # define drop down list
        self.known_names_drop_down = QComboBox()
        self.set_up_known_names_box()

        # buttons
        self.add_to_lookupbook = QCheckBox()
        self.add_to_lookupbook.setChecked(True)
        self.submit_button = QPushButton("Accept")
        self.submit_button.clicked[bool].connect(self.accept)

        self.initUI()
        self.try_to_set_drop_box()

    def initUI(self):

        outer_layout = QVBoxLayout()
        header_bar = QHBoxLayout()
        upper_grid = QGridLayout()
        mid_bar = QHBoxLayout()
        known_facility_bar = QHBoxLayout()
        second_mid_bar = QHBoxLayout()
        le_grid = QGridLayout()
        third_mid_bar = QHBoxLayout()
        bottom_bar = QHBoxLayout()

        header_label = QLabel("Select Facility Names")
        header_bar.addWidget(header_label)

        upper_grid.addWidget(QLabel("File Name:"), 0, 0)
        upper_grid.addWidget(QLabel(self.file_name), 0, 1)

        upper_grid.addWidget(QLabel("Source Type:"), 1, 0)
        upper_grid.addWidget(QLabel(self.source_type), 1, 1)

        upper_grid.addWidget(QLabel("Expected Name:"), 2, 0)
        upper_grid.addWidget(QLabel(self.initial_name), 2, 1)

        mid_bar.addWidget(QLabel("Known Facilities:"))
        mid_bar.addStretch(1)

        known_facility_bar.addWidget(QLabel("Select from known facilities:"))
        known_facility_bar.addWidget(self.known_names_drop_down)

        second_mid_bar.addWidget(QLabel("Inputs:"))
        second_mid_bar.addStretch(1)

        le_grid.addWidget(QLabel("PVS P&DC Name:*"), 0, 0)
        le_grid.addWidget(QLabel("PVS Name:*"), 1, 0)
        le_grid.addWidget(QLabel("Short Name:*"), 2, 0)
        le_grid.addWidget(QLabel("Address:"), 3, 0)
        le_grid.addWidget(QLabel("Region:"), 4, 0)
        le_grid.addWidget(QLabel("HCR P&DC Name:"), 5, 0)
        le_grid.addWidget(QLabel("HTML Postal Facility:"), 6, 0)
        le_grid.addWidget(QLabel("HTML PVS Name:"), 7, 0)
        le_grid.addWidget(QLabel("HTML P&DC Name:"), 8, 0)
        le_grid.addWidget(QLabel("JDA Name:"), 9, 0)

        le_grid.addWidget(self.pvs_pdc_le, 0, 1)
        le_grid.addWidget(self.pvs_name_le, 1, 1)
        le_grid.addWidget(self.short_name_le, 2, 1)
        le_grid.addWidget(self.address_le, 3, 1)
        le_grid.addWidget(self.region_le, 4, 1)
        le_grid.addWidget(self.hcr_pdc_le, 5, 1)
        le_grid.addWidget(self.html_pf_le, 6, 1)
        le_grid.addWidget(self.html_pvs_le, 7, 1)
        le_grid.addWidget(self.html_pdc_le, 8, 1)
        le_grid.addWidget(self.jda_name_le, 9, 1)

        third_mid_bar.addStretch(1)
        third_mid_bar.addWidget(QLabel("Asterisks indicate required fields"))
        third_mid_bar.addStretch(1)

        bottom_bar.addWidget(QLabel("Add new facility to lookups?"))
        bottom_bar.addWidget(self.add_to_lookupbook)
        bottom_bar.addStretch(1)
        bottom_bar.addWidget(self.submit_button)

        outer_layout.addLayout(header_bar)
        outer_layout.addLayout(upper_grid)
        outer_layout.addLayout(mid_bar)
        outer_layout.addLayout(known_facility_bar)
        outer_layout.addLayout(second_mid_bar)
        outer_layout.addLayout(le_grid)
        outer_layout.addLayout(third_mid_bar)
        outer_layout.addLayout(bottom_bar)
        self.setLayout(outer_layout)

    def set_up_known_names_box(self):

        self.known_names_drop_down.addItem("New Facility")

        for name in self.lb.short_names:
            self.known_names_drop_down.addItem(name)

        self.known_names_drop_down.activated.connect(self.populate_items_from_existing_facility)

    def set_initial_name_from_source(self):

        if self.source_type == "hcr_pdc":
            self.initial_name = self.hcr_pdc
        if self.source_type == "pvs_postal_facility":
            self.initial_name = self.html_postal_facility

    def populate_items_from_existing_facility(self):

        self.short_name = self.known_names_drop_down.currentText()

        if self.short_name == "New Facility":
            return

        index = self.lb.short_names.index(self.short_name)

        self.pvs_pdc = self.lb.pvs_pdcs[index]
        self.pvs_name = self.lb.pvs_names[index]
        self.address = self.lb.pdc_addresses[index]
        self.region = self.lb.regions[index]
        self.hcr_pdc = self.lb.hcr_pdcs[index]
        self.html_postal_facility = self.lb.postal_facility_names[index]
        self.html_pvs = self.lb.alternate_pvs_names[index]
        self.html_pdc = self.lb.alternate_pdc_names[index]
        self.jda_name = self.lb.jda_facility_names[index]
        self.set_les_to_names()

    def set_les_to_names(self):

        self.pvs_name_le.setText(self.pvs_name)
        self.pvs_pdc_le.setText(self.pvs_pdc)
        self.address_le.setText(self.address)
        self.short_name_le.setText(self.short_name)
        self.region_le.setText(self.region)
        self.hcr_pdc_le.setText(self.hcr_pdc)
        self.html_pf_le.setText(self.html_postal_facility)
        self.html_pvs_le.setText(self.html_pvs)
        self.html_pdc_le.setText(self.html_pdc)
        self.jda_name_le.setText(self.jda_name)

    def set_names_to_les(self):

        self.pvs_name = self.pvs_name_le.text()
        self.pvs_pdc = self.pvs_pdc_le.text()
        self.address = self.address_le.text()
        self.short_name = self.short_name_le.text()
        self.region = self.region_le.text()
        self.hcr_pdc = self.hcr_pdc_le.text()
        self.html_postal_facility = self.html_pf_le.text()
        self.html_pvs = self.html_pvs_le.text()
        self.html_pdc = self.html_pdc_le.text()
        self.jda_name = self.jda_name_le.text()

    def add_names_to_lookups(self):

        self.set_names_to_les()

        if self.short_name in self.lb.short_names:
            index = self.lb.short_names.index(self.short_name)
            self.lb.pvs_pdcs[index] = self.pvs_pdc
            self.lb.pvs_names[index] = self.pvs_name
            self.lb.pdc_addresses[index] = self.address
            self.lb.regions[index] = self.region
            self.lb.hcr_pdcs[index] = self.hcr_pdc
            self.lb.postal_facility_names[index] = self.html_postal_facility
            self.lb.alternate_pvs_names[index] = self.html_pvs
            self.lb.alternate_pdc_names[index] = self.html_pdc
            self.lb.jda_facility_names[index] = self.jda_name

            if self.add_to_lookupbook.isChecked():
                self.lb.update_from_list(index)

            return

        self.lb.short_names.append(self.short_name)
        self.lb.pvs_pdcs.append(self.pvs_pdc)
        self.lb.pvs_names.append(self.pvs_name)
        self.lb.pdc_addresses.append(self.address)
        self.lb.regions.append(self.region)
        self.lb.hcr_pdcs.append(self.hcr_pdc)
        self.lb.postal_facility_names.append(self.html_postal_facility)
        self.lb.alternate_pvs_names.append(self.html_pvs)
        self.lb.alternate_pdc_names.append(self.html_pdc)
        self.lb.jda_facility_names.append(self.jda_name)

        index = self.lb.short_names.index(self.short_name)
        if self.add_to_lookupbook.isChecked():
            self.lb.update_from_list(index)

    def try_to_set_drop_box(self):

        pass

    def exec_(self):
        super().exec_()
        self.close()
        self.add_names_to_lookups()


class AddFrequencyCode(QDialog):

    def __init__(self, freq_code, annual_trips, description, lb, plate_num=None, parent=None):
        super(AddFrequencyCode, self).__init__(parent)

        self.setWindowTitle("New Code on Plate: " + str(plate_num))

        self.lb = lb
        self.freq_code = freq_code
        self.description = description
        if not self.description:
            self.description = "Unknown"

        self.annual_trips = annual_trips
        if not self.annual_trips:
            self.annual_trips = "Unknown"

        self.check_all = QCheckBox()
        self.hol = QCheckBox()
        self.check_boxes = []

        self.accept_button = QPushButton("Accept", self)

        self.initUI()

    def initUI(self):

        vbox = QVBoxLayout()
        grid = QGridLayout()
        other_hbox = QHBoxLayout()

        freq_code_label = QLabel("Frequency Code: " + str(self.freq_code))
        annual_trips_label = QLabel("Annual Trips: " + str(self.annual_trips))
        description_label = QLabel("Description : " + str(self.description))

        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        for x, day in enumerate(days):
            grid.addWidget(QLabel(day), 0, x)
            day_box = QCheckBox()
            self.check_boxes.append(day_box)
            grid.addWidget(day_box, 1, x)

        self.check_all.clicked.connect(self.check_all_days)

        other_hbox.addWidget(QLabel("Check All"))
        other_hbox.addWidget(self.check_all)
        other_hbox.addStretch(1)
        other_hbox.addWidget(QLabel("Holiday"))
        other_hbox.addWidget(self.hol)

        vbox.addWidget(freq_code_label)
        vbox.addWidget(annual_trips_label)
        vbox.addWidget(description_label)
        vbox.addLayout(grid)
        vbox.addLayout(other_hbox)

        self.accept_button.clicked[bool].connect(self.accept)

        vbox.addWidget(self.accept_button)
        self.setLayout(vbox)

    def check_all_days(self):

        if self.check_all.isChecked():
            for box in self.check_boxes:
                box.setChecked(True)
        else:
            for box in self.check_boxes:
                box.setChecked(False)

    def exec_(self):
        super().exec_()

        bin_string = ""
        for box in self.check_boxes:
            if box.isChecked():
                bin_string += "1"
            else:
                bin_string += "0"

        if self.hol.isChecked():
            bin_string += "1"

        if self.lb and "9" not in self.freq_code:
            self.lb.add_frequency_code(freq_code=self.freq_code, annual_trips=self.annual_trips,
                                       description=self.description, bin_string=bin_string)

        self.close()
        return bin_string
