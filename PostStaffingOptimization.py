# import pandas as pd
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant # , QSortFilterProxyModel
from PyQt5.QtWidgets import QTableView
from openpyxl import load_workbook
import operator
from GeneralMethods import cminute_to_tod, night_differential, sun_minutes, today


class PostStaffingProcess:

    def __init__(self, file_name, ip):

        print("starting staffing print")
        self.nd_am = ip.nd_am
        self.nd_pm = ip.nd_pm

        self.vehicle_type = "Unknown"
        eleven_ton_list = ["11-Ton", "11-TON", "11TON", "11Ton", "11ton", "11-ton"]
        single_list = ["Single", "SINGLE", "single"]

        if any(x in file_name for x in eleven_ton_list):
            self.vehicle_type = "11-Ton"
        elif any(x in file_name for x in single_list):
            self.vehicle_type = "Single"

        try:
            self.pdc_name = file_name.split("_")[-2]
        except:
            self.pdc_name = "Unknown"

        self.file_name = "FinalStaffing_" + self.pdc_name + "_" + self.vehicle_type + "_" + today() + ".xlsx"

        self.wb = load_workbook(file_name)
        self.wb.save(self.file_name)

        self.schedule_nums_and_lunches = []
        self.results_list = []

        self.employees = [] # this list is all schedules compiled to each employee
        self.minutes_list = [] # this list is all weekly times by employee
        self.total_hours = [] # this list is total annual hours by employee type

        self.read_in_lunches()
        self.read_in_results()

        self.categorize_employees()
        self.calculate_minutes()
        self.calculate_total_hours()

        self.print_schedules()
        self.print_schedule_summary()
        self.print_employee_summary()
        # self.print_total_hours()

        self.wb.save(self.file_name)
        self.wb.close()

        print("finished staffing print")

    def read_in_lunches(self):

        ws = self.wb["Data"]
        max_row = ws.max_row

        x = 2
        while x <= max_row:
            dow = ws["A" + str(x)].value
            schedule_num = ws["B" + str(x)].value
            schedule_num = str(dow) + "-" + str(schedule_num)
            duration = ws["E" + str(x)].value
            lunch_minutes = ws["F" + str(x)].value
            lunch_start = ws["G" + str(x)].value
            lunch_end = ws["H" + str(x)].value
            if schedule_num not in [x[0] for x in self.schedule_nums_and_lunches]:
                self.schedule_nums_and_lunches.append([schedule_num, duration, lunch_minutes, lunch_start, lunch_end])
            x += 1

    def read_in_results(self):

        ws = self.wb["Results"]

        max_row = ws.max_row
        x = 2

        while x <= max_row:
            employee = ws["B" + str(x)].value
            dow = int(ws["C" + str(x)].value)
            schedule_num = ws["D" + str(x)].value
            start_time = int(ws["E" + str(x)].value)
            end_time = int(ws["F" + str(x)].value)
            part_time = ws["J" + str(x)].value
            lunch_row = next(x for x in self.schedule_nums_and_lunches if x[0] == str(dow) + "-" + schedule_num)
            duration = int(lunch_row[1])
            lunch_minutes = int(lunch_row[2])
            lunch_start = None
            lunch_end = None
            if lunch_minutes > 0:
                lunch_start = int(lunch_row[3])
                lunch_end = int(lunch_row[4])
            row = [employee, dow, schedule_num, start_time, end_time, duration, lunch_minutes, lunch_start, lunch_end,
                   part_time]
            self.results_list.append(row)
            x += 1

    def categorize_employees(self):

        employee_nums = list(set([x[0] for x in self.results_list]))
        employee_nums.sort()
        pre_sort = []

        for employee in employee_nums:
            schedules = [x for x in self.results_list if x[0] == employee]
            total_minutes = 0
            lunch_minutes = 0
            employee_schedules = []
            for schedule in schedules:
                employee_schedules.append(schedule)
                total_minutes += int(schedule[5])
                lunch_minutes += int(schedule[6])
            pre_sort.append([employee, total_minutes, lunch_minutes, employee_schedules])

        self.employees = sorted(pre_sort, key=operator.itemgetter(1))
        self.employees.reverse()

        for employee in self.employees:
            part_time = next(x[9] for x in self.results_list if x[0] == employee[0])
            if part_time in ("0", None, 0):
                employee.append("FT")
            else:
                employee.append("PTF")

    def calculate_minutes(self):
        # this function takes the employee list and calculates things...

        for employee in self.employees:
            total_worked_weekly_minutes = employee[1] - employee[2]
            total_over_minutes = max(0, total_worked_weekly_minutes - 2400)
            total_penalty_minutes = max(0, total_worked_weekly_minutes - 3360)
            days_worked = len(employee[3])
            max_days = 5
            if employee[4] == "FT":
                max_days = 5
            elif employee[4] == "PTF":
                max_days = 6
            over_days = max(0, days_worked - max_days)
            daily_minutes = []
            weekly_over_duration_minutes = 0
            weekly_night_minutes = 0
            weekly_sunday_minutes = 0
            weekly_penalty_minutes = 0

            for schedule in employee[3]:
                total_minutes = schedule[5]
                lunch_minutes = schedule[6]
                work_minutes = total_minutes - lunch_minutes
                over_duration_minutes = max(0, work_minutes - 480)
                penalty_minutes = max(0, work_minutes - 600)
                night_diff_minutes = night_differential(self.nd_am, self.nd_pm, schedule[3], schedule[5])
                night_diff_lunch = night_differential(self.nd_am, self.nd_pm, schedule[7], schedule[6])
                night_diff_minutes -= night_diff_lunch

                sunday_minutes = sun_minutes(schedule[3], schedule[5])
                sunday_lunch = sun_minutes(schedule[7], schedule[6])
                sunday_minutes -= sunday_lunch

                daily_minutes.append(work_minutes)
                weekly_night_minutes += night_diff_minutes
                weekly_sunday_minutes += sunday_minutes
                weekly_penalty_minutes += penalty_minutes
                weekly_over_duration_minutes += over_duration_minutes

            over_day_minutes = 0
            daily_minutes.sort(key=int)
            for day in range(0, over_days):
                over_day_minutes += daily_minutes.pop(0)

            weekly_overtime_minutes = max(weekly_over_duration_minutes + over_day_minutes, total_over_minutes, 0)
            weekly_penalty_minutes = max(weekly_penalty_minutes, total_penalty_minutes)
            regular_minutes = total_worked_weekly_minutes - weekly_overtime_minutes
            weekly_overtime_minutes -= weekly_penalty_minutes
            minutes_row = [employee[0], employee[4], employee[1], employee[2], total_worked_weekly_minutes,
                           regular_minutes, weekly_night_minutes, weekly_sunday_minutes, weekly_overtime_minutes,
                           weekly_penalty_minutes]
            self.minutes_list.append(minutes_row)

    def calculate_total_hours(self):
        # [total, lunch, worked, regular, night, sunday, ot, penalty]
        ft_row = [0, 0, 0, 0, 0, 0, 0, 0]
        ptf_row = [0, 0, 0, 0, 0, 0, 0, 0]

        ft_adds = [x for x in self.minutes_list if x[1] == "FT"]

        for row in ft_adds:
            for x in range(0, 8):
                ft_row[x] += row[x+2]
        ptf_adds = [x for x in self.minutes_list if x[1] == "PTF"]
        for row in ptf_adds:
            for x in range(0, 8):
                ptf_row[x] += row[x+2]

        self.total_hours.append([x*52/60 for x in ft_row])
        self.total_hours.append([x*52/60 for x in ptf_row])

    def print_schedules(self):

        ws = self.wb["Schedules"]

        ft_count = 0
        ft_hours = 0
        ft_average = 0
        ptf_count = 0
        ptf_hours = 0
        ptf_average = 0

        start_row = 7
        for x, employee in enumerate(self.employees):
            row = str(x + start_row)
            if employee[-1] == "FT":
                ft_count += 1
                ft_hours += (employee[1] - employee[2])/60
                # ft_average = ft_hours/ft_count
            elif employee[-1] == "PTF":
                ptf_count += 1
                ptf_hours += (employee[1] - employee[2])/60
                # ptf_average = ptf_hours/ptf_count
            ws["B" + row].value = employee[0]
            schedules = employee[3]
            for schedule in schedules:
                dow = schedule[1]
                y = 1 + (dow * 2)
                ws.cell(row=int(row), column=y).value = cminute_to_tod(schedule[3])
                ws.cell(row=int(row), column=(y + 1)).value = cminute_to_tod(schedule[4])
            ws["Q" + row].value = employee[1]/60
            ws["R" + row].value = (employee[1] - employee[2])/60
            ws["S" + row].value = employee[4]

    def print_schedule_summary(self):

        ws = self.wb["Staffing"]

        start_row = 6
        for x, employee in enumerate(self.employees):
            row = str(x + start_row)
            ws["B" + row].value = employee[0]
            ws["C" + row].value = employee[4]
            dows = [x[0:3] for x in employee[3]]
            # print(dows)
            for y in range (1, 8):
                schedule_num = [x[2] for x in dows if x[1] == y]
                if schedule_num:
                    ws.cell(row=int(row), column=(3+y)).value = schedule_num[0]

    def print_employee_summary(self):

        ws = self.wb["By Employee"]
        start_row = 2
        for x, row in enumerate(self.minutes_list):
            for y, item in enumerate(row):
                ws.cell(row=start_row+x, column=y+1).value = item

    def print_total_hours(self):

        ws = self.wb["Hours for Cost Model"]

        for y, row in enumerate(self.total_hours):
            for x, item in enumerate(row):
                ws.cell(row=x+2, column=y+2).value = item


class ReadInHuasOutput:

    def __init__(self, hua_file_name, optimizer_file_name):

        print("reading in Hua")

        self.wb = load_workbook(hua_file_name)
        self.output_wb = load_workbook(optimizer_file_name)

        print("workbooks loaded")

        self.employees = []
        self.schedules = []
        print("reading in data")
        self.read_in_data()
        print("reading in schedules")
        self.read_in_schedules()
        self.wb.close()

        self.print_to_results_page()
        self.output_wb.save(optimizer_file_name)
        self.output_wb.close()
        print("done!")

    def read_in_data(self):

        ws = self.wb["Staffing Input"]

        x = 2
        max_row = ws.max_row

        while x <= max_row:
            dow = ws["A" + str(x)].value
            if dow in (None, "", " "):
                break
            schedule = ws["B" + str(x)].value
            start = int(ws["C" + str(x)].value)
            stop = int(ws["D" + str(x)].value)
            duration = int(ws["E" + str(x)].value)
            self.schedules.append([dow, schedule, start, stop, duration])
            x += 1

    def read_in_schedules(self):

        ws = self.wb["SCHEDULES"]
        self.employees = []

        x = 2
        max_row = ws.max_row

        all_rows = []
        while x <= max_row:
            if ws["D" + str(x)].value == "SOURCE":
                continue
            employee_num = ws["A" + str(x)].value
            if employee_num in (None, "", " "):
                break
            part_time = ws["B" + str(x)].value
            schedule = ws["D" + str(x)].value
            all_rows.append([employee_num, part_time, schedule])
            x += 1

        employee_nums = list(set([x[0] for x in all_rows]))

        for employee_num in employee_nums:
            rows = [x for x in all_rows if x[0] == employee_num]
            part_time = rows[0][1]
            schedules = []
            for row in rows:
                schedules.append(row[2].split("_"))
            self.employees.append([employee_num, part_time, schedules])

    def print_to_results_page(self):

        ws = self.output_wb["Results"]

        x = 2

        for employee in self.employees:
            for schedule in employee[2]:
                schedule_row = next(x for x in self.schedules if int(x[0]) == int(schedule[0]) and
                                    int(x[1]) == int(schedule[1]))
                ws["B" + str(x)].value = employee[0]
                ws["C" + str(x)].value = schedule[0]
                ws["D" + str(x)].value = schedule[1]
                ws["E" + str(x)].value = schedule_row[2]
                ws["F" + str(x)].value = schedule_row[3]
                ws["G" + str(x)].value = schedule_row[4]
                ws["H" + str(x)].value = cminute_to_tod(schedule_row[2])
                ws["I" + str(x)].value = cminute_to_tod(schedule_row[3])
                ws["J" + str(x)].value = employee[1]

                x += 1


class StaffingOutputModel(QAbstractTableModel):

    def __init__(self, data, parent=None):
        QAbstractTableModel.__init__(self, parent)
        self._data = data.values
        self._headers = data.columns

    def rowCount(self, parent=None):
        if not self._data[0][0]:
            return 0
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data[0])

    def data(self, index, role):
        if not index.isValid():
            return None
        elif role == Qt.DisplayRole:
            value = self._data[index.row()][index.column()]
            return str(value)
        else:
            return None

    def headerData(self, col, orientation, role):

        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self._headers[col]

        return QVariant()


class StaffingOutputView(QTableView):

    def __init__(self, parent=None):
        super().__init__()


class StaffingOutputToPandas:

    pass