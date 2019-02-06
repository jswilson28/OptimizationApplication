# This file contains postalizers (which compile individual plates) and other schedule manipulators and compilers.

from AddressCompilation import ExistingAddressBook, NewAddressBook
from openpyxl import load_workbook
from GeneralMethods import today
import os


def day_name_from_day_num(day_num):

    day_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    return day_list[day_num]


def get_optidata_day_nums(day_num):

    low = day_num * 1440
    high = day_num * 1440

    if day_num > 6:
        return "ERROR"

    return 1200 + low, 3360 + high


def get_a_day_string(trips):

    mon = []
    tue = []
    wed = []
    thu = []
    fri = []
    sat = []
    sun = []

    for trip in trips:
        mon.append(trip.bin_string[0])
        tue.append(trip.bin_string[1])
        wed.append(trip.bin_string[2])
        thu.append(trip.bin_string[3])
        fri.append(trip.bin_string[4])
        sat.append(trip.bin_string[5])
        sun.append(trip.bin_string[6])

    print_mon = True
    print_tue = True
    print_wed = True
    print_thu = True
    print_fri = True
    print_sat = True
    print_sun = True
    sun_schedules = False

    if tue == mon:
        print_tue = False
    if wed == tue:
        print_wed = False
    if thu == wed:
        print_thu = False
    if fri == thu:
        print_fri = False
    if sat == fri:
        print_sat = False
    if sun == sat:
        print_sun = False

    if "1" in sun:
        sun_schedules = True

    return [print_mon, print_tue, print_wed, print_thu, print_fri, print_sat, print_sun], sun_schedules


class Postalizer:

    def __init__(self, input_passer, pvs_name):

        self.ip = input_passer

        self.pvs_name = pvs_name
        self.short_name = None
        self.pdc_name = None
        self.readers = []
        self.plate_nums = []
        self.html_files = []
        self.jda_files = []
        self.sources = []
        self.schedules = []
        self.output_list = []
        self.round_trips = []
        self.num_round_trips = None

        # get address book
        folder = "Sites"
        sites = [x for x in os.listdir(folder) if x[-4:] == "xlsx"]
        site_names = [x[:-5] for x in sites]

        if self.pvs_name in site_names:
            self.address_book = ExistingAddressBook(self.pvs_name)
        else:
            self.address_book = NewAddressBook(self.pvs_name)
            self.address_book = ExistingAddressBook(self.pvs_name)

        self.set_output_list()

    def set_output_list(self):

        source_string = ""
        if len(self.sources) > 0:
            for source in self.sources:
                source_string += source + ", "
            source_string = source_string[:-2]

        plate_num_string = ""
        if len(self.plate_nums) > 0:
            for plate_num in self.plate_nums:
                plate_num_string += plate_num[:6] + ", "
            plate_num_string = plate_num_string[:-2]

        num_schedules = len(self.schedules)

        if num_schedules > 0:
            compliant = self.postal_compliant()
            compliant_schedules, non_compliant_schedules = self.count_schedules()
        else:
            compliant = False
            compliant_schedules = 0
            non_compliant_schedules = 0

        self.output_list = [self.short_name, source_string, plate_num_string, num_schedules, compliant,
                            compliant_schedules, non_compliant_schedules, self.num_round_trips]

    def add_reader(self, reader):

        if not reader or reader.pvs_name != self.pvs_name:
            return

        source = reader.source
        if source == "HCR":
            plate_num = reader.plate_number
            if plate_num in self.plate_nums:
                return
            self.pdc_name = reader.pvs_pdc
        elif source == "PVS":
            plate_num = "html"
            file_name = reader.source_file
            if file_name in self.html_files:
                return
            else:
                self.html_files.append(file_name)
            self.pdc_name = reader.pvs_pdc
        elif source == "JDA":
            plate_num = "JDA"
            file_name = reader.source_file
            if file_name in self.jda_files:
                return
            else:
                self.jda_files.append(file_name)
            self.pdc_name = reader.pdc_name
        else:
            print("UHOH")

        self.short_name = reader.short_name
        self.sources.append(source)
        self.plate_nums.append(plate_num)
        self.schedules += reader.schedules
        self.compile_round_trips()
        self.set_output_list()

    def postal_compliant(self):

        for schedule in self.schedules:
            schedule.is_postalized = schedule.postal_compliance_check(self.ip)

        for schedule in self.schedules:
            if not schedule.is_postalized:
                return False

        return True

    def count_schedules(self):

        compliant = 0
        non_compliant = 0

        for schedule in self.schedules:
            if schedule.is_postalized:
                compliant += 1
            else:
                non_compliant += 1

        return compliant, non_compliant

    def can_postalize(self):

        for schedule in self.schedules:
            schedule.cant_postalize = []
            schedule.can_postalize = schedule.postal_compliance_possible(self.ip, self.address_book)

        for schedule in self.schedules:
            if not schedule.can_postalize:
                # print(schedule.schedule_name)
                # print(schedule.cant_postalize)
                return False

        return True

    def postalize_schedules(self):

        if self.postal_compliant():
            return

        if not self.can_postalize():
            print("We'll do our best!")

        for schedule in self.schedules:
            schedule.postalize(self.ip, self.address_book)

        self.set_output_list()

    def compile_round_trips(self):

        self.round_trips = []

        for schedule in self.schedules:
            for round_trip in schedule.round_trips:
                self.round_trips.append(round_trip)

        self.num_round_trips = len(self.round_trips)

    def print_cplex_scheduler_input(self):

        eleven_ton_schedule_count, single_schedule_count = self.get_schedule_day_counts()

        round_trips_to_print = [x for x in self.round_trips if not x.holiday and x.trip_type == 1 and x.is_selected]
        print("Printing", len(round_trips_to_print), "round trips.")
        eleven_ton = [x for x in round_trips_to_print if x.vehicle_type in ("11-Ton", "11-TON")]
        single = [x for x in round_trips_to_print if x.vehicle_type == "Single"]

        spotter_round_trips_to_print = [x for x in self.round_trips if not x.holiday and x.trip_type == 3
                                        and x.is_selected]
        eleven_ton_spotters = [x for x in spotter_round_trips_to_print if x.vehicle_type in ("11-Ton", "11-TON")]
        single_spotters = [x for x in spotter_round_trips_to_print if x.vehicle_type in ("Single", "SINGLE")]

        holiday_round_trips_to_print = [x for x in self.round_trips if x.holiday and x.is_selected]
        eleven_ton_holiday = [x for x in holiday_round_trips_to_print if x.vehicle_type in ("11-Ton", "11-TON")]
        single_holiday = [x for x in holiday_round_trips_to_print if x.vehicle_type in ("Single", "SINGLE")]

        for x, trip in enumerate(eleven_ton):
            trip.optimizer_trip_num = (x + 1)

        for x, trip in enumerate(eleven_ton_spotters):
            trip.optimizer_trip_num = (x + 1)

        for x, trip in enumerate(eleven_ton_holiday):
            trip.optimizer_trip_num = (x + 1)

        for x, trip in enumerate(single):
            trip.optimizer_trip_num = (x + 1)

        for x, trip in enumerate(single_spotters):
            trip.optimizer_trip_num = (x + 1)

        for x, trip in enumerate(single_holiday):
            trip.optimizer_trip_num = (x + 1)

        eleven_ton_day_list, eleven_ton_sun = get_a_day_string(eleven_ton)
        single_day_list, single_sun = get_a_day_string(single)

        eleven_ton_file_name = "OptidataSchedules_" + self.pvs_name + "_11-ton_" + today() + ".xlsx"
        single_file_name = "OptidataSchedules_" + self.pvs_name + "_Single_" + today() + ".xlsx"

        if eleven_ton:
            wb = load_workbook("Optimization Formats/optidata.xlsx")
            source_sheet = wb['Optidata Format']
            source_sheet_two = wb['Solution Format']
            wb.save(eleven_ton_file_name)
            for day_num, print_day in enumerate(eleven_ton_day_list):
                if print_day:
                    day_name = day_name_from_day_num(day_num)
                    ws = wb.copy_worksheet(source_sheet)
                    ws.sheet_view.showGridLines = False
                    ws.title = "Optidata " + day_name
                    ws["B2"].value = day_name + " CPLEX Input Table"
                    self.print_one_optidata_day(ws, eleven_ton, day_num, wb)
                    ws = wb.copy_worksheet(source_sheet_two)
                    ws["B2"].value = day_name + " Solution"
                    ws.title = "Solution " + day_name

            wb.remove_sheet(wb['Optidata Format'])
            wb.remove_sheet(wb['Solution Format'])

            ws = wb["Data Format"]
            row = 2
            for trip in eleven_ton:
                trip.print_optidata_data(ws, row)
                row += len(trip.stops)

            ws.title = "Trips"

            ws = wb["Summary"]

            ws["B1"].value = self.pvs_name
            ws["B2"].value = self.pdc_name
            ws["B3"].value = str(eleven_ton_sun)

            for x, day in enumerate(eleven_ton_schedule_count):
                ws["B" + str(x+7)].value = day

            if eleven_ton_spotters:
                ws = wb["Spotter Format"]
                ws.title = "Spotter Trips"
                row = 2
                for trip in eleven_ton_spotters:
                    trip.print_optidata_data(ws, row)
                    row += len(trip.stops)
            else:
                wb.remove_sheet(wb["Spotter Format"])

            if eleven_ton_holiday:
                ws = wb["Holiday Format"]
                ws.title = "Holiday Trips"
                row = 2
                for trip in eleven_ton_holiday:
                    trip.print_optidata_data(ws, row)
                    row += len(trip.stops)
            else:
                wb.remove_sheet(wb["Holiday Format"])

            wb.save(eleven_ton_file_name)
            wb.close()

        if single:
            wb = load_workbook("Optimization Formats/optidata.xlsx")
            source_sheet = wb['Optidata Format']
            source_sheet_two = wb['Solution Format']
            wb.save(single_file_name)
            for day_num, print_day in enumerate(single_day_list):
                if print_day:
                    day_name = day_name_from_day_num(day_num)
                    ws = wb.copy_worksheet(source_sheet)
                    ws.sheet_view.showGridLines = False
                    ws.title = "Optidata " + day_name
                    ws["B2"].value = day_name + " CPLEX Input Table"
                    self.print_one_optidata_day(ws, single, day_num, wb)
                    ws = wb.copy_worksheet(source_sheet_two)
                    ws["B2"].value = day_name + " Solution"
                    ws.title = "Solution " + day_name

            wb.remove_sheet(wb['Optidata Format'])
            wb.remove_sheet(wb['Solution Format'])

            ws = wb["Data Format"]
            row = 2
            for trip in single:
                trip.print_optidata_data(ws, row)
                row += len(trip.stops)

            ws.title = "Trips"

            ws = wb["Summary"]
            ws["B1"].value = self.pvs_name
            ws["B2"].value = self.pdc_name
            ws["B3"].value = str(single_sun)

            for x, day in enumerate(single_schedule_count):
                ws["B" + str(x+7)].value = day

            if single_spotters:
                ws = wb["Spotter Format"]
                ws.title = "Spotter Trips"
                row = 2
                for trip in single_spotters:
                    trip.print_optidata_data(ws, row)
                    row += len(trip.stops)
            else:
                wb.remove_sheet(wb["Spotter Format"])

            if single_holiday:
                ws = wb["Holiday Format"]
                ws.title = "Holiday Trips"
                row = 2
                for trip in single_holiday:
                    trip.print_optidata_data(ws, row)
                    row += len(trip.stops)
            else:
                wb.remove_sheet(wb["Holiday Format"])

            wb.save(single_file_name)
            wb.close()

    def print_schedules(self):

        file_name = "Output Formats/ShortScheduleFormat.xlsx"
        wb = load_workbook(file_name)
        new_file_name = "Schedules " + str(self.short_name) + ".xlsx"
        wb.save(filename=new_file_name)

        summary_ws = wb["Schedule Summaries"]
        as_read_ws = wb["As Read Schedules"]
        postalized_ws = wb["Postalized Schedules"]

        orig_row = 2
        post_row = 2
        for x, schedule in enumerate(self.schedules):
            schedule.short_print_summary(summary_ws, x + 2)

            schedule.short_print_original(as_read_ws, orig_row)
            orig_row += len(schedule.original_stops)

            schedule.short_print_postalized(postalized_ws, post_row)
            post_row += len(schedule.postalized_stops)

        wb.save(new_file_name)

    def get_schedule_day_counts(self):

        eleven_ton_schedule_count = [0, 0, 0, 0, 0, 0, 0]
        single_schedule_count = [0, 0, 0, 0, 0, 0, 0]

        eleven_ton = [x for x in self.schedules if x.vehicle_category in ("11-TON", "11-Ton", "11-ton")]
        single = [x for x in self.schedules if x.vehicle_category in ("SINGLE", "Single", "single")]

        for schedule in eleven_ton:
            for x in range(0, 7):
                if schedule.schedule_type == 1 and schedule.bin_string[x] in (1, "1"):
                    eleven_ton_schedule_count[x] += 1

        for schedule in single:
            for x in range(0, 7):
                if schedule.schedule_type == 1 and schedule.bin_string[x] in (1, "1"):
                    single_schedule_count[x] += 1

        return eleven_ton_schedule_count, single_schedule_count

    @staticmethod
    def print_one_optidata_day(ws, round_trips, day_num, wb):

        low, high = get_optidata_day_nums(day_num)

        rows = []

        low_list = [0, 0, 0, low, low, 0]
        high_list = [0, 999, 0, high, high, 0]

        rows.append(low_list)
        rows.append(high_list)

        trips_to_print = [x for x in round_trips if x.bin_string[day_num] in (1, "1")]

        for trip in trips_to_print:
            rows.append(trip.optidata_row(day_num))

        for x, row in enumerate(rows):
            for y, item in enumerate(row):
                ws.cell(row=x+7, column=y+2).value = item

        max_row = 6 + len(rows)
        ws.sheet_view.showGridLines = False
        wb.create_named_range("TRIPS" + str(day_num+1), ws, "$B$7:$G$" + str(max_row))


class InSourceCompiler:

    def __init__(self, hcr_contract, vehicle_contract, hcr_reader, input_passer, cost_model):
        # this thing needs to compile whether plates are eligible for import, postalizable,
        # and then the cost.
        self.hcr_contract = hcr_contract
        self.cost_model = cost_model

        self.total_postalized_mileage = 0
        self.total_postalized_duration = 0
        self.total_postalized_night_diff = 0
        self.total_calculated_mileage = 0
        self.total_calculated_duration = 0
        self.total_calculated_night_diff = 0

        if hcr_contract:
            self.hcr_id = hcr_contract.hcr_id
            area = hcr_contract.area
        else:
            self.hcr_id = hcr_reader.plate_number
            area = "u/k"

        self.vehicle_contract = vehicle_contract
        self.hcr_reader = hcr_reader
        self.input_passer = input_passer

        if not self.hcr_reader:
            self.pvs_name = "None"
            self.short_name = "None"
        else:
            self.pvs_name = hcr_reader.pvs_name
            self.short_name = hcr_reader.short_name

        folder = "Sites"
        sites = [x for x in os.listdir(folder) if x[-4:] == "xlsx"]
        site_names = [x[:-5] for x in sites]

        if self.pvs_name in site_names:
            self.address_book = ExistingAddressBook(self.pvs_name)
        else:
            self.address_book = NewAddressBook(self.pvs_name)
            self.address_book = ExistingAddressBook(self.pvs_name)

        self.costs_enough = True

        self.total_trips = 0
        self.insourceble_trips = 0
        self.postalizable_trips = 0
        self.network_trips = 0
        self.one_state_trips = 0

        self.current_cost = self.get_current_cost()
        self.is_eligible = self.find_eligibility()
        self.is_postalizable = self.can_postalize()
        self.postalize_schedules()
        self.compile_network_trips()

        if self.hcr_contract:
            self.hcr_contract.total_calculated_duration = self.total_calculated_duration
            self.hcr_contract.total_postalized_duration = self.total_postalized_duration
            self.hcr_contract.total_calculated_mileage = self.total_calculated_mileage
            self.hcr_contract.total_postalized_mileage = self.total_postalized_mileage
            self.hcr_contract.add_night_diff(self.total_postalized_night_diff, self.total_calculated_night_diff)

        self.current_cost, self.dep_cost, self.full_cost, self.lease_cost, self.dep_cost_postalized, \
            self.full_cost_postalized, self.lease_cost_postalized = self.get_costs()

        self.output_list = [self.short_name, self.hcr_id, self.is_eligible, self.is_postalizable,
                            self.total_trips, self.insourceble_trips, self.postalizable_trips, self.network_trips,
                            self.one_state_trips, self.current_cost, self.full_cost, self.dep_cost, self.lease_cost]

        self.output_list_postalized = [self.short_name, self.hcr_id, self.is_eligible, self.is_postalizable,
                                       self.total_trips, self.insourceble_trips, self.postalizable_trips,
                                       self.network_trips, self.one_state_trips, self.current_cost,
                                       self.full_cost_postalized, self.dep_cost_postalized, self.lease_cost_postalized]

    def find_eligibility(self):

        if not self.hcr_reader:
            return False

        for schedule in self.hcr_reader.schedules:
            schedule.cant_eligible = []
            schedule.is_eligible = schedule.insource_eligible_check(self.input_passer)

        if not self.costs_enough:
            for schedule in self.hcr_reader.schedules:
                schedule.cant_eligible.append("Minimum cost")
                schedule.is_eligible = False

        return_var = True

        for schedule in self.hcr_reader.schedules:
            if schedule.is_eligible:
                self.insourceble_trips += 1
            else:
                return_var = False

        return return_var

    def can_postalize(self):

        if not self.hcr_reader:
            return False

        for schedule in self.hcr_reader.schedules:
            schedule.cant_postalize = []
            schedule.can_postalize = schedule.postal_compliance_possible(self.input_passer, self.address_book)

        return_var = True

        for schedule in self.hcr_reader.schedules:
            if schedule.can_postalize:
                self.postalizable_trips += 1
            else:
                return_var = False

        return return_var

    def compile_network_trips(self):

        if not self.hcr_reader:
            return False

        for schedule in self.hcr_reader.schedules:
            self.total_trips += 1
            if schedule.network_schedule:
                self.network_trips += 1
            if not schedule.cross_state_lines:
                self.one_state_trips += 1

    def get_current_cost(self):

        if not self.hcr_contract or not self.vehicle_contract:
            return -1

        cc = self.hcr_contract.total_annual_rate

        if cc <= 100000:
            self.costs_enough = False

        return cc

    def get_costs(self):

        if not self.hcr_contract or not self.vehicle_contract:
            return -1, -1, -1, -1, -1, -1, -1

        self.cost_model.process_contract(self.hcr_id, self.hcr_contract, self.vehicle_contract)
        cc, dc, fc, lc, dcp, fcp, lcp = self.cost_model.get_table_outputs()

        if cc <= 100000:
            self.costs_enough = False

        return cc, dc, fc, lc, dcp, fcp, lcp

    def postalize_schedules(self):

        if not self.hcr_reader:
            return

        for schedule in self.hcr_reader.schedules:
            schedule.postalize(self.input_passer, self.address_book)

            self.total_calculated_duration += schedule.annual_calculated_duration
            self.total_postalized_duration += schedule.annual_postalized_duration
            self.total_calculated_night_diff += schedule.calculated_night_hours
            self.total_postalized_night_diff += schedule.postalized_night_hours
            self.total_calculated_mileage += schedule.annual_calculated_mileage
            self.total_postalized_mileage += schedule.annual_postalized_mileage
