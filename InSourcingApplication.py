# This file contains the GUI that will be given to the client. To read plates, and determine which can be insourced,
# how much they're estimated to cost, etc...
import sys
import os
from PyQt5.QtWidgets import (QApplication, QWidget, QHBoxLayout, QVBoxLayout, QLabel, QPushButton, QLineEdit,
                             QFileDialog, QMainWindow, QTabWidget, QDialog, QMessageBox, QRadioButton, QComboBox,
                             QProgressBar, QGridLayout, QCheckBox)
from PyQt5.QtCore import QRegExp
from LookupReader import LookupBook
from GeneralMethods import InputPasser, money, QuickSpin
from datetime import datetime
from HCRPlateReaderV2 import HCRReader, CustomHCRReaderPopUp
from HCRContractReading import HCRContractReader, HCRVehicleContract
from ScheduleCompilation import InSourceCompiler
from InSorcerorTable import (CompilersToPandas, CompilerView, PopUpToPandas, PopUpView,
                             SchedulePopUpToPandas, SchedulePopUpView)
from CostEstimationModel import PythonCostModel
from openpyxl import load_workbook
from CostModelInputsTable import CostInputsView, CostInputsToLists


class App(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("InSorceror Application")
        self.setGeometry(100, 100, 1000, 500)

        self.lb = LookupBook()

        self.tabs = QTabWidget()

        self.tab1 = InSorceror(self.lb)
        self.tabs.addTab(self.tab1, "Model Main Page")

        self.tab2 = CostModelInputsGUI(self.lb)
        self.tabs.addTab(self.tab2, "Cost Model Inputs")

        self.setCentralWidget(self.tabs)
        self.show()


class InSorceror(QWidget):

    def __init__(self, lb):
        super().__init__()

        self.lb = lb
        self.ip = InputPasser(source="InSorceror")
        self.hcr_check_list = [1, 8, 24]
        self.strict_merge = True

        self.plate_nums = []
        self.readers = []
        self.compilers = []
        self.set_check_times_button = None
        self.set_new_paybook_button = None
        self.set_new_cost_model_button = None
        self.populate_table_button = None
        self.populate_by_selection_button = None
        self.custom_populate_button = None
        self.clear_button = None
        self.print_summaries_button = None
        self.print_schedules_button = None
        self.show_estimates = None
        self.show_postalized = None
        self.case = 1

        self.search_bar = None
        self.plates = None
        self.main_table = None

        self.cost_model = None
        self.contracts = None
        self.contracts_as_of_date = None
        self.update_label = None

        self.initUI()

    def initUI(self):

        outer_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        radio_hbox = QHBoxLayout()
        top_right_hbox = QHBoxLayout()
        right_layout = QVBoxLayout()

        self.main_table = CompilerView()
        self.main_table.doubleClicked.connect(self.pop_up_plate)

        self.set_new_paybook_button = QPushButton("Select New Paybook", self)
        self.set_new_cost_model_button = QPushButton("Select New Cost Model Inputs", self)
        self.populate_table_button = QPushButton("Populate by Directory", self)
        self.populate_by_selection_button = QPushButton("Populate by Selection", self)
        self.custom_populate_button = QPushButton("Custom Populate by Selection", self)
        self.print_summaries_button = QPushButton("Print Summaries", self)
        self.print_schedules_button = QPushButton("Print Schedules", self)
        self.clear_button = QPushButton("Clear Table", self)
        self.set_check_times_button = QPushButton("HCR Reader Settings", self)

        self.show_estimates = QRadioButton("Estimates")
        self.show_estimates.setChecked(True)
        self.show_estimates.case = 1
        self.show_postalized = QRadioButton("Postalized")
        self.show_postalized.case = 2
        self.show_estimates.toggled.connect(self.change_costs)
        self.show_postalized.toggled.connect(self.change_costs)

        radio_hbox.addWidget(self.show_estimates)
        rbe_tt = "Show in-sourced costs with mileage and hours estimates"
        self.show_estimates.setToolTip(rbe_tt)
        radio_hbox.addWidget(self.show_postalized)
        rbp_tt = "Show in-sourced costs with mileage and hours based on postalization"
        self.show_postalized.setToolTip(rbp_tt)

        tt1 = "Select new paybook (must be *.xlsx)"
        tt3 = "Select new cost model inputs file (must be *.xlsx)"
        tt4 = "Select a folder and populate the table with every HCR plate in that folder"
        tt5 = "Select one or more individual HCR plates to populate the table"
        tt6 = "Print cost summaries for current selection, you will be prompted to select which scenarios to print"
        tt7 = "Clear the table of all plates"
        tt8 = "Print the schedules of the selected plates to an excel file"
        tt9 = "Set the layover and combined times that the contract reader will use to merge parts of a 'U'"

        self.set_new_paybook_button.setToolTip(tt1)
        self.set_new_cost_model_button.setToolTip(tt3)
        self.populate_table_button.setToolTip(tt4)
        self.populate_by_selection_button.setToolTip(tt5)
        self.print_summaries_button.setToolTip(tt6)
        self.print_schedules_button.setToolTip(tt8)
        self.clear_button.setToolTip(tt7)
        self.set_check_times_button.setToolTip(tt9)

        self.set_new_paybook_button.clicked[bool].connect(self.set_new_paybook)
        self.set_new_cost_model_button.clicked[bool].connect(self.set_new_cost_model)
        self.populate_table_button.clicked[bool].connect(self.populate_table_by_directory)
        self.populate_by_selection_button.clicked[bool].connect(self.populate_table_by_selection)
        self.custom_populate_button.clicked[bool].connect(self.custom_populate)
        self.print_summaries_button.clicked[bool].connect(self.print_summaries)
        self.print_schedules_button.clicked[bool].connect(self.print_schedules)
        self.clear_button.clicked[bool].connect(self.deep_clear)
        self.set_check_times_button.clicked[bool].connect(self.set_check_times)

        search_label = QLabel("Search Plate Numbers")
        self.search_bar = QLineEdit()
        self.search_bar.textChanged.connect(self.filter_pattern)
        self.search_bar.setFixedWidth(100)
        self.search_bar.setToolTip("Filter on HCR contract number")

        left_layout.addWidget(QLabel("Select HCR Plates"))
        left_layout.addWidget(self.populate_table_button)
        left_layout.addWidget(self.populate_by_selection_button)
        left_layout.addWidget(self.custom_populate_button)
        left_layout.addWidget(self.clear_button)

        left_layout.addStretch(1)

        left_layout.addWidget(search_label)
        left_layout.addWidget(self.search_bar)

        # left_layout.addStretch(1)
        # left_layout.addLayout(radio_hbox)

        top_right_hbox.addStretch(4)
        top_right_hbox.addWidget(QLabel("Cost Estimation Method:"))
        top_right_hbox.addLayout(radio_hbox)
        top_right_hbox.addStretch(1)

        left_layout.addStretch(1)

        left_layout.addWidget(QLabel("Generate Output (Selected Rows)"))
        left_layout.addWidget(self.print_summaries_button)
        left_layout.addWidget(self.print_schedules_button)

        left_layout.addStretch(2)

        left_layout.addWidget(QLabel("Select Model Inputs"))
        self.update_label = QLabel()
        self.update_label_now()
        left_layout.addWidget(self.update_label)
        left_layout.addWidget(self.set_new_paybook_button)
        left_layout.addWidget(self.set_new_cost_model_button)
        left_layout.addWidget(self.set_check_times_button)

        outer_layout.addLayout(left_layout)

        right_layout.addLayout(top_right_hbox)
        right_layout.addWidget(self.main_table)

        outer_layout.addLayout(right_layout)

        self.setLayout(outer_layout)

        try:
            self.set_cost_model()
        except:
            QMessageBox.question(self, "Cost Model Load", "Please select a cost model", QMessageBox.Ok)
            self.set_new_cost_model()

        try:
            self.read_existing_contract_csv()
        except:
            QMessageBox.question(self, "Paybook Load", "Please select a paybook (xlsx)", QMessageBox.Ok)
            self.set_new_paybook()

        self.update_table()

    def update_label_now(self):
        self.update_label.setText("(Current paybook: " + str(self.contracts_as_of_date) + ")")

    def filter_pattern(self):

        reg_exp = QRegExp(self.search_bar.text())
        if self.main_table.sorter:
            self.main_table.sorter.setFilterRegExp(reg_exp)
        else:
            return

    def populate_table_by_directory(self):

        if not self.contracts:
            QMessageBox.question(self, "Paybook Load", "Please select a paybook (xlsx)", QMessageBox.Ok)
            self.set_new_paybook()
            return

        # if not self.vehicle_contracts:
        #     print("Vehicles not loaded")
        #     return

        if not self.cost_model:
            QMessageBox.question(self, "Cost Model Load", "Please select a cost model", QMessageBox.Ok)
            self.set_new_cost_model()
            return

        self.clear()

        file_name = QFileDialog.getExistingDirectory(self, "Select Directory Containing Plates")

        if file_name:
            items = os.listdir(file_name)
            self.plates = [x for x in items if x[-3:] in ('PDF', 'Pdf', 'pdf')]
            self.plates = [file_name + "/" + x for x in self.plates]
            self.read_plates()
        else:
            print("No files selected")
            return

        self.update_table()

    def populate_table_by_selection(self):

        if not self.contracts:
            QMessageBox.question(self, "Paybook Load", "Please select a paybook (.xlsx)", QMessageBox.Ok)
            self.set_new_paybook()
            return

        # if not self.vehicle_contracts:
        #     print("Vehicles not loaded")
        #     return

        if not self.cost_model:
            QMessageBox.question(self, "Cost Model Load", "Please select a cost model", QMessageBox.Ok)
            self.set_new_cost_model()
            return

        self.clear()

        # options = QFileDialog.Options()
        # options |= QFileDialog.DontUseNativeDialog
        files, _ = QFileDialog.getOpenFileNames(self, "Select HCR Plate(s)",
                                                "", "PDF Files (*.pdf);;PDF Files (*.pdf)")
                                                # options=options)
        if files:
            self.plates = files
            self.read_plates()
        else:
            print("No files selected")
            return

        self.update_table()

    def custom_populate(self):

        if not self.contracts:
            QMessageBox.question(self, "Paybook Load", "Please select a paybook (xlsx)", QMessageBox.Ok)
            self.set_new_paybook()
            return

        if not self.cost_model:
            QMessageBox.question(self, "Cost Model Load", "Please select a cost model", QMessageBox.Ok)
            self.set_new_cost_model()
            return

        self.clear()

        files, _ = QFileDialog.getOpenFileNames(self, "Select HCR Plate(s)",
                                                "", "PDF Files (*.pdf);;PDF Files (*.pdf)")
        if files:
            self.plates = files
            # for pdf in self.plates:
            #     test_reader = HCRReader(pdf, self.lb, "Test", self.hcr_check_list)
            #     possible_names, possible_addresses = test_reader.get_potential_pdcs_and_addresses()
            self.read_plates()
        else:
            print("No files selected")
            return

        self.update_table()

    def change_costs(self):

        radiobutton = self.sender()

        if radiobutton.isChecked():
            self.case = radiobutton.case
            self.clear()
            self.update_table()

    def update_table(self):

        case = self.case

        for x, plate_num in enumerate(self.plate_nums):
            reader = self.readers[x]
            contract = self.contracts.get_contract(plate_num)
            # vehicle_contract = self.vehicle_contracts.get_contract(plate_num)
            if reader:
                vehicle_contract = HCRVehicleContract(plate_num, reader.cargo_van_count, reader.tractor_sa_count,
                                                      reader.tractor_ta_count, reader.trailer_count)
            else:
                vehicle_contract = HCRVehicleContract(plate_num, 0, 0, 0, 0)
            self.compilers.append(InSourceCompiler(contract, vehicle_contract, reader, self.ip, self.cost_model))

        # print("all compilers made")
        compiled_df = CompilersToPandas(self.compilers, case).df
        self.main_table.set_model(compiled_df)

    def clear(self):

        del self.plates
        del self.compilers

        self.plates = []
        self.compilers = []
        # compiled_df = CompilersToPandas(self.compilers).df
        # self.main_table.set_model(compiled_df)
        # self.main_table.update()

    def deep_clear(self):

        del self.plates
        del self.plate_nums
        del self.readers
        del self.compilers

        self.plates = []
        self.plate_nums = []
        self.readers = []
        self.compilers = []
        compiled_df = CompilersToPandas(self.compilers, self.case).df
        self.main_table.set_model(compiled_df)
        self.main_table.update()

    def read_plates(self):

        progress_bar = ProgressBarWindow("Progress Reading Plates", len(self.plates))
        QApplication.processEvents()

        for x, pdf in enumerate(self.plates):

            progress_bar.update(x)
            plate_num = pdf.split('/')[-1][:5]

            if plate_num not in self.plate_nums:
                self.plate_nums.append(plate_num)
                reader = HCRReader(pdf, self.lb, "InSorceror", [num*60 for num in self.hcr_check_list],
                                   strict_merge=self.strict_merge)
                if reader.is_readable:
                    self.readers.append(reader)
                else:
                    self.readers.append(None)
            else:
                print("Duplicate plate " + plate_num + " not added.")

    def set_new_paybook(self):

        file_name, _ = QFileDialog.getOpenFileName(self, "Select HCR Contract File", '', 'Excel Files (*.xlsx)', None,
                                                   QFileDialog.DontUseNativeDialog)

        if file_name:
            try:
                self.contracts = HCRContractReader(file_name)
            except:
                QMessageBox.question(self, "Paybook Selection", "Paybook could not be read!", QMessageBox.Ok)
                return
            self.contracts_as_of_date = self.contracts.as_of_date
            print("here", str(self.contracts_as_of_date))
            self.update_label_now()
            # self.lb.update_paybook_file(file_name)

    def read_existing_contract_csv(self):

        self.contracts = HCRContractReader('hcr_contracts.csv')
        self.contracts_as_of_date = self.contracts.as_of_date
        self.update_label_now()

    def pop_up_plate(self):

        index = self.main_table.selectedIndexes()[0]
        index = self.main_table.model().mapToSource(index)

        plate_num = self.plate_nums[index.row()]
        reader = self.compilers[index.row()].hcr_reader
        short_name = self.compilers[index.row()].short_name

        PlatePopUpGUI(short_name, plate_num, reader).exec_()

    def set_cost_model(self):

        self.cost_model = PythonCostModel("Cost Model Inputs.xlsx")

    def set_new_cost_model(self):

        file_name, _ = QFileDialog.getOpenFileName(self, "Select Cost Model Inputs", '', 'Excel Files (*.xlsx)', None,
                                                   QFileDialog.DontUseNativeDialog)

        if file_name:
            wb = load_workbook(file_name)
            if "Inputs" not in wb.get_sheet_names():
                QMessageBox.question(self, "Cost Model", "Selected file does not have Inputs tab!", QMessageBox.Ok)
                return
            wb.save("Cost Model Inputs.xlsx")
            for ws in wb.worksheets:
                if ws.title != "Inputs":
                    wb.remove_sheet(ws)
            wb.save("Cost Model Inputs.xlsx")
            self.set_cost_model()
            # self.lb.update_cost_model_file(file_name)

    def print_summaries(self):

        if len(self.compilers) == 0:
            return

        compilers_to_print = []

        indices = self.main_table.selectedIndexes()
        unique_rows = []

        for index in indices:
            row = self.main_table.model().mapToSource(index).row()
            if row not in unique_rows:
                unique_rows.append(row)

        for row in unique_rows:
            compilers_to_print.append(self.compilers[row])

        if len(compilers_to_print) == 0:
            QMessageBox.question(self, "Print Selection", "No rows selected!", QMessageBox.Ok)
            return

        default_name = compilers_to_print[0].hcr_id
        default_name = "Output " + str(datetime.now().date()) + " " + str(default_name)

        wodep, wdep, lease, name, cancel = PrintOptionsPopUp(default_name, self.case).exec_()

        if cancel:
            return

        if len(name) < 1:
            name = None

        case_list = [wodep, wdep, lease]
        self.print_table(case_list, name, compilers_to_print)

    def print_table(self, case_list, input_file_name, compilers_to_print):

        if input_file_name:
            new_file_name = input_file_name
        else:
            QMessageBox.question(self, "No Name", "Must input a name!", QMessageBox.Ok)
            return

        if new_file_name[:-5] != ".xlsx":
            new_file_name += ".xlsx"

        file_name = "Output Formats/CostModelOutputsCompiled.xlsx"

        try:
            wb = load_workbook(file_name)
            wb.save(filename=new_file_name)
        except:
            QMessageBox.question(self, "Close Workbook", "Close workbook, and try again.", QMessageBox.Ok)
            return

        cases = []
        source_sheets = []
        title_strings = []

        if case_list[0]:
            cases.append(1)
            source_sheets.append(wb["Summary with No Depreciation"])
            title_strings.append(" wo Dep")
        if case_list[1]:
            cases.append(2)
            source_sheets.append(wb["Summary with Depreciation"])
            title_strings.append(" w Dep")
        if case_list[2]:
            cases.append(3)
            source_sheets.append(wb["Summary with Leased Fleet"])
            title_strings.append(" w Lease")

        summary_list = []
        for x, case in enumerate(cases):
            for compiler in compilers_to_print:
                if x == 0:
                    if self.case == 1:
                        summary_list.append(compiler.output_list)
                    if self.case == 2:
                        summary_list.append(compiler.output_list_postalized)
                contract = compiler.hcr_contract
                if contract:
                    ws = wb.copy_worksheet(source_sheets[x])
                    ws.title = compiler.hcr_id + title_strings[x]
                    if self.case == 1:
                        contract.print_one_summary_estimated(ws, case)
                    if self.case == 2:
                        contract.print_one_summary_postalized(ws, case)

        ws = wb['Summary Sheet']

        for x, row in enumerate(summary_list):
            for y, item in enumerate(row):
                if y in (0, 1, 2, 3, 4, 5, 6, 7, 8):
                    ws.cell(row=x+2, column=y+1).value = item
                else:
                    ws.cell(row=x+2, column=y+1).value = money(item)

        wb.save(filename=new_file_name)
        if not case_list[0]:
            wb.remove_sheet(wb["CheatSheet with No Depreciation"])
        if not case_list[1]:
            wb.remove_sheet(wb["CheatSheet with Depreciation"])
        if not case_list[2]:
            wb.remove_sheet(wb["CheatSheet with Leased Fleet"])

        wb.remove_sheet(wb["Summary with No Depreciation"])
        wb.remove_sheet(wb["Summary with Depreciation"])
        wb.remove_sheet(wb["Summary with Leased Fleet"])
        wb.save(filename=new_file_name)

    def print_schedules(self):

        if len(self.compilers) == 0:
            return

        compilers_to_print = []

        indices = self.main_table.selectedIndexes()
        unique_rows = []

        for index in indices:
            row = self.main_table.model().mapToSource(index).row()
            if row not in unique_rows:
                unique_rows.append(row)

        for row in unique_rows:
            compilers_to_print.append(self.compilers[row])

        if len(compilers_to_print) == 0:
            QMessageBox.question(self, "Print Selection", "No rows selected!", QMessageBox.Ok)
            return

        file_name = "Output Formats/ShortScheduleFormat.xlsx"

        for compiler in compilers_to_print:

            wb = load_workbook(file_name)
            new_file_name = "Schedules " + str(compiler.hcr_id) + ".xlsx"
            wb.save(filename=new_file_name)

            summary_ws = wb["Schedule Summaries"]
            as_read_ws = wb["As Read Schedules"]
            postalized_ws = wb["Postalized Schedules"]

            orig_row = 2
            post_row = 2
            if compiler.hcr_reader:
                for x, schedule in enumerate(compiler.hcr_reader.schedules):
                    schedule.short_print_summary(summary_ws, x+2)

                    schedule.short_print_original(as_read_ws, orig_row)
                    orig_row += len(schedule.original_stops)

                    schedule.short_print_postalized(postalized_ws, post_row)
                    post_row += len(schedule.postalized_stops)
            else:
                QMessageBox.question(self, "Schedule Printer", "Plate not read correctly", QMessageBox.Ok)

            wb.save(new_file_name)

    def set_check_times(self):

        self.strict_merge, self.hcr_check_list = HCRReaderSettingsPopUp(self.strict_merge, self.hcr_check_list).exec_()


class CostModelInputsGUI(QWidget):

    def __init__(self, lb):
        super().__init__()

        file_name = "Cost Model Inputs.xlsx"
        self.lb = lb

        try:
            self.wb = load_workbook(file_name)
        except:
            return

        self.ws = self.wb['Inputs']
        self.save_button = None
        self.undo_button = None
        self.table_selector = None
        # self.central_table = CostModelInputsTable(self.ws)
        # self.central_table.set_model(self.ws)
        self.central_table = CostInputsView()
        self.initUI()

    def initUI(self):

        left_vbox = QVBoxLayout()
        right_vbox = QVBoxLayout()
        hbox = QHBoxLayout()
        # self.save_button = QPushButton("Save Changes", self)
        # self.undo_button = QPushButton("Undo Changes", self)
        # self.save_button.clicked[bool].connect(self.save_changes)
        # self.undo_button.clicked[bool].connect(self.undo_changes)

        self.table_selector = QComboBox()
        self.table_selector.addItem("Hourly Wages")
        self.table_selector.addItem("Labor Splits")
        self.table_selector.addItem("Other Inputs")
        self.table_selector.addItem("Vehicle Acquisition Costs")
        self.table_selector.addItem("Vehicle Lease Details")
        self.table_selector.addItem("Cost per Mile by Region")
        self.table_selector.addItem("Trailer Maintenance by Region")
        self.table_selector.addItem("Annual Fuel Costs")
        self.table_selector.currentIndexChanged.connect(self.update_table)

        notelabels = ["- Annual STO Cost is Fully Loaded FY17 Cost for EAS-17 Level from National Average Labor Rates "
                      "January 22, 2018.",
                      "- Leasing costs come from Ryder USPS Pricing 2016. Cargo Van costs align to costs of CDL "
                      "Heavy Duty Straight Truck, Single Axle is Single Daycab and Tandem Axle is Tandem Daycab. "
                      "Monthly Costs reflect the cost of a 1 Year ebuy.",
                      "- Cost per Mile from 5505 Annual Average Q2 FY17 - Q1 FY18. Cargo Van costs are 11-Ton. "
                      "National costs are applied to HQ contracts."]

        left_vbox.addWidget(self.table_selector)
        left_vbox.addWidget(self.central_table)
        right_vbox.addWidget(QLabel("Notes: "))
        for var in notelabels:
            label = QLabel(var)
            label.setWordWrap(True)
            right_vbox.addWidget(label)
        right_vbox.addStretch(1)
        hbox.addLayout(left_vbox, 2)
        hbox.addLayout(right_vbox, 1)
        # hbox.addStretch(2)
        self.setLayout(hbox)
        self.update_table()

    def undo_changes(self):

        self.central_table.undo_changes()

    def save_changes(self):

        # self.central_table.save_changes()
        print("hi!")

    def update_table(self):
        index = self.table_selector.currentIndex() + 1
        data_list, headers = CostInputsToLists(self.ws).get_a_table(index)
        self.central_table.set_model(data_list, headers, index)


class PlatePopUpGUI(QDialog):

    def __init__(self, short_name, plate_num, reader, parent=None):
        super(PlatePopUpGUI, self).__init__(parent)

        self.plate_num = plate_num
        self.setWindowTitle(short_name + ": " + self.plate_num)

        if not reader:
            self.no_reader()
            return

        self.schedules = reader.schedules
        self.close_button = None
        self.central_table = None
        self.initUI()

    def initUI(self):

        vbox = QVBoxLayout()

        self.setGeometry(200, 200, 850, 500)
        self.central_table = PopUpView()
        self.central_table.doubleClicked.connect(self.pop_up_schedule)
        self.central_table.set_model(PopUpToPandas(self.schedules).df)

        self.close_button = QPushButton("Close", self)
        self.close_button.clicked[bool].connect(self.accept)

        vbox.addWidget(self.central_table)
        vbox.addWidget(self.close_button)

        self.setLayout(vbox)
        self.show()

    def no_reader(self):

        vbox = QVBoxLayout()
        vbox.addWidget(QLabel("Plate not read in correctly"))
        close_button = QPushButton("Close", self)
        close_button.clicked[bool].connect(self.accept)
        vbox.addWidget(close_button)
        self.setGeometry(500, 300, 200, 100)
        self.setLayout(vbox)

    def pop_up_schedule(self):

        index = self.central_table.selectedIndexes()[0]
        index = self.central_table.model().mapToSource(index)

        schedule = self.schedules[index.row()]

        SchedulePopUpGUI(schedule).exec_()


class SchedulePopUpGUI(QDialog):

    def __init__(self, schedule, parent=None):
        super(SchedulePopUpGUI, self).__init__(parent)

        self.schedule = schedule
        self.setWindowTitle(schedule.schedule_name)
        self.setGeometry(250, 250, 750, 400)
        self.close_button = None
        self.left_table = None
        self.right_table = None
        self.initUI()

    def initUI(self):
        vbox = QVBoxLayout()
        left_vbox = QVBoxLayout()
        right_vbox = QVBoxLayout()
        hbox = QHBoxLayout()

        left_label = QLabel("Schedule as read:")
        self.left_table = SchedulePopUpView()
        self.left_table.set_model(SchedulePopUpToPandas(self.schedule, 1).df)

        if self.schedule.is_postalized:
            postalized_string = "Compliant"
        else:
            postalized_string = "Non-Compliant"

        right_label = QLabel("Postalized Schedule: " + postalized_string)
        self.right_table = SchedulePopUpView()
        self.right_table.set_model(SchedulePopUpToPandas(self.schedule, 2).df)
        # self.right_table.set_color(self.schedule.is_postalized)

        self.close_button = QPushButton("Close", self)
        self.close_button.clicked[bool].connect(self.accept)

        left_vbox.addWidget(left_label)
        left_vbox.addWidget(self.left_table)
        right_vbox.addWidget(right_label)
        right_vbox.addWidget(self.right_table)

        hbox.addLayout(left_vbox)
        hbox.addLayout(right_vbox)
        vbox.addLayout(hbox)
        vbox.addWidget(self.close_button)

        self.setLayout(vbox)
        self.show()


class ProgressBarWindow(QWidget):

    def __init__(self, title, length):
        super().__init__()
        self.setWindowTitle(title)
        vbox = QVBoxLayout()
        self.length = length
        self.pbar = QProgressBar(self)
        self.pbar.setMaximum(length)
        self.pbar.setGeometry(10, 10, 200, 50)
        vbox.addWidget(self.pbar)
        self.label = QLabel("Reading Plate: x/" + str(length))
        vbox.addWidget(self.label)
        self.setLayout(vbox)
        self.setGeometry(300, 300, 300, 50)
        self.show()

    def update(self, value):

        self.pbar.setValue(value)
        self.label.setText("Reading Plate: " + str(value) + "/" + str(self.length))


class PrintOptionsPopUp(QDialog):

    def __init__(self, default_name, case, parent=None):
        super(PrintOptionsPopUp, self).__init__(parent)

        self.setWindowTitle("Print Options")
        self.default_name = default_name
        self.case = case
        self.wodep = None
        self.wdep = None
        self.lease = None
        self.file_name = None
        self.accept_button = None
        self.cancel = False
        self.cancel_button = None

        self.initUI()

    def initUI(self):

        main_layout = QVBoxLayout()
        button_grid = QGridLayout()
        file_hbox = QHBoxLayout()
        button_hbox = QHBoxLayout()

        grid_label = QLabel("Select which summaries to print:")
        if self.case == 1:
            next_label = QLabel("(Using mileage and duration inflation estimates)")
        elif self.case == 2:
            next_label = QLabel("(Using postalized mileage and duration)")
        else:
            next_label = "Uh oh"

        labels = ["w/o Depreciation", "w/ Depreciation", "w/ Lease"]
        for x, label in enumerate(labels):
            button_grid.addWidget(QLabel(label), x, 0)

        self.wodep = QCheckBox()
        self.wdep = QCheckBox()
        self.lease = QCheckBox()

        button_grid.addWidget(self.wodep, 0, 1)
        button_grid.addWidget(self.wdep, 1, 1)
        button_grid.addWidget(self.lease, 2, 1)

        self.file_name = QLineEdit()
        self.file_name.setText(self.default_name)
        file_label = QLabel("Enter Output File Name: ")

        file_hbox.addWidget(file_label)
        file_hbox.addWidget(self.file_name)

        self.accept_button = QPushButton("Accept", self)
        self.accept_button.clicked[bool].connect(self.accept)

        self.cancel_button = QPushButton("Cancel", self)
        self.cancel_button.clicked[bool].connect(self.cancel_function)

        button_hbox.addWidget(self.cancel_button)
        button_hbox.addWidget(self.accept_button)

        main_layout.addWidget(grid_label)
        main_layout.addWidget(next_label)
        main_layout.addLayout(button_grid)
        main_layout.addLayout(file_hbox)
        main_layout.addLayout(button_hbox)

        self.setLayout(main_layout)
        self.show()

    def cancel_function(self):

        self.cancel = True
        self.accept()

    def exec_(self):
        super().exec_()

        wodep = self.wodep.isChecked()
        wdep = self.wdep.isChecked()
        lease = self.lease.isChecked()
        name = self.file_name.text()

        self.close()

        return wodep, wdep, lease, name, self.cancel


class HCRReaderSettingsPopUp(QDialog):

    def __init__(self, current_u_part_merge, current_check_times, parent=None):
        super(HCRReaderSettingsPopUp, self).__init__(parent)

        self.setWindowTitle("Set HCR Reader Settings Times")

        self.current_check_times = current_check_times
        self.current_u_part_merge = current_u_part_merge

        self.strict_merge = QCheckBox()
        self.strict_merge.setChecked(self.current_u_part_merge)

        self.min_time_spin = QuickSpin(0, 100, current_check_times[0])
        self.max_layover_spin = QuickSpin(0, 100, current_check_times[1])
        self.max_combined_spin = QuickSpin(0, 100, current_check_times[2])

        self.cancel_button = QPushButton("Cancel", self)
        self.submit_button = QPushButton("Submit", self)

        self.cancel_button.clicked[bool].connect(self.cancel)
        self.submit_button.clicked[bool].connect(self.accept)

        self.initUI()

    def initUI(self):

        strict_row = QHBoxLayout()
        grid = QGridLayout()
        button_row = QHBoxLayout()
        main_layout = QVBoxLayout()

        strict_row.addWidget(QLabel("Merge all possible U Parts?"))
        strict_row.addWidget(self.strict_merge)

        grid.addWidget(QLabel("Minimum layover hours to require check: "), 0, 0)
        grid.addWidget(self.min_time_spin, 0, 1)

        grid.addWidget(QLabel("Maximum layover hours to allow check: "), 1, 0)
        grid.addWidget(self.max_layover_spin, 1, 1)

        grid.addWidget(QLabel("Combined maximum hours duration:: "), 2, 0)
        grid.addWidget(self.max_combined_spin, 2, 1)

        button_row.addWidget(self.cancel_button)
        button_row.addWidget(self.submit_button)

        main_layout.addLayout(strict_row)
        main_layout.addLayout(grid)
        main_layout.addLayout(button_row)
        self.setLayout(main_layout)

    def cancel(self):

        self.strict_merge.setChecked(self.current_u_part_merge)

        self.min_time_spin.setValue(self.current_check_times[0])
        self.max_layover_spin.setValue(self.current_check_times[1])
        self.max_combined_spin.setValue(self.current_check_times[2])

        self.accept()

    def exec_(self):
        super().exec_()

        bool0 = self.strict_merge.isChecked()

        val1 = self.min_time_spin.value()
        val2 = self.max_layover_spin.value()
        val3 = self.max_combined_spin.value()

        return bool0, [val1, val2, val3]


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
